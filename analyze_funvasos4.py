import openpyxl
wb = openpyxl.load_workbook('DocumentRepository/funvasos/FIN160226.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Each "unit" has 4 rows: PO (programado), RP (reprogramado), RP, PR (progreso/realizado)
# Column X has the state: PO, RP, RP, PR

# Check if col C has name in merged cells spanning the rows
print('=== Merged cells involving col C (names) ===')
for mc in ws.merged_cells.ranges:
    if mc.min_col <= 3 and mc.max_col >= 3 and mc.min_row >= 19:
        print(f'  {mc}: val={ws.cell(row=mc.min_row, column=3).value}')

print('\n=== Merged cells involving col I (unit codes) ===')
for mc in ws.merged_cells.ranges:
    if mc.min_col <= 9 and mc.max_col >= 9 and mc.min_row >= 19:
        print(f'  {mc}: val={ws.cell(row=mc.min_row, column=9).value}')

print('\n=== Merged cells involving col M (mantto type) ===')
for mc in ws.merged_cells.ranges:
    if mc.min_col <= 13 and mc.max_col >= 13 and mc.min_row >= 19:
        print(f'  {mc}: val={ws.cell(row=mc.min_row, column=13).value}')

# Row 143 has I=MMT04 but C=None - check merged C around row 143
print('\n=== Searching C values around rows 130-150 ===')
for mc in ws.merged_cells.ranges:
    if mc.min_col <= 3 and mc.max_col >= 8 and mc.min_row >= 130 and mc.min_row <= 200:
        print(f'  {mc}: val={ws.cell(row=mc.min_row, column=3).value}')

# Summary: Check rows 297-312 for totals
print('\n=== ROWS 297-312 (totals area) ===')
for r in range(297, 313):
    for c in range(1, 30):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            print(f'  R{r} {cl}: {repr(v)}')

# Check the color coding of the bars
# Check which dates have actual data values for row 19 (PO)
print('\n=== Row 19 (ANG01 PO) - checking which date columns have non-date values ===')
# The Z-NZ columns = 365 days. Check if any has a number vs a date
date_count = 0
num_count = 0
for c in range(26, 396):  # Z=26 to NZ=390
    v = ws.cell(row=19, column=c).value
    if v is not None:
        import datetime
        if isinstance(v, datetime.datetime):
            date_count += 1
        elif isinstance(v, (int, float)):
            num_count += 1
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            print(f'  {cl}{19}: {v}')
        else:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            print(f'  {cl}{19}: {repr(v)} (type: {type(v).__name__})')
print(f'  date_count={date_count}, num_count={num_count}')

# Check cell fill/formatting for the gantt bars
print('\n=== Cell background colors for row 19 (few sample dates) ===')
for c in range(26, 90):  # Z to CK (Jan-Feb-Mar approx)
    cell = ws.cell(row=19, column=c)
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.rgb and str(fill.start_color.rgb) != '00000000':
        from openpyxl.utils import get_column_letter
        cl = get_column_letter(c)
        print(f'  {cl}{19}: fill={fill.start_color.rgb}, val={cell.value}')
