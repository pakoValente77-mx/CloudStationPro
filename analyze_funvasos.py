import openpyxl
wb = openpyxl.load_workbook('DocumentRepository/funvasos/FIN160226.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
print(f'Dimensions: {ws.dimensions}, Rows: {ws.max_row}, Cols: {ws.max_column}')

# Print first 15 rows, columns A-Z only (first 26 cols)
print('\n=== HEADER AREA (rows 1-15, cols A-Z) ===')
for r in range(1, 16):
    row_data = {}
    for c in range(1, 27):
        cell = ws.cell(row=r, column=c)
        if cell.value is not None:
            row_data[cell.coordinate] = cell.value
    if row_data:
        print(f'Row {r}: {row_data}')

# Print column X (variable types) for rows 1-60
print('\n=== Column X (Variable Type), rows 1-60 ===')
for r in range(1, 61):
    val = ws.cell(row=r, column=24).value  # X = col 24
    if val:
        print(f'Row {r}: X={val}')

# Print key columns for rows 1-60
print('\n=== Key columns B,I,J,M,O,P,S rows 1-60 ===')
for r in range(1, 61):
    b = ws.cell(row=r, column=2).value
    i_val = ws.cell(row=r, column=9).value
    j = ws.cell(row=r, column=10).value
    m = ws.cell(row=r, column=13).value
    o = ws.cell(row=r, column=15).value
    p = ws.cell(row=r, column=16).value
    s = ws.cell(row=r, column=19).value
    if any([b, i_val, j, m, o, p, s]):
        print(f'Row {r}: B={b}, I={i_val}, J={j}, M={m}, O={o}, P={p}, S={s}')

# Print merged cells
print('\n=== Merged Cells ===')
for mc in ws.merged_cells.ranges:
    print(mc)

# Columns OF, OG, OH, OI for rows 1-10
print('\n=== Far-right columns (OF=col395, OG, OH, OI) header row ===')
for r in range(1, 5):
    for c_name in ['OA','OB','OC','OD','OE','OF','OG','OH','OI']:
        col_idx = 0
        for i, ch in enumerate(reversed(c_name)):
            col_idx += (ord(ch) - ord('A') + 1) * (26 ** i)
        val = ws.cell(row=r, column=col_idx).value
        if val is not None:
            print(f'Row {r}, {c_name} (col {col_idx}): {val}')

# Check what columns Z through AF contain in row 1 (header)
print('\n=== Column Z-AF (cols 26-32) in header rows (dates?) ===')
for r in [1, 2, 3, 4, 5]:
    for c in range(26, 33):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(c)
            print(f'Row {r}, {col_letter}{r}: {val}')

# Check how many vasos/dams there are
print('\n=== Column A values (all rows) ===')
for r in range(1, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value
    if a is not None:
        print(f'Row {r}: A={a}')
