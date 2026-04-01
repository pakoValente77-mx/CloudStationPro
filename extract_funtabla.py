import openpyxl
import json

wb = openpyxl.load_workbook(
    '/Users/subgerenciagrijalva/Downloads/0_simulacion_Ene-Dic 2025_Predespacho_Año SM_01Ene2025.xlsx',
    data_only=True
)
ws = wb['FUNTABLA']

presas = {
    'Angostura':    {'col_elev': 1, 'col_cap': 2, 'col_area': 3, 'col_cesp': 4},
    'Chicoasen':    {'col_elev': 7, 'col_cap': 8, 'col_area': 9, 'col_cesp': 10},
    'Malpaso':      {'col_elev': 13, 'col_cap': 14, 'col_area': 15, 'col_cesp': 16},
    'Penitas':      {'col_elev': 19, 'col_cap': 20, 'col_area': 21, 'col_cesp': 22},
    'JGrijalva':    {'col_elev': 26, 'col_cap': 27, 'col_area': 28},
}

result = {}
for name, cols in presas.items():
    records = []
    for row in range(3, ws.max_row + 1):
        e = ws.cell(row=row, column=cols['col_elev']).value
        c = ws.cell(row=row, column=cols['col_cap']).value
        if e is None or c is None:
            continue
        rec = {'elevation': round(float(e), 2), 'capacity_mm3': round(float(c), 4)}
        a = ws.cell(row=row, column=cols['col_area']).value
        if a is not None:
            rec['area_km2'] = round(float(a), 4)
        if 'col_cesp' in cols:
            ce = ws.cell(row=row, column=cols['col_cesp']).value
            if ce is not None:
                rec['specific_consumption'] = round(float(ce), 4)
        records.append(rec)
    result[name] = records
    print(f"{name}: {len(records)} registros, Elev: {records[0]['elevation']} - {records[-1]['elevation']} msnm")

with open('/Users/subgerenciagrijalva/CFE/CloudStation/Datos/funtabla_curves.json', 'w') as f:
    json.dump(result, f)

print("\nJSON guardado en Datos/funtabla_curves.json")
print(f"Total registros: {sum(len(v) for v in result.values())}")
