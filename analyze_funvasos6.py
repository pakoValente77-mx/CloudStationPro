import openpyxl
wb = openpyxl.load_workbook('/Users/subgerenciagrijalva/Downloads/FIN240326.xlsx', data_only=True)

print(f'Sheets: {wb.sheetnames}')
for sn in wb.sheetnames:
    ws = wb[sn]
    # find actual max row (skip empty)
    actual_max = 0
    for r in range(1, 200):
        has_data = False
        for c in range(1, 96):
            if ws.cell(row=r, column=c).value is not None:
                has_data = True
                break
        if has_data:
            actual_max = r
    print(f'\nSheet "{sn}": actual rows with data up to {actual_max}, cols: {ws.max_column}')

ws = wb['FIN']

# Identify all section headers (where row has "Elev. vaso")
print('\n=== SECTION HEADERS ===')
for r in range(1, 200):
    for c in range(1, 96):
        v = ws.cell(row=r, column=c).value
        if v is not None and isinstance(v, str) and 'Elev' in v:
            from openpyxl.utils import get_column_letter
            print(f'  R{r}: {get_column_letter(c)}={repr(v)}')

# Get dam names and section starts
print('\n=== SECTION LABELS (rows 8-11, 40-44, 73-77, 106-110, 139-143) ===')
for scan_start, scan_end in [(7,14), (40,48), (73,81), (106,114), (139,148)]:
    for r in range(scan_start, scan_end):
        for c in range(1, 26):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                from openpyxl.utils import get_column_letter
                print(f'  R{r} {get_column_letter(c)}: {repr(v)}')
        
# Full columns for ANGOSTURA section (rows 15-38)
print('\n=== ANGOSTURA FULL DATA (R15-R38) A-P ===')
for r in range(15, 39):
    vals = []
    for c in range(1, 17):
        v = ws.cell(row=r, column=c).value
        from openpyxl.utils import get_column_letter
        cl = get_column_letter(c)
        vals.append(f'{cl}={v}')
    print(f'  R{r}: {", ".join(vals)}')

# Right side data for ANGOSTURA (cols U-AC)
print('\n=== ANGOSTURA RIGHT SIDE (R15-R38) cols U-AC ===')
for r in range(15, 39):
    vals = []
    for c in range(21, 30):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            vals.append(f'{cl}={v}')
    if vals:
        print(f'  R{r}: {", ".join(vals)}')

# CHICOASEN data structure
print('\n=== CHICOASEN SECTION (R45-R70) A-P ===')
for r in range(45, 71):
    vals = []
    for c in range(1, 17):
        v = ws.cell(row=r, column=c).value
        from openpyxl.utils import get_column_letter
        cl = get_column_letter(c)
        vals.append(f'{cl}={v}')
    if any(ws.cell(row=r, column=c).value is not None for c in range(1, 17)):
        print(f'  R{r}: {", ".join(vals)}')

# Check all columns present for any row
print('\n=== ALL COLUMNS WITH DATA (scan R15 all 95 cols) ===')
for c in range(1, 96):
    v = ws.cell(row=15, column=c).value
    if v is not None:
        from openpyxl.utils import get_column_letter
        print(f'  {get_column_letter(c)}(c{c}): {repr(v)}')

# Check columns Q-T (any section)
print('\n=== COLS Q-T for rows 12-38 ===')
for r in range(12, 39):
    for c in range(17, 21):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            print(f'  R{r} {get_column_letter(c)}: {repr(v)}')
