import openpyxl
wb = openpyxl.load_workbook('/Users/subgerenciagrijalva/Downloads/FIN240326.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]
print(f'Sheet: {wb.sheetnames[0]}, Rows: {ws.max_row}, Cols: {ws.max_column}')

# Headers rows 1-18
print('\n=== ROWS 1-18 (all non-empty, cols A-Y) ===')
for r in range(1, 19):
    for c in range(1, 26):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            print(f'  R{r} {cl}(c{c}): {repr(v)}')

# All dam/unit rows with B values
print('\n=== All B values (unit rows) ===')
for r in range(1, ws.max_row+1):
    b = ws.cell(row=r, column=2).value
    if b is not None:
        c = ws.cell(row=r, column=3).value
        i = ws.cell(row=r, column=9).value
        j = ws.cell(row=r, column=10).value
        m = ws.cell(row=r, column=13).value
        o = ws.cell(row=r, column=15).value
        p = ws.cell(row=r, column=16).value
        s = ws.cell(row=r, column=19).value
        v = ws.cell(row=r, column=22).value
        x = ws.cell(row=r, column=24).value
        print(f'  R{r}: B={b}, C={repr(c)}, I={repr(i)}, J={j}, M={repr(m)}, O={repr(o)}, P={p}, S={s}, V={v}, X={repr(x)}')

# Check merged cells for C (central names)
print('\n=== Merged cells for col C (central names) ===')
for mc in sorted(ws.merged_cells.ranges, key=lambda x: x.min_row):
    if mc.min_col <= 3 and mc.max_col >= 3 and mc.min_row >= 19:
        val = ws.cell(row=mc.min_row, column=3).value
        if val:
            print(f'  {mc}: {repr(val)}')

# Check row 19 data columns - are there actual values or just dates?
print('\n=== Row 19 Z-AE data check ===')
for c in range(26, 32):
    v = ws.cell(row=19, column=c).value
    from openpyxl.utils import get_column_letter
    cl = get_column_letter(c)
    print(f'  {cl}19: {repr(v)} (type: {type(v).__name__})')

# Check OG-OI columns (summary) for first few dam rows
print('\n=== Summary columns OA-OI headers (rows 11-18) ===')
for col_name, col_idx in [('OA',391),('OB',392),('OC',393),('OD',394),('OE',395),('OF',396),('OG',397),('OH',398),('OI',399)]:
    for r in range(11, 19):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None:
            print(f'  R{r} {col_name}(c{col_idx}): {repr(v)}')

print('\n=== Summary data rows 19-22 (OE-OI) ===')
for col_name, col_idx in [('OE',395),('OF',396),('OG',397),('OH',398),('OI',399)]:
    for r in range(19, 27):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None:
            print(f'  R{r} {col_name}: {repr(v)}')

# Check bottom rows
print('\n=== BOTTOM ROWS (last 25) ===')
for r in range(ws.max_row-24, ws.max_row+1):
    for c in range(1, 30):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            print(f'  R{r} {cl}: {repr(v)}')
