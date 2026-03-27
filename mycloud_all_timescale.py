"""
VERSIÓN OPTIMIZADA PARA TIMESCALEDB
- Compresión automática 20:1
- Hypertables (consultas 100x más rápidas)
- Retención automática
- Batch inserts optimizados
"""

import socket
import struct
import hashlib
import time
import os
import threading
import schedule
import psycopg2
import psycopg2.pool
import psycopg2.extensions
from psycopg2.extras import execute_values
import configparser
import datetime
from datetime import timedelta
import pyodbc
import collections
import json as json_module
from flask import Flask, jsonify, request, Response

# ==================== LOG RING BUFFER ====================
_log_buffer = collections.deque(maxlen=500)
_log_lock = threading.Lock()
_start_time = datetime.datetime.now()
_fetch_ok_count = 0
_fetch_fail_count = 0

_original_print = print
def print(*args, **kwargs):
    """Override print para capturar mensajes en el ring buffer."""
    msg = " ".join(str(a) for a in args)
    with _log_lock:
        _log_buffer.append({
            "ts": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "msg": msg
        })
    _original_print(*args, **kwargs)

# ==================== CONFIGURACIÓN ====================
# Cuando se empaqueta con PyInstaller, el cwd puede ser un directorio temporal.
# Usamos la ubicación real del ejecutable para encontrar config.ini
import sys
if getattr(sys, 'frozen', False):
    _base_dir = os.path.dirname(sys.executable)
else:
    _base_dir = os.path.dirname(os.path.abspath(__file__))

config = configparser.ConfigParser()
config.read(os.path.join(_base_dir, 'config.ini'))

# TimescaleDB (PostgreSQL)
PG_HOST = config.get('timescaledb', 'host', fallback='localhost')
PG_PORT = config.getint('timescaledb', 'port', fallback=5432)
PG_USER = config.get('timescaledb', 'user', fallback='postgres')
PG_PASSWORD = config.get('timescaledb', 'password', fallback='Cfe2026##')
PG_DATABASE = config.get('timescaledb', 'database', fallback='mycloud_timescale')

# ==================== CONFIGURACIÓN SQL SERVER ====================
SQL_MODE = config.get('settings', 'sql_mode', fallback='remoto').lower()

if SQL_MODE == 'remoto':
    sql_section = 'sql_server_remote'
else:
    sql_section = 'sql_server_local'

SQL_SERVER = config.get(sql_section, 'server', fallback='127.0.0.1')
SQL_PORT = config.get(sql_section, 'port', fallback='1433')
SQL_DB = config.get(sql_section, 'database', fallback='IGSCLOUD_PRO')
SQL_USER = config.get(sql_section, 'user', fallback='sa')
SQL_PASSWORD = config.get(sql_section, 'password', fallback='Cfe2026$$')

print(f"[CONFIG] SQL Server Mode: {SQL_MODE.upper()}")
print(f"[CONFIG] Conectando a SQL Server: {SQL_SERVER}:{SQL_PORT}/{SQL_DB}")

SQL_CONN_STR = (
    'DRIVER={ODBC Driver 18 for SQL Server};'
    f'SERVER={SQL_SERVER},{SQL_PORT};'
    f'DATABASE={SQL_DB};'
    f'UID={SQL_USER};'
    f'PWD={SQL_PASSWORD};'
    'TrustServerCertificate=yes;'
    'Encrypt=no;'
)



# LRGS (desde config.ini)
DEFAULT_USER = config.get('lrgs', 'default_user', fallback='cilasur')
DEFAULT_PASSWORD = config.get('lrgs', 'default_password', fallback='ksip-CWNG-05')

HOSTS = [h.strip() for h in config.get('lrgs', 'hosts', fallback='10.41.75.100,lrgseddn1.cr.usgs.gov,lrgseddn2.cr.usgs.gov,lrgseddn3.cr.usgs.gov').split(',')]

# Cargar credenciales especiales por host desde config.ini
CREDENTIALS = {}
for key, value in config.items('lrgs') if config.has_section('lrgs') else []:
    if key.startswith('credentials_'):
        host_name = key[len('credentials_'):]
        parts = value.split(':')
        if len(parts) == 2:
            CREDENTIALS[host_name] = {"user": parts[0], "password": parts[1]}

OUT_DIR = os.path.join(_base_dir, "dds_last")
os.makedirs(OUT_DIR, exist_ok=True)

import platform
if platform.system() == 'Windows':
    MIS_OUT_DIR = config.get('paths', 'mis_output_dir_windows', fallback='C:\\IGSCLOUD\\Datos\\GOES')
else:
    MIS_OUT_DIR = config.get('paths', 'mis_output_dir_mac', fallback='./Datos/GOES')
os.makedirs(MIS_OUT_DIR, exist_ok=True)

# ==================== FUNCIONES AUXILIARES ====================
def generar_archivo_mis(dcp_id, meta, raw_header, raw_payload, valores, estacion_data, sensores_list):
    """
    Genera un archivo .mis con el formato específico solicitado.
    """
    try:
        timestamp = meta['timestamp']
        # Nombre de archivo: DCPID_YYYYMMDDHHMMEX.mis
        filename = f"{dcp_id}_{timestamp.strftime('%Y%m%d%H%M')}EX.mis"
        filepath = os.path.join(MIS_OUT_DIR, filename)
        
        id_asignado = estacion_data.get("id_asignado", "")
        
        # Preparar datos para el cuerpo
        # Reutilizamos lógica de cálculo de valores y timestamps
        lines_body = []
        
        # Asegurar que los sensores estén ordenados por número
        sensores_ordenados = sorted(sensores_list, key=lambda s: s["numero_sensor"])

        for sensor in sensores_ordenados:
            sensor_id_str = sensor["numero_sensor"] # Ya viene con zfill(4) desde la carga
            
            # XML Header por sensor
            lines_body.append(f"<STATION>{id_asignado}</STATION><SENSOR>{sensor_id_str}</SENSOR><DATEFORMAT>YYYYMMDD</DATEFORMAT>")
            
            for i, pos in enumerate(sensor["posiciones"]):
                if pos >= len(valores):
                    continue
                
                # Calcular Valor
                valor_raw = valores[pos]
                valor = round(valor_raw / (10 ** sensor["decimales"]), sensor["decimales"])
                # valor += sensor.get("valor_cota", 0.0)
                
                # Calcular Timestamp
                ts = timestamp - datetime.timedelta(minutes=i * sensor["periodo"])
                ts = hora_concurrente(ts, sensor["periodo"])
                
                # Formato: YYYYMMDD;HH:MM:SS;Valor
                date_str = ts.strftime('%Y%m%d')
                time_str = ts.strftime('%H:%M:%S')
                # Formatear valor: eliminar .0 si es entero, o mostrar decimales si corresponde. 
                # El ejemplo muestra -1310.72 y 300.8 y 0.
                valor_str = f"{valor:g}" 
                
                lines_body.append(f"{date_str};{time_str};{valor_str}")

        with open(filepath, "w", encoding="utf-8") as f:
            # 1. Encabezado (Línea 1)
            # Formato: DCP_ID,DD/MM/YYYY HH:MM:SS,SIGNAL,Good Msg!,CHANNEL
            # Signal: +XXdBm
            sig = meta.get('sig', '00')
            try:
                sig_val = int(sig)
                sig_str = f"+{sig_val}dBm"
            except:
                sig_str = f"{sig}dBm"
            
            channel = str(meta.get('channel', '000')).zfill(3)
            date_header = timestamp.strftime('%d/%m/%Y %H:%M:%S')
            
            header_line = f"{dcp_id},{date_header},{sig_str},Good Msg!,{channel}"
            f.write(header_line + "\n")
            
            # 2. Cuerpo
            for line in lines_body:
                f.write(line + "\n")
            
            # 3. Pie de página (Footer)
            # {<RAW_HEADER><RAW_PAYLOAD>}
            # Raw header son 37 bytes. Raw payload es el resto.
            # Concatenamos bytes y decodificamos a ascii (latin-1 para preservar bytes) o escribimos en modo binario?
            # El archivo original es texto, así que decodificamos 'latin-1' para no perder bytes si hubiera raros,
            # aunque DOMSAT suele ser ASCII.
            
            raw_content = raw_header + raw_payload
            footer_str = raw_content.decode('latin-1', errors='replace')
            f.write(f"{{{footer_str}}}\n")
            
        print(f"[MIS] Generado: {filename}")

    except Exception as e:
        print(f"[ERROR MIS] Falló generación de archivo .mis: {e}")


def importar_archivo_mis(filepath, sensores_global, estaciones_global):
    """
    Importa un archivo .mis histórico e inserta los datos en PostgreSQL.
    Formato esperado:
      Línea 1: DCP_ID,DD/MM/YYYY HH:MM:SS,signal,quality,channel
      Bloques: <STATION>ID</STATION><SENSOR>NNNN</SENSOR><DATEFORMAT>YYYYMMDD</DATEFORMAT>
               YYYYMMDD;HH:MM:SS;valor
               ...
      Footer (opcional): {raw_data}
    """
    import re
    registros = []
    nombre_archivo = os.path.basename(filepath)

    try:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            contenido = f.read()
    except Exception as e:
        print(f"  [ERROR] No se pudo leer {nombre_archivo}: {e}")
        return 0

    lines = contenido.split('\n')
    if not lines:
        print(f"  [SKIP] {nombre_archivo}: archivo vacío")
        return 0

    # === Parsear Header (línea 1) ===
    header_parts = lines[0].split(',')
    dcp_id = header_parts[0].strip().upper()

    # Buscar id_asignado desde ESTACIONES
    id_asignado = None
    for eid, edata in estaciones_global.items():
        if eid == dcp_id:
            id_asignado = edata.get("id_asignado", "")
            break

    if not id_asignado:
        # Intentar extraer de las líneas <STATION>
        match = re.search(r'<STATION>([^<]+)</STATION>', contenido)
        if match:
            id_asignado = match.group(1).strip()
        else:
            print(f"  [SKIP] {nombre_archivo}: no se encontró id_asignado para {dcp_id}")
            return 0

    # Obtener sensores configurados
    sensores_cfg = sensores_global.get(id_asignado, [])
    sensor_info = {}
    for s in sensores_cfg:
        sensor_info[s["numero_sensor"]] = {
            "nombre": s["nombre"].lower().replace(" ", "_"),
            "tipo": s.get("tipo", s["nombre"]).lower(),
            "valor_minimo": s.get("valor_minimo"),
            "valor_maximo": s.get("valor_maximo"),
        }

    # === Parsear bloques de datos ===
    current_sensor_id = None
    station_re = re.compile(r'<STATION>([^<]+)</STATION><SENSOR>([^<]+)</SENSOR>')

    for line in lines[1:]:
        line = line.strip()
        if not line:
            continue

        # Footer: línea que empieza con {
        if line.startswith('{'):
            break

        # Línea de bloque sensor
        m = station_re.search(line)
        if m:
            current_sensor_id = m.group(2).strip().zfill(4)  # Normalizar a 4 dígitos como en SENSORES
            continue

        # Línea de datos: YYYYMMDD;HH:MM:SS;valor
        parts = line.split(';')
        if len(parts) == 3 and current_sensor_id:
            try:
                fecha_str, hora_str, valor_str = parts
                ts = datetime.datetime.strptime(
                    f"{fecha_str} {hora_str}", "%Y%m%d %H:%M:%S"
                ).replace(tzinfo=datetime.timezone.utc)
                valor = float(valor_str)

                # Validar contra límites del sensor
                info = sensor_info.get(current_sensor_id, {})
                variable = info.get("nombre", f"sensor_{current_sensor_id}")
                tipo = info.get("tipo", variable)
                valido = True
                descripcion = "OK (importado)"

                vmin = info.get("valor_minimo")
                vmax = info.get("valor_maximo")
                if vmin is not None and valor < vmin:
                    valido = False
                    descripcion = f"Bajo: {valor} < {vmin}"
                elif vmax is not None and valor > vmax:
                    valido = False
                    descripcion = f"Alto: {valor} > {vmax}"

                registros.append((
                    ts, dcp_id, id_asignado, current_sensor_id,
                    variable, valor, tipo, valido, descripcion[:200],
                    ts.year, ts.month, ts.day, ts.hour, ts.minute
                ))
            except (ValueError, IndexError):
                continue

    if not registros:
        print(f"  [SKIP] {nombre_archivo}: sin registros válidos")
        return 0

    # === Insertar en PostgreSQL ===
    conn = get_pg_conn()
    if not conn:
        print(f"  [ERROR] {nombre_archivo}: sin conexión PG")
        return 0
    try:
        cur = conn.cursor()
        execute_values(cur, """
            INSERT INTO dcp_datos
            (ts, dcp_id, id_asignado, sensor_id, variable, valor, tipo,
             valido, descripcion, anio, mes, dia, hora, minuto)
            VALUES %s
            ON CONFLICT (ts, dcp_id, sensor_id) DO UPDATE SET
                valor = EXCLUDED.valor,
                valido = EXCLUDED.valido,
                descripcion = EXCLUDED.descripcion
        """, registros, page_size=500)
        conn.commit()
        cur.close()
        print(f"  [OK] {nombre_archivo}: {len(registros)} registros importados ({dcp_id} / {id_asignado})")
        return len(registros)
    except Exception as e:
        conn.rollback()
        print(f"  [ERROR] {nombre_archivo}: {e}")
        return 0
    finally:
        release_pg_conn(conn)


def importar_directorio_mis(directorio, sensores_global, estaciones_global, patron="*.mis"):
    """Importa todos los archivos .mis de un directorio."""
    import glob
    archivos = sorted(glob.glob(os.path.join(directorio, patron)))
    if not archivos:
        print(f"[IMPORT] No se encontraron archivos {patron} en {directorio}")
        return

    print(f"[IMPORT] {len(archivos)} archivos encontrados en {directorio}")
    total_registros = 0
    exitosos = 0
    for i, filepath in enumerate(archivos, 1):
        print(f"[IMPORT] [{i}/{len(archivos)}] {os.path.basename(filepath)}")
        n = importar_archivo_mis(filepath, sensores_global, estaciones_global)
        total_registros += n
        if n > 0:
            exitosos += 1

    print(f"\n[IMPORT] Resumen: {exitosos}/{len(archivos)} archivos, {total_registros} registros totales")


def importar_mis_automatico():
    """Importa archivos .mis desde la carpeta de entrada y los mueve a procesados.
    Se ejecuta periódicamente desde el scheduler."""
    import glob
    import shutil

    mis_import_dir = config.get('paths', 'mis_import_dir_windows' if platform.system() == 'Windows' else 'mis_import_dir_mac', fallback='')
    if not mis_import_dir or not os.path.isdir(mis_import_dir):
        return  # No configurado o no existe, silenciosamente ignorar

    procesados_dir = os.path.join(mis_import_dir, 'procesados')
    errores_dir = os.path.join(mis_import_dir, 'errores')

    archivos = sorted(glob.glob(os.path.join(mis_import_dir, '*.mis')))
    if not archivos:
        return  # Sin archivos pendientes

    os.makedirs(procesados_dir, exist_ok=True)
    os.makedirs(errores_dir, exist_ok=True)

    print(f"[AUTO-IMPORT] {len(archivos)} archivos .mis pendientes en {mis_import_dir}")
    total_reg = 0
    for filepath in archivos:
        nombre = os.path.basename(filepath)
        n = importar_archivo_mis(filepath, SENSORES, ESTACIONES)
        if n > 0:
            shutil.move(filepath, os.path.join(procesados_dir, nombre))
            total_reg += n
        else:
            shutil.move(filepath, os.path.join(errores_dir, nombre))

    print(f"[AUTO-IMPORT] Finalizado: {total_reg} registros importados, {len(archivos)} archivos procesados")


# ==================== IMPORTACIÓN AUTOMÁTICA DE EXCEL FUNVASOS ====================

# Secciones del Excel FunVasos: presa → (headerRow, dataStartRow, dataEndRow, isTapon)
_FUNVASOS_SECCIONES = {
    "Angostura":           (12, 15, 38, False),
    "Chicoasén":           (45, 48, 70, False),
    "Malpaso":             (78, 81, 103, False),
    "Tapón Juan Grijalva": (111, 114, 136, True),
    "Peñitas":             (144, 147, 170, False),
}

# Meses en español para parseo manual de fecha
_MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12
}


def _funvasos_extraer_fecha(ws):
    """Extrae la fecha del reporte desde la hoja FIN del Excel.
    Intenta celdas H4 (datetime directo) o H3 (texto en español)."""
    from datetime import datetime as dt_cls
    # Intenta H4 como datetime
    val = ws.cell(row=4, column=8).value
    if isinstance(val, dt_cls):
        return val.replace(hour=0, minute=0, second=0, microsecond=0)
    # Intenta H3 como texto "24 de Marzo de 2026"
    txt = ws.cell(row=3, column=8).value
    if txt and isinstance(txt, str):
        partes = txt.strip().split()
        if len(partes) >= 5:
            try:
                dia = int(partes[0])
                mes = _MESES_ES.get(partes[2].lower())
                anio = int(partes[4])
                if mes:
                    return dt_cls(anio, mes, dia)
            except (ValueError, IndexError):
                pass
    return None


def _funvasos_get_float(ws, row, col):
    """Lee un valor float de una celda, retorna None si vacía o no numérica."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _funvasos_get_short(ws, row, col):
    """Lee un valor entero corto de una celda."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _funvasos_parsear_excel(filepath):
    """Parsea un archivo Excel de FunVasos y retorna (fecha, registros, errores).
    Replica la lógica de FunVasosService.ParseAndStoreAsync() del C#."""
    from openpyxl import load_workbook

    errores = []
    registros = []

    try:
        # NO usar read_only=True: el acceso aleatorio a celdas por fila/columna
        # falla en modo streaming (retorna None para celdas no iteradas secuencialmente).
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return None, [], [f"Error abriendo Excel: {e}"]

    if "FIN" not in wb.sheetnames:
        wb.close()
        return None, [], ["No se encontró la hoja 'FIN' en el archivo Excel."]

    ws = wb["FIN"]

    # Extraer fecha del reporte
    fecha = _funvasos_extraer_fecha(ws)
    if fecha is None:
        wb.close()
        return None, [], ["No se pudo extraer la fecha del reporte."]

    # Parsear datos por presa
    for presa, (header_row, data_start, data_end, is_tapon) in _FUNVASOS_SECCIONES.items():
        presa_count = 0
        for row in range(data_start, data_end + 1):
            hora_val = _funvasos_get_float(ws, row, 1)  # Columna A = Hora
            if hora_val is None or hora_val < 1 or hora_val > 24:
                continue

            hora = int(hora_val)

            # Si elevación vacía, saltar fila
            elev = _funvasos_get_float(ws, row, 2)  # Columna B
            if elev is None:
                continue

            registro = {
                "ts": fecha,
                "presa": presa,
                "hora": hora,
                "elevacion": elev,
                "almacenamiento": _funvasos_get_float(ws, row, 3),
                "diferencia": _funvasos_get_float(ws, row, 4),
                "aportaciones_q": _funvasos_get_float(ws, row, 5),
                "aportaciones_v": _funvasos_get_float(ws, row, 6),
                "extracciones_turb_q": _funvasos_get_float(ws, row, 7),
                "extracciones_turb_v": _funvasos_get_float(ws, row, 8),
                "extracciones_vert_q": _funvasos_get_float(ws, row, 9),
                "extracciones_vert_v": _funvasos_get_float(ws, row, 10),
                "extracciones_total_q": _funvasos_get_float(ws, row, 11),
                "extracciones_total_v": _funvasos_get_float(ws, row, 12),
                "generacion": None if is_tapon else _funvasos_get_float(ws, row, 13),
                "num_unidades": None if is_tapon else _funvasos_get_short(ws, row, 14),
                "aportacion_cuenca_propia": _funvasos_get_float(ws, row, 15),
                "aportacion_promedio": _funvasos_get_float(ws, row, 16),
            }
            registros.append(registro)
            presa_count += 1

        print(f"[FUNVASOS]   {presa}: {presa_count} horas parseadas")

    wb.close()

    if not registros:
        errores.append("No se encontraron datos horarios en el archivo.")

    return fecha, registros, errores


def _funvasos_insertar_pg(fecha, registros):
    """Inserta registros de FunVasos en PostgreSQL (upsert: elimina fecha existente + inserta)."""
    conn = get_pg_conn()
    try:
        cur = conn.cursor()
        # Eliminar datos existentes para esta fecha
        cur.execute("DELETE FROM public.funvasos_horario WHERE ts = %s", (fecha,))

        # Insertar todos los registros
        insert_sql = """
            INSERT INTO public.funvasos_horario
            (ts, presa, hora, elevacion, almacenamiento, diferencia,
             aportaciones_q, aportaciones_v, extracciones_turb_q, extracciones_turb_v,
             extracciones_vert_q, extracciones_vert_v, extracciones_total_q, extracciones_total_v,
             generacion, num_unidades, aportacion_cuenca_propia, aportacion_promedio)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        for r in registros:
            cur.execute(insert_sql, (
                r["ts"], r["presa"], r["hora"], r["elevacion"],
                r["almacenamiento"], r["diferencia"],
                r["aportaciones_q"], r["aportaciones_v"],
                r["extracciones_turb_q"], r["extracciones_turb_v"],
                r["extracciones_vert_q"], r["extracciones_vert_v"],
                r["extracciones_total_q"], r["extracciones_total_v"],
                r["generacion"], r["num_unidades"],
                r["aportacion_cuenca_propia"], r["aportacion_promedio"],
            ))

        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        print(f"[FUNVASOS] Error insertando en PostgreSQL: {e}")
        return False
    finally:
        release_pg_conn(conn)


def importar_funvasos_automatico():
    """Monitorea carpeta de entrada, parsea Excel FunVasos, inserta en DB y archiva.
    Se ejecuta periódicamente desde el scheduler."""
    import glob as glob_mod
    import shutil

    inbox_key = 'funvasos_inbox_windows' if platform.system() == 'Windows' else 'funvasos_inbox_mac'
    inbox_dir = config.get('paths', inbox_key, fallback='')
    if not inbox_dir or not os.path.isdir(inbox_dir):
        return

    repo_key = 'funvasos_repo_windows' if platform.system() == 'Windows' else 'funvasos_repo_mac'
    repo_dir = config.get('paths', repo_key, fallback='')

    procesados_dir = os.path.join(inbox_dir, 'procesados')
    errores_dir = os.path.join(inbox_dir, 'errores')

    # Buscar archivos .xlsx y .xls
    archivos = sorted(
        glob_mod.glob(os.path.join(inbox_dir, '*.xlsx')) +
        glob_mod.glob(os.path.join(inbox_dir, '*.xls')) +
        glob_mod.glob(os.path.join(inbox_dir, '*.xlsm'))
    )
    if not archivos:
        return

    os.makedirs(procesados_dir, exist_ok=True)
    os.makedirs(errores_dir, exist_ok=True)
    if repo_dir:
        os.makedirs(repo_dir, exist_ok=True)

    print(f"[FUNVASOS] {len(archivos)} archivos Excel pendientes en {inbox_dir}")

    for filepath in archivos:
        nombre = os.path.basename(filepath)
        print(f"[FUNVASOS] Procesando: {nombre}")

        fecha, registros, errores = _funvasos_parsear_excel(filepath)

        if errores:
            for err in errores:
                print(f"[FUNVASOS]   Error: {err}")

        if not fecha or not registros:
            print(f"[FUNVASOS]   → Movido a errores/")
            shutil.move(filepath, os.path.join(errores_dir, nombre))
            continue

        # Insertar en PostgreSQL
        ok = _funvasos_insertar_pg(fecha, registros)
        if not ok:
            print(f"[FUNVASOS]   → Error BD, movido a errores/")
            shutil.move(filepath, os.path.join(errores_dir, nombre))
            continue

        # Copiar al repositorio de documentos con nombre estándar FINddmmyy.xlsx
        if repo_dir:
            ext = os.path.splitext(nombre)[1]
            repo_name = f"FIN{fecha.strftime('%d%m%y')}{ext}"
            repo_path = os.path.join(repo_dir, repo_name)
            try:
                shutil.copy2(filepath, repo_path)
                print(f"[FUNVASOS]   → Archivado en repo: {repo_name}")
            except Exception as e:
                print(f"[FUNVASOS]   Advertencia copiando a repo: {e}")

        # Mover a procesados
        shutil.move(filepath, os.path.join(procesados_dir, nombre))
        print(f"[FUNVASOS]   → {len(registros)} registros importados ({fecha.strftime('%d/%m/%Y')})")

    print(f"[FUNVASOS] Procesamiento completado.")


# Pool de conexiones a TimescaleDB (reutiliza conexiones en lugar de abrir/cerrar cada vez)
_pg_pool = None
_pg_pool_lock = threading.Lock()

def _create_pg_pool():
    global _pg_pool
    _pg_pool = psycopg2.pool.ThreadedConnectionPool(
        minconn=2,
        maxconn=10,
        host=PG_HOST,
        port=PG_PORT,
        user=PG_USER,
        password=PG_PASSWORD,
        database=PG_DATABASE,
        client_encoding='UTF8',
        options='-c client_encoding=UTF8',
        # Evitar conexiones colgadas
        connect_timeout=10,
        keepalives=1,
        keepalives_idle=30,
        keepalives_interval=10,
        keepalives_count=3
    )

_create_pg_pool()

def get_pg_conn():
    """Obtener conexión del pool de TimescaleDB con recuperación automática."""
    global _pg_pool
    for intento in range(3):
        try:
            conn = _pg_pool.getconn()
            # Verificar que la conexión sigue viva
            try:
                if conn.closed:
                    raise Exception("Conexión cerrada")
                cur = conn.cursor()
                cur.execute("SELECT 1")
                cur.close()
            except Exception:
                # Conexión muerta, descartarla y pedir otra
                try:
                    _pg_pool.putconn(conn, close=True)
                except Exception:
                    pass
                if intento < 2:
                    continue
                raise
            return conn
        except psycopg2.pool.PoolError as e:
            print(f"[PG POOL] Pool agotado intento {intento+1}/3 ({e})")
            if intento < 2:
                time.sleep(2)  # Esperar a que se liberen conexiones
                continue
            print(f"[PG POOL] Recreando pool...")
            with _pg_pool_lock:
                try:
                    _pg_pool.closeall()
                except Exception:
                    pass
                _create_pg_pool()
            return _pg_pool.getconn()
    raise Exception("No se pudo obtener conexión PG después de 3 intentos")

def release_pg_conn(conn):
    """Devolver conexión al pool, reseteando si está en mal estado."""
    if conn:
        try:
            # Si la conexión tiene una transacción pendiente, hacer rollback
            if conn.status == psycopg2.extensions.TRANSACTION_STATUS_INERROR:
                conn.rollback()
            elif conn.status == psycopg2.extensions.TRANSACTION_STATUS_INTRANS:
                conn.rollback()
            _pg_pool.putconn(conn)
        except Exception:
            # Si no se puede devolver, cerrarla para no bloquear el pool
            try:
                _pg_pool.putconn(conn, close=True)
            except Exception:
                try:
                    conn.close()
                except Exception:
                    pass

def get_sql_conn():
    """Conexión a SQL Server (solo para config)"""
    return pyodbc.connect(SQL_CONN_STR)

def get_credentials(host):
    creds = CREDENTIALS.get(host)
    if creds:
        return creds["user"], creds["password"]
    return DEFAULT_USER, DEFAULT_PASSWORD


# ==================== CREACIÓN AUTOMÁTICA DE TABLAS ====================
_tables_ensured = False

def _ensure_pg_tables():
    """Crea todas las tablas en TimescaleDB si no existen (idempotente)."""
    global _tables_ensured
    if _tables_ensured:
        return
    conn = get_pg_conn()
    try:
        cur = conn.cursor()

        # --- dcp_datos (hypertable) ---
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.dcp_datos (
                ts TIMESTAMP WITH TIME ZONE NOT NULL,
                dcp_id CHARACTER VARYING(20) NOT NULL,
                id_asignado CHARACTER VARYING(50),
                sensor_id CHARACTER VARYING(20) NOT NULL,
                variable CHARACTER VARYING(50),
                valor REAL,
                tipo CHARACTER VARYING(50),
                valido BOOLEAN,
                descripcion CHARACTER VARYING(200),
                anio SMALLINT,
                mes SMALLINT,
                dia SMALLINT,
                hora SMALLINT,
                minuto SMALLINT,
                CONSTRAINT dcp_datos_pkey PRIMARY KEY (ts, dcp_id, sensor_id)
            );
        """)
        # Intentar crear hypertable (ignora si ya existe)
        try:
            cur.execute("SELECT create_hypertable('public.dcp_datos', 'ts', if_not_exists => TRUE);")
        except Exception:
            conn.rollback()

        # --- resumen_horario (hypertable) ---
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.resumen_horario (
                ts TIMESTAMP WITH TIME ZONE NOT NULL,
                dcp_id CHARACTER VARYING(20) NOT NULL,
                sensor_id CHARACTER VARYING(20) NOT NULL,
                id_asignado CHARACTER VARYING(50),
                variable CHARACTER VARYING(50),
                tipo CHARACTER VARYING(50),
                suma REAL,
                conteo INTEGER,
                promedio REAL,
                minimo REAL,
                maximo REAL,
                acumulado REAL,
                CONSTRAINT resumen_horario_pkey PRIMARY KEY (ts, dcp_id, sensor_id)
            );
        """)
        try:
            cur.execute("SELECT create_hypertable('public.resumen_horario', 'ts', if_not_exists => TRUE);")
        except Exception:
            conn.rollback()

        # --- resumen_diario (hypertable) ---
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.resumen_diario (
                fecha DATE NOT NULL,
                dcp_id CHARACTER VARYING(20) NOT NULL,
                sensor_id CHARACTER VARYING(20) NOT NULL,
                id_asignado CHARACTER VARYING(50),
                variable CHARACTER VARYING(50),
                tipo CHARACTER VARYING(50),
                suma REAL,
                conteo INTEGER,
                promedio REAL,
                minimo REAL,
                maximo REAL,
                acumulado REAL,
                CONSTRAINT resumen_diario_pkey PRIMARY KEY (fecha, dcp_id, sensor_id)
            );
        """)
        try:
            cur.execute("SELECT create_hypertable('public.resumen_diario', 'fecha', if_not_exists => TRUE);")
        except Exception:
            conn.rollback()

        # --- bitacora_goes ---
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.bitacora_goes (
                id SERIAL PRIMARY KEY,
                dcp_id CHARACTER VARYING(20) NOT NULL,
                timestamp_utc TIMESTAMP WITH TIME ZONE NOT NULL,
                timestamp_msg TIMESTAMP WITH TIME ZONE,
                servidor CHARACTER VARYING(200),
                exito BOOLEAN
            );
        """)

        # --- dcp_headers ---
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.dcp_headers (
                dcp_id CHARACTER VARYING(20) NOT NULL,
                timestamp_msg TIMESTAMP WITH TIME ZONE NOT NULL,
                ts TIMESTAMP WITH TIME ZONE,
                failure_code CHARACTER VARYING(10),
                signal_strength SMALLINT,
                frequency_offset CHARACTER VARYING(10),
                mod_index CHARACTER VARYING(10),
                data_quality CHARACTER VARYING(10),
                channel SMALLINT,
                spacecraft CHARACTER VARYING(10),
                data_source CHARACTER VARYING(10),
                CONSTRAINT dcp_headers_pkey PRIMARY KEY (dcp_id, timestamp_msg)
            );
        """)

        # --- ultimas_mediciones ---
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.ultimas_mediciones (
                dcp_id CHARACTER VARYING NOT NULL,
                variable CHARACTER VARYING NOT NULL,
                ts TIMESTAMP WITH TIME ZONE,
                valor REAL,
                tipo CHARACTER VARYING,
                PRIMARY KEY (dcp_id, variable)
            );
        """)

        # --- estatus_estaciones ---
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.estatus_estaciones (
                dcp_id CHARACTER VARYING(20) NOT NULL PRIMARY KEY,
                fecha_calculo TIMESTAMP WITH TIME ZONE,
                intervalo_minutos INTEGER DEFAULT 60,
                mensajes_esperados_48h INTEGER,
                mensajes_recibidos_48h INTEGER DEFAULT 0,
                color_estatus CHARACTER VARYING(20) DEFAULT 'NEGRO',
                fecha_ultima_tx TIMESTAMP WITH TIME ZONE,
                servidor CHARACTER VARYING(50)
            );
        """)

        # --- Trigger para actualizar estatus ---
        cur.execute("""
            CREATE OR REPLACE FUNCTION public.trg_actualizar_estatus()
            RETURNS trigger
            LANGUAGE plpgsql
            AS $function$
            DECLARE
                v_rec_count INT;
                v_intervalo INT;
                v_status VARCHAR(20);
                v_window_start TIMESTAMP WITH TIME ZONE;
            BEGIN
                v_window_start := NOW() - INTERVAL '48 hours';
                SELECT COALESCE(intervalo_minutos, 60) INTO v_intervalo
                FROM public.estatus_estaciones
                WHERE dcp_id = NEW.dcp_id;
                IF v_intervalo IS NULL THEN v_intervalo := 60; END IF;
                SELECT COUNT(*) INTO v_rec_count
                FROM public.bitacora_goes
                WHERE dcp_id = NEW.dcp_id
                  AND timestamp_utc >= v_window_start;
                IF v_rec_count = 0 THEN
                    v_status := 'NEGRO';
                ELSIF v_rec_count >= (48 * 60 / v_intervalo) THEN
                    v_status := 'VERDE';
                ELSIF (48 * 60 / v_intervalo) - v_rec_count <= 2 THEN
                    v_status := 'AMARILLO';
                ELSE
                    v_status := 'ROJO';
                END IF;
                INSERT INTO public.estatus_estaciones (
                    dcp_id, fecha_calculo, mensajes_recibidos_48h,
                    fecha_ultima_tx, servidor, color_estatus
                )
                VALUES (
                    NEW.dcp_id, NOW(), v_rec_count,
                    NEW.timestamp_utc, NEW.servidor, v_status
                )
                ON CONFLICT (dcp_id) DO UPDATE SET
                    fecha_calculo = EXCLUDED.fecha_calculo,
                    mensajes_recibidos_48h = EXCLUDED.mensajes_recibidos_48h,
                    fecha_ultima_tx = EXCLUDED.fecha_ultima_tx,
                    servidor = EXCLUDED.servidor,
                    color_estatus = EXCLUDED.color_estatus;
                RETURN NEW;
            END;
            $function$;
        """)

        # Crear trigger solo si no existe
        cur.execute("""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM pg_trigger WHERE tgname = 'trg_update_estatus_on_insert'
                ) THEN
                    CREATE TRIGGER trg_update_estatus_on_insert
                    AFTER INSERT ON bitacora_goes
                    FOR EACH ROW EXECUTE FUNCTION trg_actualizar_estatus();
                END IF;
            END $$;
        """)

        # --- Trigger para actualizar ultimas_mediciones ---
        cur.execute("""
            CREATE OR REPLACE FUNCTION public.trg_actualizar_ultima_medicion()
            RETURNS trigger
            LANGUAGE plpgsql
            AS $function$
            BEGIN
                INSERT INTO public.ultimas_mediciones (dcp_id, variable, ts, valor, tipo)
                VALUES (NEW.dcp_id, NEW.variable, NEW.ts, NEW.valor, NEW.tipo)
                ON CONFLICT (dcp_id, variable) DO UPDATE SET
                    ts = CASE WHEN EXCLUDED.ts >= ultimas_mediciones.ts THEN EXCLUDED.ts ELSE ultimas_mediciones.ts END,
                    valor = CASE WHEN EXCLUDED.ts >= ultimas_mediciones.ts THEN EXCLUDED.valor ELSE ultimas_mediciones.valor END,
                    tipo = CASE WHEN EXCLUDED.ts >= ultimas_mediciones.ts THEN EXCLUDED.tipo ELSE ultimas_mediciones.tipo END;
                RETURN NEW;
            END;
            $function$;
        """)

        cur.execute("""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM pg_trigger WHERE tgname = 'trg_update_last_val'
                ) THEN
                    CREATE TRIGGER trg_update_last_val
                    AFTER INSERT ON dcp_datos
                    FOR EACH ROW EXECUTE FUNCTION trg_actualizar_ultima_medicion();
                END IF;
            END $$;
        """)

        # --- funvasos_horario (datos horarios de presas) ---
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.funvasos_horario (
                ts TIMESTAMP NOT NULL,
                presa CHARACTER VARYING(100) NOT NULL,
                hora SMALLINT NOT NULL,
                elevacion REAL,
                almacenamiento REAL,
                diferencia REAL,
                aportaciones_q REAL,
                aportaciones_v REAL,
                extracciones_turb_q REAL,
                extracciones_turb_v REAL,
                extracciones_vert_q REAL,
                extracciones_vert_v REAL,
                extracciones_total_q REAL,
                extracciones_total_v REAL,
                generacion REAL,
                num_unidades SMALLINT,
                aportacion_cuenca_propia REAL,
                aportacion_promedio REAL,
                CONSTRAINT funvasos_horario_pkey PRIMARY KEY (ts, presa, hora)
            );
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_funvasos_presa
            ON public.funvasos_horario(presa);
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_funvasos_ts
            ON public.funvasos_horario(ts DESC);
        """)

        conn.commit()
        cur.close()
        _tables_ensured = True
        print("[PG] ✅ Tablas y triggers verificados/creados")

    except Exception as e:
        print(f"[ERROR] Falló verificación de tablas: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        release_pg_conn(conn)


# ==================== CARGA ESTACIONES ====================
def load_estaciones_from_sql():
    try:
        print("[DB] Conectando a SQL Server para configuración...")
        conn = get_sql_conn()
        cursor = conn.cursor()
        
        query = """
        SELECT [IdSatelital], [Minuto], [Nombre], [IdAsignado], [RangoTransmision]
        FROM [dbo].[NV_GoesSGD]
        WHERE [Minuto] IS NOT NULL 
          AND [Minuto] BETWEEN 0 AND 59
          AND [IdEstacion] IS NOT NULL
          AND [Decodificador] = 'GS300'          
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        
        estaciones = {}
        for row in rows:
            dcp_id = str(row.IdSatelital).strip().upper()
            minuto = int(row.Minuto)
            nombre = str(row.Nombre or "Sin nombre").strip()
            id_asignado = str(row.IdAsignado or "").strip()
            
            # Parsear RangoTransmision de HH:MM:SS a minutos totales
            rango_transmision = 60  # Default: 1 hora
            if row.RangoTransmision:
                try:
                    partes = str(row.RangoTransmision).split(':')
                    if len(partes) == 3:
                        horas = int(partes[0])
                        minutos = int(partes[1])
                        rango_transmision = (horas * 60) + minutos
                except Exception as e:
                    print(f"[WARN] Error parseando RangoTransmision '{row.RangoTransmision}': {e}")
            
            estaciones[dcp_id] = {
                "minuto": minuto,
                "nombre": nombre,
                "id_asignado": id_asignado,
                "rango_transmision": rango_transmision
            }
        
        conn.close()
        print(f"[DB] {len(estaciones)} estaciones GS300 cargadas")
        return estaciones

    except Exception as e:
        print(f"[ERROR] Falló SQL: {e}")
        return {}


# ==================== CARGA CUENCAS / SUBCUENCAS ====================
def load_cuencas_from_sql():
    """Carga mapeo de IdAsignado → (cuenca, subcuenca) desde SQL Server.
    Retorna dict keyed por id_asignado → {'cuenca': str, 'subcuenca': str}."""
    try:
        conn = get_sql_conn()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT e.IdAsignado,
                   ISNULL(c.Nombre, '') AS Cuenca,
                   ISNULL(sc.Nombre, '') AS Subcuenca
            FROM Estacion e
            LEFT JOIN Cuenca c ON e.IdCuenca = c.Id
            LEFT JOIN Subcuenca sc ON e.IdSubcuenca = sc.Id
            WHERE e.Activo = 1
        """)
        cuencas = {}
        for row in cursor.fetchall():
            id_asignado = str(row.IdAsignado).strip()
            cuencas[id_asignado] = {
                "cuenca": str(row.Cuenca).strip(),
                "subcuenca": str(row.Subcuenca).strip()
            }
        conn.close()
        cuencas_unicas = set(v["cuenca"] for v in cuencas.values() if v["cuenca"])
        subcuencas_unicas = set(v["subcuenca"] for v in cuencas.values() if v["subcuenca"])
        print(f"[CUENCAS] {len(cuencas)} estaciones mapeadas | "
              f"{len(cuencas_unicas)} cuencas, {len(subcuencas_unicas)} subcuencas")
        return cuencas
    except Exception as e:
        print(f"[ERROR] Falló carga de cuencas: {e}")
        return {}

# ==================== CARGA SENSORES ====================
def load_sensores_from_sql():
    try:
        conn = get_sql_conn()
        cursor = conn.cursor()
        
        # Leer directo de Sensor + Estacion + TipoSensor + CotaSensor
        # para obtener los ValorMinimo/ValorMaximo actualizados por el web admin.
        # NV_SensoresSGD es una vista/tabla legacy que puede no reflejar cambios.
        query = """
        SELECT 
            e.IdAsignado,
            s.NumeroSensor,
            ts.Nombre AS Sensor,
            ISNULL(s.PuntoDecimal, 2) AS PuntoDecimal,
            s.PeriodoMuestra,
            s.ValorMinimo,
            s.ValorMaximo,
            ISNULL(
                (SELECT TOP 1 
                    CASE WHEN cs.Operador = '-' THEN -1 * cs.ValorCota ELSE cs.ValorCota END
                 FROM CotaSensor cs 
                 WHERE cs.IdSensor = s.Id 
                   AND (cs.Fin = 0 OR cs.Fin IS NULL)
                 ORDER BY cs.FechaRegistro DESC), 0.0
            ) AS ValorCota
        FROM Sensor s
        INNER JOIN Estacion e ON s.IdEstacion = e.Id
        INNER JOIN TipoSensor ts ON s.IdTipoSensor = ts.Id
        WHERE s.Activo = 1 AND e.Activo = 1
        """
        try:
            cursor.execute(query)
        except Exception as e_new:
            # Fallback: si PuntoDecimal no existe en Sensor, usar vista legacy
            print(f"[SENSORES] Query directo falló ({e_new}), usando NV_SensoresSGD")
            cursor.execute("""
                SELECT IdAsignado, NumeroSensor, Sensor, PuntoDecimal,
                       PeriodoMuestra, ValorMinimo, ValorMaximo, ValorCota
                FROM [dbo].[NV_SensoresSGD]
            """)
        rows = cursor.fetchall()
        
        sensores_por_estacion = {}
        for row in rows:
            id_asignado = str(row.IdAsignado).strip()
            if id_asignado not in sensores_por_estacion:
                sensores_por_estacion[id_asignado] = []
            valor_cota = row.ValorCota if row.ValorCota is not None else 0.0
            sensores_por_estacion[id_asignado].append({
                "numero_sensor": str(row.NumeroSensor).zfill(4),
                "nombre": str(row.Sensor),
                "decimales": int(row.PuntoDecimal),
                "periodo": int(row.PeriodoMuestra),
                "valor_minimo": float(row.ValorMinimo),
                "valor_maximo": float(row.ValorMaximo),
                "valor_cota": float(valor_cota),
                "posiciones": []
            })
        
        conn.close()
        print(f"[SENSORES] Cargados {sum(len(v) for v in sensores_por_estacion.values())} sensores")
        return sensores_por_estacion

    except Exception as e:
        print(f"[ERROR] Falló carga de sensores: {e}")
        return {}


def revalidar_datos_invalidos(sensores_dict, estaciones_dict, dias=3):
    """Re-valida datos marcados como inválidos en PostgreSQL usando los límites
    actuales de SENSORES. Útil cuando se cambian ValorMinimo/ValorMaximo desde el web admin."""
    conn = get_pg_conn()
    try:
        cur = conn.cursor()
        total_corregidos = 0

        # Mapear id_asignado → dcp_id
        asignado_a_dcp = {}
        for dcp_id, datos in estaciones_dict.items():
            ia = datos.get("id_asignado", "")
            if ia:
                asignado_a_dcp[ia] = dcp_id

        for id_asignado, sensores in sensores_dict.items():
            dcp_id = asignado_a_dcp.get(id_asignado)
            if not dcp_id:
                continue
            for sensor in sensores:
                vmin = sensor.get("valor_minimo")
                vmax = sensor.get("valor_maximo")
                sid = sensor["numero_sensor"]
                if vmin is None or vmax is None:
                    continue
                # Corregir registros que están DENTRO de los límites actuales pero marcados inválidos
                cur.execute("""
                    UPDATE dcp_datos
                    SET valido = true, descripcion = 'OK (revalidado)'
                    WHERE dcp_id = %s AND sensor_id = %s
                      AND valido = false
                      AND valor >= %s AND valor <= %s
                      AND ts > NOW() - make_interval(days := %s)
                """, (dcp_id, sid, vmin, vmax, dias))
                n = cur.rowcount
                if n > 0:
                    total_corregidos += n
                    print(f"[REVALID] {dcp_id}/{sid} ({sensor['nombre']}): {n} registros corregidos "
                          f"(rango actual: {vmin} a {vmax})")

        conn.commit()
        cur.close()
        if total_corregidos > 0:
            print(f"[REVALID] Total corregidos: {total_corregidos}")
        else:
            print("[REVALID] Sin registros por corregir.")
    except Exception as e:
        print(f"[ERROR REVALID] {e}")
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        release_pg_conn(conn)


# ==================== CARGA UMBRALES DE PRECIPITACIÓN ====================
def load_umbrales_precipitacion():
    """Carga umbrales activos de precipitación desde SQL Server (UmbralAlertas).
    Retorna dict keyed por (id_asignado, numero_sensor) → lista de umbrales."""
    try:
        conn = get_sql_conn()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT u.Id, u.ValorReferencia, u.Umbral, u.Operador, u.Periodo,
                   u.Nombre, u.Color,
                   s.NumeroSensor, e.IdAsignado, e.Nombre as EstacionNombre
            FROM UmbralAlertas u
            JOIN Sensor s ON u.IdSensor = s.Id
            JOIN TipoSensor ts ON s.IdTipoSensor = ts.Id
            JOIN Estacion e ON s.IdEstacion = e.Id
            WHERE u.Activo = 1 AND ts.Nombre = N'Precipitación'
        """)
        umbrales = {}
        count = 0
        for row in cursor.fetchall():
            num_sensor = str(row.NumeroSensor).zfill(4)
            id_asignado = str(row.IdAsignado).strip()
            key = (id_asignado, num_sensor)
            if key not in umbrales:
                umbrales[key] = []
            umbrales[key].append({
                "umbral_id": int(row.Id),
                "valor_referencia": float(row.ValorReferencia),
                "umbral": float(row.Umbral),
                "operador": str(row.Operador).strip(),
                "periodo": int(row.Periodo) if row.Periodo else 60,
                "nombre": str(row.Nombre).strip(),
                "color": str(row.Color or "").strip(),
                "estacion_nombre": str(row.EstacionNombre).strip(),
            })
            count += 1
        conn.close()
        print(f"[UMBRALES] {count} umbrales de precipitación cargados ({len(umbrales)} sensores)")
        return umbrales
    except Exception as e:
        print(f"[ERROR] Falló carga de umbrales: {e}")
        return {}


def asignar_posiciones_sensores(sensores_lista, rango_transmision=60):
    """Asignar posiciones secuenciales por periodo
    
    Args:
        sensores_lista: Lista de sensores a procesar
        rango_transmision: Cantidad de minutos de datos que trae cada TX (default 60)
    """
    sensores_ordenados = sorted(sensores_lista, key=lambda s: s["numero_sensor"])
    
    posicion = 0
    for sensor in sensores_ordenados:
        periodo = sensor["periodo"]
        
        if periodo <= 0:
            print(f"[WARN] Periodo inválido ({periodo}) para sensor {sensor['numero_sensor']}. Ignorando.")
            continue
        
        if rango_transmision % periodo != 0:
            print(f"[WARN] Periodo {periodo} no divide rango {rango_transmision}. Ignorando {sensor['numero_sensor']}")
            continue
        
        valores_por_rango = int(rango_transmision // periodo)
        sensor["posiciones"] = list(range(posicion, posicion + valores_por_rango))
        posicion += valores_por_rango
        
        print(f"[POS] {sensor['numero_sensor']} | {sensor['nombre']} | Periodo: {periodo} min | Rango TX: {rango_transmision} min | Posiciones: {sensor['posiciones']}")
    
    return sensores_ordenados

def hora_concurrente(utc_now: datetime.datetime, periodo: int) -> datetime.datetime:
    """Alinear timestamp al múltiplo del periodo"""
    try:
        hora_actual = utc_now.hour
        min_actual = utc_now.minute
        
        if not isinstance(periodo, (int, float)) or periodo <= 0:
            raise ValueError(f"Periodo inválido: {periodo}")
        
        min_alineado = int((min_actual // periodo) * periodo)
        return datetime.datetime(utc_now.year, utc_now.month, utc_now.day, 
                                hora_actual, min_alineado, 0, tzinfo=datetime.timezone.utc)
    except Exception as e:
        print(f"[ERROR hora_concurrente] {e}")
        raise

# ==================== LLUVIA ACUMULADA (24h y periodo actual) ====================
def actualizar_lluvia_acumulada(dcp_id: str, id_asignado: str, sensores: list, conn):
    """
    Calcula precipitación acumulada en dos periodos:
    - '24h':    de 6:00 AM (UTC-6) ayer  a  6:00 AM (UTC-6) hoy
    - 'actual': de 6:00 AM (UTC-6) hoy   a  hora actual
    
    Horario de acumulación: minuto :01 de cada hora a minuto :00 de la siguiente
    (consistente con resumen_horario: ts - INTERVAL '1 minute')
    
    Solo procesa sensores cuyo nombre sea exactamente 'Precipitación'.
    """
    # Filtrar solo sensores de precipitación (exacto, no "precipitación acumulada")
    sensores_precip = [s for s in sensores 
                       if s["nombre"].strip().lower() == 'precipitación']
    if not sensores_precip:
        return
    
    try:
        cur = conn.cursor()
        
        # Crear tabla si no existe
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.lluvia_acumulada (
                id_asignado     CHARACTER VARYING(50) NOT NULL,
                dcp_id          CHARACTER VARYING(20) NOT NULL,
                sensor_id       CHARACTER VARYING(20) NOT NULL,
                variable        CHARACTER VARYING(50),
                periodo_inicio  TIMESTAMP WITH TIME ZONE NOT NULL,
                periodo_fin     TIMESTAMP WITH TIME ZONE NOT NULL,
                tipo_periodo    CHARACTER VARYING(10) NOT NULL,
                acumulado       REAL DEFAULT 0,
                horas_con_dato  INTEGER DEFAULT 0,
                ultima_actualizacion TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
                CONSTRAINT lluvia_acumulada_pkey 
                    PRIMARY KEY (id_asignado, sensor_id, tipo_periodo)
            );
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_lluvia_tipo_periodo 
            ON public.lluvia_acumulada (tipo_periodo);
        """)
        conn.commit()
        
        ahora_utc = datetime.datetime.now(datetime.timezone.utc)
        
        # Hora 6:00 AM tiempo local (UTC-6) = 12:00 PM UTC
        hoy_6am_utc = ahora_utc.replace(hour=12, minute=0, second=0, microsecond=0)
        if ahora_utc < hoy_6am_utc:
            # Antes de las 6am local: el "hoy 6am" es ayer a las 12 UTC
            hoy_6am_utc -= datetime.timedelta(days=1)
        
        ayer_6am_utc = hoy_6am_utc - datetime.timedelta(days=1)
        
        for sensor in sensores_precip:
            variable = sensor["nombre"].lower().replace(" ", "_")
            
            # ---- Periodo 24h (ayer 6am a hoy 6am) ----
            cur.execute("""
                SELECT COALESCE(SUM(acumulado), 0), COUNT(*)
                FROM resumen_horario
                WHERE id_asignado = %s AND sensor_id = %s
                  AND ts >= %s AND ts < %s
            """, (id_asignado, sensor["numero_sensor"], ayer_6am_utc, hoy_6am_utc))
            row_24h = cur.fetchone()
            acum_24h = row_24h[0] if row_24h else 0
            horas_24h = row_24h[1] if row_24h else 0
            
            cur.execute("""
                INSERT INTO lluvia_acumulada 
                (id_asignado, dcp_id, sensor_id, variable, 
                 periodo_inicio, periodo_fin, tipo_periodo,
                 acumulado, horas_con_dato, ultima_actualizacion)
                VALUES (%s, %s, %s, %s, %s, %s, '24h', %s, %s, NOW())
                ON CONFLICT (id_asignado, sensor_id, tipo_periodo) DO UPDATE SET
                    dcp_id = EXCLUDED.dcp_id,
                    variable = EXCLUDED.variable,
                    periodo_inicio = EXCLUDED.periodo_inicio,
                    periodo_fin = EXCLUDED.periodo_fin,
                    acumulado = EXCLUDED.acumulado,
                    horas_con_dato = EXCLUDED.horas_con_dato,
                    ultima_actualizacion = NOW();
            """, (id_asignado, dcp_id, sensor["numero_sensor"], variable,
                  ayer_6am_utc, hoy_6am_utc, acum_24h, horas_24h))
            
            # ---- Periodo actual (hoy 6am a ahora) ----
            cur.execute("""
                SELECT COALESCE(SUM(acumulado), 0), COUNT(*)
                FROM resumen_horario
                WHERE id_asignado = %s AND sensor_id = %s
                  AND ts >= %s AND ts < %s
            """, (id_asignado, sensor["numero_sensor"], hoy_6am_utc, ahora_utc))
            row_act = cur.fetchone()
            acum_actual = row_act[0] if row_act else 0
            horas_actual = row_act[1] if row_act else 0
            
            cur.execute("""
                INSERT INTO lluvia_acumulada 
                (id_asignado, dcp_id, sensor_id, variable, 
                 periodo_inicio, periodo_fin, tipo_periodo,
                 acumulado, horas_con_dato, ultima_actualizacion)
                VALUES (%s, %s, %s, %s, %s, %s, 'actual', %s, %s, NOW())
                ON CONFLICT (id_asignado, sensor_id, tipo_periodo) DO UPDATE SET
                    dcp_id = EXCLUDED.dcp_id,
                    variable = EXCLUDED.variable,
                    periodo_inicio = EXCLUDED.periodo_inicio,
                    periodo_fin = EXCLUDED.periodo_fin,
                    acumulado = EXCLUDED.acumulado,
                    horas_con_dato = EXCLUDED.horas_con_dato,
                    ultima_actualizacion = NOW();
            """, (id_asignado, dcp_id, sensor["numero_sensor"], variable,
                  hoy_6am_utc, ahora_utc, acum_actual, horas_actual))
        
        print(f"[LLUVIA] {id_asignado} | 24h: {acum_24h:.1f}mm ({horas_24h}h) | Actual: {acum_actual:.1f}mm ({horas_actual}h)")
        cur.close()
        
    except Exception as e:
        print(f"[ERROR LLUVIA] {id_asignado}: {e}")
        try:
            conn.rollback()
        except Exception:
            pass


# ==================== ALERTAS DE PRECIPITACIÓN ====================
def evaluar_alertas_precipitacion(dcp_id: str, id_asignado: str, sensores: list, conn):
    """
    Evalúa si la precipitación acumulada en el periodo del umbral supera el valor de referencia.
    Lee umbrales de UMBRALES_PRECIP (cargados desde SQL Server).
    Escribe alertas en alertas_precipitacion (TimescaleDB).
    """
    # Filtrar sensores de precipitación
    sensores_precip = [s for s in sensores
                       if s["nombre"].strip().lower() == 'precipitación']
    if not sensores_precip:
        return

    try:
        cur = conn.cursor()

        # Crear tabla si no existe
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.alertas_precipitacion (
                ts                  TIMESTAMPTZ NOT NULL,
                id_asignado         VARCHAR(50) NOT NULL,
                dcp_id              VARCHAR(20) NOT NULL,
                sensor_id           VARCHAR(20) NOT NULL,
                estacion_nombre     VARCHAR(200),
                umbral_id           BIGINT NOT NULL,
                umbral_nombre       VARCHAR(100),
                valor_referencia    REAL,
                valor_medido        REAL,
                operador            VARCHAR(2),
                periodo_minutos     INT,
                color               VARCHAR(50),
                activa              BOOLEAN DEFAULT TRUE,
                notificada          BOOLEAN DEFAULT FALSE,
                CONSTRAINT alertas_precipitacion_pkey
                    PRIMARY KEY (ts, dcp_id, sensor_id, umbral_id)
            );
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_alertas_precip_estacion
            ON public.alertas_precipitacion (id_asignado, ts DESC);
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_alertas_precip_activa
            ON public.alertas_precipitacion (activa, notificada);
        """)
        conn.commit()

        ahora_utc = datetime.datetime.now(datetime.timezone.utc)

        for sensor in sensores_precip:
            key = (id_asignado, sensor["numero_sensor"])
            umbrales = UMBRALES_PRECIP.get(key, [])
            if not umbrales:
                continue

            for umb in umbrales:
                periodo = umb["periodo"]  # minutos
                desde = ahora_utc - datetime.timedelta(minutes=periodo)

                # Sumar precipitación en el periodo usando resumen_horario si periodo >= 60,
                # o dcp_datos si periodo < 60
                if periodo >= 60:
                    cur.execute("""
                        SELECT COALESCE(SUM(acumulado), 0)
                        FROM resumen_horario
                        WHERE id_asignado = %s AND sensor_id = %s
                          AND ts >= %s AND ts <= %s
                    """, (id_asignado, sensor["numero_sensor"], desde, ahora_utc))
                else:
                    cur.execute("""
                        SELECT COALESCE(SUM(valor), 0)
                        FROM dcp_datos
                        WHERE id_asignado = %s AND sensor_id = %s
                          AND ts >= %s AND ts <= %s
                          AND valido = true
                    """, (id_asignado, sensor["numero_sensor"], desde, ahora_utc))

                row = cur.fetchone()
                valor_medido = float(row[0]) if row else 0.0

                # Evaluar condición según operador
                limite = umb["valor_referencia"] + umb["umbral"]
                operador = umb["operador"]
                alerta = False

                if operador == '+' and valor_medido >= limite:
                    alerta = True
                elif operador == '-' and valor_medido <= (umb["valor_referencia"] - umb["umbral"]):
                    alerta = True

                if alerta:
                    # Insertar alerta (UPSERT para no duplicar en la misma hora)
                    ts_alerta = ahora_utc.replace(minute=0, second=0, microsecond=0)
                    cur.execute("""
                        INSERT INTO alertas_precipitacion
                        (ts, id_asignado, dcp_id, sensor_id, estacion_nombre,
                         umbral_id, umbral_nombre, valor_referencia, valor_medido,
                         operador, periodo_minutos, color, activa, notificada)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE, FALSE)
                        ON CONFLICT (ts, dcp_id, sensor_id, umbral_id) DO UPDATE SET
                            valor_medido = EXCLUDED.valor_medido,
                            activa = TRUE;
                    """, (ts_alerta, id_asignado, dcp_id, sensor["numero_sensor"],
                          umb["estacion_nombre"], umb["umbral_id"], umb["nombre"],
                          umb["valor_referencia"], valor_medido, operador,
                          periodo, umb["color"]))

                    print(f"[ALERTA] {id_asignado} | {umb['estacion_nombre']} | "
                          f"{umb['nombre']} | Medido: {valor_medido:.1f}mm >= {limite:.1f}mm")

        cur.close()

    except Exception as e:
        print(f"[ERROR ALERTAS] {id_asignado}: {e}")
        try:
            conn.rollback()
        except Exception:
            pass


# ==================== EVENTOS DE LLUVIA ====================
def _detectar_sospechoso(valores_batch, periodo, acumulado_evento, duracion_evento_min):
    """
    Validación exhaustiva multi-criterio para detectar datos de precipitación sospechosos.
    Retorna (es_sospechoso: bool, motivos: list[str]).

    Criterios:
    1. Lectura individual extrema (>10mm en un periodo)
    2. Intensidad instantánea absurda (>100 mm/h)
    3. Pico abrupto: lectura >20× la mediana del lote
    4. Sensor pegado / ruido constante: >70% lecturas no-cero con misma resolución mínima
    5. Evento prolongado sin pausas naturales (>60 min, 0 ceros, promedio <0.3mm)
    6. Acumulado alto por goteo: promedio <0.25mm/lectura pero total >5mm
    """
    motivos = []
    if not valores_batch:
        return False, motivos

    vals = [float(v) for v in valores_batch]
    no_cero = [v for v in vals if v > 0]

    # --- 1. Lectura individual extrema ---
    for v in vals:
        if v >= 10.0:
            motivos.append(f"Lectura extrema {v:.1f}mm en {periodo}min (umbral 10mm)")

    # --- 2. Intensidad instantánea absurda ---
    for v in vals:
        intensidad = v * (60.0 / periodo)
        if intensidad >= 100.0:
            motivos.append(f"Intensidad {intensidad:.1f}mm/h (umbral 100mm/h)")

    # --- 3. Pico abrupto respecto al lote ---
    if len(no_cero) >= 3:
        sorted_nz = sorted(no_cero)
        mediana = sorted_nz[len(sorted_nz) // 2]
        maximo = max(no_cero)
        if mediana > 0 and maximo >= 20.0 * mediana:
            motivos.append(f"Pico abrupto: max {maximo:.1f}mm vs mediana {mediana:.2f}mm "
                           f"(ratio {maximo/mediana:.0f}×)")

    # --- 4. Sensor pegado / ruido constante ---
    if len(vals) >= 5:
        pct_no_cero = len(no_cero) / len(vals)
        if pct_no_cero >= 0.70 and no_cero:
            # Verificar si la mayoría de lecturas son el mismo valor (resolución mínima)
            from collections import Counter
            conteo = Counter(round(v, 2) for v in no_cero)
            val_freq, freq = conteo.most_common(1)[0]
            if freq / len(no_cero) >= 0.50 and val_freq <= 0.3:
                motivos.append(f"Sensor pegado/ruido: {freq}/{len(no_cero)} lecturas "
                               f"en {val_freq}mm ({pct_no_cero*100:.0f}% no-cero)")

    # --- 5. Evento prolongado sin pausas naturales ---
    if duracion_evento_min >= 60:
        ceros_en_batch = sum(1 for v in vals if v == 0)
        if ceros_en_batch == 0 and no_cero:
            promedio = sum(no_cero) / len(no_cero)
            if promedio < 0.3:
                motivos.append(f"Sin pausas >60min con promedio {promedio:.2f}mm/lectura "
                               f"(patrón de ruido)")

    # --- 6. Acumulado alto por goteo constante ---
    if no_cero and acumulado_evento >= 5.0:
        total_lecturas_estimadas = max(duracion_evento_min / periodo, len(vals))
        promedio_por_lectura = acumulado_evento / total_lecturas_estimadas if total_lecturas_estimadas > 0 else 0
        if promedio_por_lectura < 0.25 and promedio_por_lectura > 0:
            motivos.append(f"Acumulado {acumulado_evento:.1f}mm por goteo "
                           f"(promedio {promedio_por_lectura:.2f}mm/lectura, patrón de sensor con fuga)")

    # Deduplicar motivos similares
    motivos_unicos = list(dict.fromkeys(motivos))
    return len(motivos_unicos) > 0, motivos_unicos


def actualizar_eventos_lluvia(dcp_id: str, id_asignado: str, sensores: list, conn,
                              timestamp_goes: datetime.datetime, rango_tx: int):
    """
    Detecta y sigue eventos de lluvia continua por estación/sensor.
    Adaptado para estaciones GOES que transmiten lotes cada 1-2 horas.

    Reglas clave:
    - INICIO: Solo si el lote tiene >=2 lecturas con valor>0, o una lectura >=0.5mm.
              Esto filtra lecturas aisladas de ruido/rocío (0.1mm puntual).
    - FIN: Requiere ceros_consecutivos >= (rango_tx / periodo), es decir, un ciclo
           completo de transmisión sin lluvia (ej: 60min/10min = 6 ceros = 1 hora seca).
    - HUÉRFANOS: Si un evento activo lleva >4 horas sin recibir nuevos datos, se cierra
                 automáticamente.
    - Validación multi-criterio para marcar datos sospechosos.
    """
    sensores_precip = [s for s in sensores
                       if s["nombre"].strip().lower() == 'precipitación']
    if not sensores_precip:
        return

    try:
        cur = conn.cursor()

        # Crear tabla si no existe
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.eventos_lluvia (
                id                   BIGSERIAL PRIMARY KEY,
                id_asignado          VARCHAR(50) NOT NULL,
                dcp_id               VARCHAR(20) NOT NULL,
                sensor_id            VARCHAR(20) NOT NULL,
                estacion_nombre      VARCHAR(200),
                inicio               TIMESTAMPTZ NOT NULL,
                fin                  TIMESTAMPTZ,
                acumulado_mm         REAL DEFAULT 0,
                intensidad_max_mmh   REAL DEFAULT 0,
                duracion_minutos     INT DEFAULT 0,
                estado               VARCHAR(20) DEFAULT 'activo',
                sospechoso           BOOLEAN DEFAULT FALSE,
                motivo_sospecha      TEXT,
                ceros_consecutivos   INT DEFAULT 0,
                ultimo_ts_procesado  TIMESTAMPTZ,
                ultima_actualizacion TIMESTAMPTZ DEFAULT NOW()
            );
        """)
        # Migrar tabla existente: agregar columnas si faltan
        cur.execute("""
            DO $$ BEGIN
                ALTER TABLE public.eventos_lluvia ADD COLUMN IF NOT EXISTS sospechoso BOOLEAN DEFAULT FALSE;
                ALTER TABLE public.eventos_lluvia ADD COLUMN IF NOT EXISTS motivo_sospecha TEXT;
            END $$;
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_eventos_lluvia_activo
            ON public.eventos_lluvia (id_asignado, sensor_id, estado);
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_eventos_lluvia_inicio
            ON public.eventos_lluvia (inicio DESC);
        """)
        conn.commit()

        # Ceros necesarios para cerrar: al menos un ciclo completo de TX sin lluvia
        # Mínimo 6 (1 hora con periodo=10min), máximo 12 (2 horas)
        nombre_est = ESTACIONES.get(dcp_id, {}).get("nombre", "")

        for sensor in sensores_precip:
            periodo = sensor["periodo"]  # típicamente 10 min
            sensor_id = sensor["numero_sensor"]
            ceros_para_cerrar = max(6, min(12, rango_tx // max(periodo, 1)))

            # Buscar evento activo para este sensor
            cur.execute("""
                SELECT id, acumulado_mm, intensidad_max_mmh, ceros_consecutivos,
                       inicio, ultimo_ts_procesado
                FROM eventos_lluvia
                WHERE id_asignado = %s AND sensor_id = %s AND estado = 'activo'
                ORDER BY inicio DESC LIMIT 1
            """, (id_asignado, sensor_id))
            evento = cur.fetchone()

            # --- Cerrar eventos huérfanos (>4 horas sin datos nuevos) ---
            if evento and evento[5]:
                horas_sin_datos = (timestamp_goes - evento[5]).total_seconds() / 3600
                if horas_sin_datos > 4:
                    dur = int((evento[5] - evento[4]).total_seconds() / 60)
                    cur.execute("""
                        UPDATE eventos_lluvia SET
                            fin = %s, duracion_minutos = %s, estado = 'finalizado',
                            ultima_actualizacion = NOW()
                        WHERE id = %s
                    """, (evento[5], dur, evento[0]))
                    print(f"[EVENTO] 🔒 CERRADO huérfano {id_asignado} | {nombre_est} | "
                          f"Sin datos por {horas_sin_datos:.1f}h, dur={dur}min")
                    evento = None  # ya no hay evento activo

            # Determinar desde dónde leer datos nuevos
            if evento and evento[5]:  # tiene ultimo_ts_procesado
                desde = evento[5]
            else:
                desde = timestamp_goes - datetime.timedelta(minutes=rango_tx)

            # Obtener datos nuevos cronológicamente
            cur.execute("""
                SELECT ts, valor FROM dcp_datos
                WHERE id_asignado = %s AND sensor_id = %s
                  AND ts > %s AND ts <= %s
                  AND valido = true
                ORDER BY ts ASC
            """, (id_asignado, sensor_id, desde, timestamp_goes))
            datos = cur.fetchall()
            if not datos:
                continue

            # Estado actual del evento
            evento_id = evento[0] if evento else None
            acumulado = float(evento[1]) if evento else 0.0
            intensidad_max = float(evento[2]) if evento else 0.0
            ceros = int(evento[3]) if evento else 0
            inicio = evento[4] if evento else None

            # Pre-análisis del lote: cuántas lecturas positivas hay y el total
            lecturas_positivas = [float(v) for _, v in datos if float(v) > 0]
            num_positivas = len(lecturas_positivas)
            total_lote = sum(lecturas_positivas)

            # Colectar valores del lote para análisis posterior
            valores_batch = []

            for ts_dato, valor in datos:
                valor = float(valor)
                valores_batch.append(valor)

                if valor > 0:
                    intensidad = valor * (60.0 / periodo)  # mm/h
                    ceros = 0

                    if evento_id is None:
                        # Filtro anti-ruido: solo crear evento si el lote tiene
                        # >=2 lecturas positivas O una lectura >= 0.5mm
                        if num_positivas >= 2 or valor >= 0.5:
                            # NUEVO EVENTO
                            cur.execute("""
                                INSERT INTO eventos_lluvia
                                (id_asignado, dcp_id, sensor_id, estacion_nombre,
                                 inicio, acumulado_mm, intensidad_max_mmh,
                                 estado, ceros_consecutivos, ultimo_ts_procesado)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, 'activo', 0, %s)
                                RETURNING id
                            """, (id_asignado, dcp_id, sensor_id, nombre_est,
                                  ts_dato, valor, intensidad, ts_dato))
                            evento_id = cur.fetchone()[0]
                            acumulado = valor
                            intensidad_max = intensidad
                            inicio = ts_dato
                            print(f"[EVENTO] 🌧️ INICIO lluvia {id_asignado} | "
                                  f"{nombre_est} | {ts_dato} | "
                                  f"{num_positivas} lecturas+ en lote, total={total_lote:.1f}mm")
                        # else: lectura aislada insignificante, ignorar
                    else:
                        # CONTINÚA: acumular
                        acumulado += valor
                        intensidad_max = max(intensidad_max, intensidad)
                else:
                    # valor == 0
                    if evento_id is not None:
                        ceros += 1

            # Actualizar estado final después de procesar todos los datos
            ultimo_ts = datos[-1][0]

            if evento_id is not None:
                duracion = int((ultimo_ts - inicio).total_seconds() / 60)

                # === VALIDACIÓN EXHAUSTIVA MULTI-CRITERIO ===
                sospechoso, motivos = _detectar_sospechoso(
                    valores_batch, periodo, acumulado, duracion)

                motivo_txt = '; '.join(motivos) if motivos else None
                marca_sosp = '⚠️  SOSPECHOSO ' if sospechoso else ''

                if sospechoso:
                    for m in motivos:
                        print(f"[EVENTO] ⚠️  SOSPECHOSO {id_asignado} | {nombre_est} | {m}")

                if ceros >= ceros_para_cerrar:
                    # CERRAR EVENTO: suficientes ceros (1 ciclo TX completo sin lluvia)
                    cur.execute("""
                        UPDATE eventos_lluvia SET
                            fin = %s, acumulado_mm = %s, intensidad_max_mmh = %s,
                            duracion_minutos = %s, estado = 'finalizado',
                            sospechoso = sospechoso OR %s,
                            motivo_sospecha = COALESCE(motivo_sospecha, %s),
                            ceros_consecutivos = %s, ultimo_ts_procesado = %s,
                            ultima_actualizacion = NOW()
                        WHERE id = %s
                    """, (ultimo_ts, acumulado, intensidad_max, duracion,
                          sospechoso, motivo_txt,
                          ceros, ultimo_ts, evento_id))
                    print(f"[EVENTO] ✅ {marca_sosp}FIN lluvia {id_asignado} | {nombre_est} | "
                          f"Acum: {acumulado:.1f}mm | Duración: {duracion}min | "
                          f"Int.max: {intensidad_max:.1f}mm/h | ceros={ceros}/{ceros_para_cerrar}")
                else:
                    # ACTUALIZAR evento activo
                    cur.execute("""
                        UPDATE eventos_lluvia SET
                            acumulado_mm = %s, intensidad_max_mmh = %s,
                            duracion_minutos = %s,
                            sospechoso = sospechoso OR %s,
                            motivo_sospecha = COALESCE(motivo_sospecha, %s),
                            ceros_consecutivos = %s,
                            ultimo_ts_procesado = %s, ultima_actualizacion = NOW()
                        WHERE id = %s
                    """, (acumulado, intensidad_max, duracion,
                          sospechoso, motivo_txt,
                          ceros, ultimo_ts, evento_id))
                    if acumulado > 0:
                        print(f"[EVENTO] 🌧️ {marca_sosp}ACTIVO {id_asignado} | {nombre_est} | "
                              f"Acum: {acumulado:.1f}mm | {duracion}min | "
                              f"Int.max: {intensidad_max:.1f}mm/h | ceros={ceros}/{ceros_para_cerrar}")

        cur.close()

    except Exception as e:
        print(f"[ERROR EVENTOS] {id_asignado}: {e}")
        try:
            conn.rollback()
        except Exception:
            pass


# ==================== PRECIPITACIÓN PROMEDIO POR CUENCA ====================
def actualizar_precipitacion_cuenca(conn):
    """Calcula precipitación promedio de la última hora por cuenca y subcuenca.
    Semaforiza según intensidad: verde < 2.5, amarillo < 7.5, naranja < 15, rojo >= 15 mm/h."""
    if not CUENCAS_ESTACION:
        return

    try:
        cur = conn.cursor()

        # Crear tabla si no existe
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.precipitacion_cuenca (
                ts                  TIMESTAMPTZ NOT NULL,
                tipo                VARCHAR(10) NOT NULL,
                nombre              VARCHAR(200) NOT NULL,
                promedio_mm         REAL DEFAULT 0,
                max_mm              REAL DEFAULT 0,
                min_mm              REAL DEFAULT 0,
                estaciones_con_dato INTEGER DEFAULT 0,
                estaciones_total    INTEGER DEFAULT 0,
                semaforo            VARCHAR(10),
                ultima_actualizacion TIMESTAMPTZ DEFAULT NOW(),
                CONSTRAINT precipitacion_cuenca_pkey
                    PRIMARY KEY (ts, tipo, nombre)
            );
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_precip_cuenca_ts
            ON public.precipitacion_cuenca (ts DESC, tipo);
        """)
        conn.commit()

        ahora_utc = datetime.datetime.now(datetime.timezone.utc)
        ts_hora = ahora_utc.replace(minute=0, second=0, microsecond=0)
        desde = ts_hora - datetime.timedelta(hours=1)

        # Obtener precipitación acumulada de la última hora por estación
        # (resumen_horario.ts = inicio de la hora, acumulado = lluvia de esa hora)
        cur.execute("""
            SELECT id_asignado, COALESCE(SUM(acumulado), 0) AS precip_mm
            FROM resumen_horario
            WHERE variable = 'precipitación'
              AND ts >= %s AND ts < %s
            GROUP BY id_asignado
        """, (desde, ts_hora))
        datos_estacion = {row[0]: float(row[1]) for row in cur.fetchall()}

        # Agrupar por cuenca y subcuenca
        cuencas_agg = {}   # nombre_cuenca → {valores:[], total:int}
        sub_agg = {}       # nombre_subcuenca → {valores:[], total:int}

        # Contar estaciones totales por cuenca/subcuenca
        for id_asignado, info in CUENCAS_ESTACION.items():
            cuenca = info["cuenca"]
            subcuenca = info["subcuenca"]
            if cuenca:
                if cuenca not in cuencas_agg:
                    cuencas_agg[cuenca] = {"valores": [], "total": 0}
                cuencas_agg[cuenca]["total"] += 1
            if subcuenca:
                if subcuenca not in sub_agg:
                    sub_agg[subcuenca] = {"valores": [], "total": 0}
                sub_agg[subcuenca]["total"] += 1

        # Asignar valores de precipitación
        for id_asignado, precip in datos_estacion.items():
            info = CUENCAS_ESTACION.get(id_asignado)
            if not info:
                continue
            if info["cuenca"] and info["cuenca"] in cuencas_agg:
                cuencas_agg[info["cuenca"]]["valores"].append(precip)
            if info["subcuenca"] and info["subcuenca"] in sub_agg:
                sub_agg[info["subcuenca"]]["valores"].append(precip)

        def semaforo(promedio):
            if promedio < 2.5:
                return "verde"
            elif promedio < 7.5:
                return "amarillo"
            elif promedio < 15.0:
                return "naranja"
            else:
                return "rojo"

        # Insertar/actualizar por cuenca y subcuenca
        registros = []
        for nombre, agg in cuencas_agg.items():
            vals = agg["valores"]
            prom = sum(vals) / len(vals) if vals else 0.0
            registros.append((ts_hora, "cuenca", nombre, prom,
                              max(vals) if vals else 0.0,
                              min(vals) if vals else 0.0,
                              len(vals), agg["total"], semaforo(prom)))

        for nombre, agg in sub_agg.items():
            vals = agg["valores"]
            prom = sum(vals) / len(vals) if vals else 0.0
            registros.append((ts_hora, "subcuenca", nombre, prom,
                              max(vals) if vals else 0.0,
                              min(vals) if vals else 0.0,
                              len(vals), agg["total"], semaforo(prom)))

        for reg in registros:
            cur.execute("""
                INSERT INTO precipitacion_cuenca
                (ts, tipo, nombre, promedio_mm, max_mm, min_mm,
                 estaciones_con_dato, estaciones_total, semaforo, ultima_actualizacion)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (ts, tipo, nombre) DO UPDATE SET
                    promedio_mm = EXCLUDED.promedio_mm,
                    max_mm = EXCLUDED.max_mm,
                    min_mm = EXCLUDED.min_mm,
                    estaciones_con_dato = EXCLUDED.estaciones_con_dato,
                    estaciones_total = EXCLUDED.estaciones_total,
                    semaforo = EXCLUDED.semaforo,
                    ultima_actualizacion = NOW();
            """, reg)

        # Log resumen
        cuencas_con_lluvia = [(n, a) for n, a in cuencas_agg.items() if a["valores"] and sum(a["valores"]) > 0]
        for nombre, agg in sorted(cuencas_con_lluvia, key=lambda x: -(sum(x[1]["valores"])/len(x[1]["valores"]))):
            vals = agg["valores"]
            prom = sum(vals) / len(vals)
            print(f"[CUENCA] {nombre} | Prom: {prom:.1f}mm/h | Max: {max(vals):.1f}mm | "
                  f"{len(vals)}/{agg['total']} estaciones | Semáforo: {semaforo(prom).upper()}")

        sub_con_lluvia = [(n, a) for n, a in sub_agg.items() if a["valores"] and sum(a["valores"]) > 0]
        for nombre, agg in sorted(sub_con_lluvia, key=lambda x: -(sum(x[1]["valores"])/len(x[1]["valores"]))):
            vals = agg["valores"]
            prom = sum(vals) / len(vals)
            print(f"[SUBCUENCA] {nombre} | Prom: {prom:.1f}mm/h | Max: {max(vals):.1f}mm | "
                  f"{len(vals)}/{agg['total']} estaciones | Semáforo: {semaforo(prom).upper()}")

        if not cuencas_con_lluvia and not sub_con_lluvia:
            print(f"[CUENCA] Sin precipitación en la última hora ({desde} → {ts_hora})")

        cur.close()

    except Exception as e:
        print(f"[ERROR CUENCA] {e}")
        try:
            conn.rollback()
        except Exception:
            pass


# ==================== INSERCIÓN OPTIMIZADA TIMESCALEDB ====================
def insertar_datos_mapeados(dcp_id: str, id_asignado: str, timestamp_goes: datetime.datetime, valores: list):
    print(f"[DEBUG] timestamp_goes (antes de insertar): {timestamp_goes} (tzinfo={getattr(timestamp_goes, 'tzinfo', None)})")
    """
    VERSIÓN OPTIMIZADA PARA TIMESCALEDB
    - Usa execute_values (batch insert 10x más rápido)
    - Inserta en hypertable con compresión automática
    - Actualiza resúmenes con UPSERT
    """
    sensores = SENSORES.get(id_asignado)
    if not sensores:
        return
    
    conn = get_pg_conn()
    try:
        cur = conn.cursor()
    
        # ==================== PREPARAR BATCH DE DATOS CRUDOS ====================
        registros_dict = {}
        for sensor in sensores:
            for i, pos in enumerate(sensor["posiciones"]):
                if pos >= len(valores):
                    continue
                valor_raw = valores[pos]
                valor = round(valor_raw / (10 ** sensor["decimales"]), sensor["decimales"])
                # valor += sensor.get("valor_cota", 0.0)
                ts = timestamp_goes - datetime.timedelta(minutes=i * sensor["periodo"])
                ts = hora_concurrente(ts, sensor["periodo"])
                valido = True
                descripcion = "OK"
                if "valor_minimo" in sensor and valor < sensor["valor_minimo"]:
                    valido = False
                    descripcion = f"Bajo: {valor} < {sensor['valor_minimo']}"
                elif "valor_maximo" in sensor and valor > sensor["valor_maximo"]:
                    valido = False
                    descripcion = f"Alto: {valor} > {sensor['valor_maximo']}"
                key = (ts, dcp_id, sensor["numero_sensor"])
                registros_dict[key] = (
                    ts,
                    dcp_id,
                    id_asignado,
                    sensor["numero_sensor"],
                    sensor["nombre"].lower().replace(" ", "_"),
                    valor,
                    sensor.get("tipo", sensor["nombre"]).lower(),
                    valido,
                    descripcion[:200],
                    ts.year,
                    ts.month,
                    ts.day,
                    ts.hour,
                    ts.minute
                )
        registros_batch = list(registros_dict.values())
        
        # ==================== INSERCIÓN MASIVA (10x más rápida) ====================
        if registros_batch:
            execute_values(cur, """
                INSERT INTO dcp_datos 
                (ts, dcp_id, id_asignado, sensor_id, variable, valor, tipo, 
                 valido, descripcion, anio, mes, dia, hora, minuto)
                VALUES %s
                ON CONFLICT (ts, dcp_id, sensor_id) DO UPDATE SET
                    id_asignado = EXCLUDED.id_asignado,
                    variable = EXCLUDED.variable,
                    valor = EXCLUDED.valor,
                    tipo = EXCLUDED.tipo,
                    valido = EXCLUDED.valido,
                    descripcion = EXCLUDED.descripcion,
                    anio = EXCLUDED.anio,
                    mes = EXCLUDED.mes,
                    dia = EXCLUDED.dia,
                    hora = EXCLUDED.hora,
                    minuto = EXCLUDED.minuto
            """, registros_batch, page_size=500)
            
            conn.commit()
            print(f"[PG] ✅ {len(registros_batch)} registros insertados en hypertable")
        
        # ==================== RESÚMENES HORARIOS ====================
        estacion_config = ESTACIONES.get(dcp_id, {})
        rango_tx = estacion_config.get("rango_transmision", 60)
        
        ts_inicio = timestamp_goes - datetime.timedelta(minutes=rango_tx)
        ts_fin = timestamp_goes
        
        ts_inicio = hora_concurrente(ts_inicio, 60)

        for sensor in sensores:
            try:
                cur.execute("""
                    INSERT INTO resumen_horario 
                    (ts, dcp_id, sensor_id, id_asignado, variable, tipo,
                     suma, conteo, promedio, minimo, maximo, acumulado)
                    SELECT 
                        date_trunc('hour', ts - INTERVAL '1 minute') AS hora,
                        dcp_id,
                        sensor_id,
                        id_asignado,
                        variable,
                        tipo,
                        SUM(valor),
                        COUNT(*),
                        AVG(valor),
                        MIN(valor),
                        MAX(valor),
                        SUM(valor)
                    FROM dcp_datos
                    WHERE id_asignado = %s AND sensor_id = %s
                      AND ts > %s AND ts <= %s
                      AND valido = true
                    GROUP BY date_trunc('hour', ts - INTERVAL '1 minute'), dcp_id, sensor_id, id_asignado, variable, tipo
                    ON CONFLICT (ts, dcp_id, sensor_id) DO UPDATE SET
                        suma = EXCLUDED.suma,
                        conteo = EXCLUDED.conteo,
                        promedio = EXCLUDED.promedio,
                        minimo = EXCLUDED.minimo,
                        maximo = EXCLUDED.maximo,
                        acumulado = EXCLUDED.acumulado;
                """, (id_asignado, sensor["numero_sensor"], ts_inicio, ts_fin))
                
            except Exception as e:
                print(f"[WARN] Resumen horario falló: {e}")
        
        # ==================== RESÚMENES DIARIOS ====================
        fecha_inicio = (ts_inicio - datetime.timedelta(minutes=1)).date()
        fecha_fin = (ts_fin - datetime.timedelta(minutes=1)).date()
        
        fechas_afectadas = set()
        fecha_cursor = fecha_inicio
        while fecha_cursor <= fecha_fin:
            fechas_afectadas.add(fecha_cursor)
            fecha_cursor += datetime.timedelta(days=1)
        
        print(f"[DIARIO] Fechas a recalcular: {fechas_afectadas}")
        
        for fecha_resumen in fechas_afectadas:
            for sensor in sensores:
                try:
                    cur.execute("""
                        INSERT INTO resumen_diario 
                        (fecha, dcp_id, sensor_id, id_asignado, variable, tipo,
                         suma, conteo, promedio, minimo, maximo, acumulado)
                        SELECT 
                            DATE(ts - INTERVAL '1 minute'),
                            dcp_id,
                            sensor_id,
                            id_asignado,
                            variable,
                            tipo,
                            SUM(valor),
                            COUNT(*),
                            AVG(valor),
                            MIN(valor),
                            MAX(valor),
                            SUM(valor)
                        FROM dcp_datos
                        WHERE id_asignado = %s AND sensor_id = %s
                          AND DATE(ts - INTERVAL '1 minute') = %s
                          AND valido = true
                        GROUP BY DATE(ts - INTERVAL '1 minute'), dcp_id, sensor_id, id_asignado, variable, tipo
                        ON CONFLICT (fecha, dcp_id, sensor_id) DO UPDATE SET
                            suma = EXCLUDED.suma,
                            conteo = EXCLUDED.conteo,
                            promedio = EXCLUDED.promedio,
                            minimo = EXCLUDED.minimo,
                            maximo = EXCLUDED.maximo,
                            acumulado = EXCLUDED.acumulado;
                    """, (id_asignado, sensor["numero_sensor"], fecha_resumen))
                    
                except Exception as e:
                    print(f"[WARN] Resumen diario falló para {fecha_resumen}: {e}")
        
        conn.commit()
        print("[PG] ✅ Resúmenes horario y diario actualizados")
        
        # ==================== LLUVIA ACUMULADA ====================
        actualizar_lluvia_acumulada(dcp_id, id_asignado, sensores, conn)
        conn.commit()

        # ==================== ALERTAS DE PRECIPITACIÓN ====================
        evaluar_alertas_precipitacion(dcp_id, id_asignado, sensores, conn)
        conn.commit()

        # ==================== EVENTOS DE LLUVIA ====================
        actualizar_eventos_lluvia(dcp_id, id_asignado, sensores, conn,
                                  timestamp_goes, rango_tx)
        conn.commit()

        # ==================== PRECIPITACIÓN POR CUENCA ====================
        actualizar_precipitacion_cuenca(conn)
        conn.commit()
        cur.close()

    except Exception as e:
        print(f"[ERROR insertar_datos_mapeados] {e}")
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        release_pg_conn(conn)

# ==================== LOG BITÁCORA ====================
def log_to_pg(dcp_id: str, success: bool, timestamp: datetime.datetime | None, host: str | None):
    """Registrar en TimescaleDB"""
    conn = None
    try:
        conn = get_pg_conn()
        cur = conn.cursor()
        
        ts_utc = datetime.datetime.now(datetime.timezone.utc)
        
        cur.execute("""
            INSERT INTO bitacora_goes 
            (dcp_id, timestamp_utc, timestamp_msg, servidor, exito)
            VALUES (%s, %s, %s, %s, %s)
        """, (dcp_id, ts_utc, timestamp, host, success))
        
        conn.commit()
        cur.close()
        
        print(f"[LOG] {dcp_id} {'✅' if success else '❌'}")
        
    except Exception as e:
        print(f"[ERROR LOG] {e}")
    finally:
        release_pg_conn(conn)

def guardar_header_goes(dcp_id: str, meta: dict):
    """Guardar datos del header GOES en TimescaleDB"""
    conn = None
    try:
        conn = get_pg_conn()
        cur = conn.cursor()
        
        # Signal strength es ASCII directo (ej: "39" = 39)
        signal_strength = None
        try:
            if meta.get("sig"):
                signal_strength = int(meta["sig"])
        except:
            pass
        
        cur.execute("""
            INSERT INTO dcp_headers 
            (timestamp_msg, ts, dcp_id, signal_strength, frequency_offset, mod_index,
             data_quality, channel, spacecraft, data_source, failure_code)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (dcp_id, timestamp_msg) DO NOTHING
        """, (
            meta["timestamp"],
            meta["timestamp"],
            dcp_id,
            signal_strength,
            meta.get("freq"),
            meta.get("mod"),
            meta.get("qual"),
            meta.get("channel"),
            meta.get("spacecraft"),
            meta.get("uplink"),
            meta.get("fail")
        ))
        
        conn.commit()
        cur.close()
        
        if signal_strength:
            print(f"[HEADER] {dcp_id} | Señal: {signal_strength} | Calidad: {meta.get('qual')}")
        
    except Exception as e:
        print(f"[ERROR HEADER] {e}")
    finally:
        release_pg_conn(conn)

# ==================== DECODIFICACIÓN (SIN CAMBIOS) ====================
def ConvertFrom6BitsSutron(binaryDataMsg):
    try:
        dataMsgIn6Bits = binaryDataMsg.decode('ascii')
        valueIn2Base = ''
        result = []
        sub = 64

        for by in dataMsgIn6Bits:
            by = ord(by)
            val = by if by < 64 else by - sub
            r = bin(val)[2:].zfill(6)
            valueIn2Base += r

        while len(valueIn2Base) >= 18:
            base6 = valueIn2Base[:18]
            valueIn2Base = valueIn2Base[18:]
            # Decodifica como entero con signo (complemento a dos, 18 bits)
            if base6[0] == '1':
                val = int(base6, 2) - (1 << 18)
            else:
                val = int(base6, 2)
            result.append(val)

        return result
    except Exception as e:
        print(f"[DECODE] Error: {e}")
        return []

# ==================== DDS PROTOCOL (SIN CAMBIOS) ====================
def compute_auth_hello_body(username: str, password: str) -> bytes:
    now_utc = datetime.datetime.now(datetime.timezone.utc).replace(second=0, microsecond=0)
    doy = now_utc.timetuple().tm_yday
    tstamp = f"{now_utc.year % 100:02d}{doy:03d}{now_utc.hour:02d}{now_utc.minute:02d}{now_utc.second:02d}"
    
    b_name = username.encode("ascii")
    b_passwd = password.encode("ascii")
    
    ms1 = bytearray()
    ms1.extend(b_name)
    ms1.extend(b_passwd)
    ms1.extend(b_name)
    ms1.extend(b_passwd)
    
    b_sha_password = hashlib.sha1(ms1).digest()
    
    dt_epoch = datetime.datetime(1970, 1, 1, tzinfo=datetime.timezone.utc)
    timet = int((now_utc - dt_epoch).total_seconds())
    b_timet = struct.pack(">I", timet)
    
    ms2 = bytearray()
    ms2.extend(b_name)
    ms2.extend(b_sha_password)
    ms2.extend(b_timet)
    ms2.extend(b_name)
    ms2.extend(b_sha_password)
    ms2.extend(b_timet)
    
    b_final_sha_password = hashlib.sha1(ms2).digest()
    auth_hex = b_final_sha_password.hex().upper()
    body_str = f"{username} {tstamp} {auth_hex} 8"
    
    return body_str.encode("ascii")

def dds_send(sock: socket.socket, type_code: str, body: bytes) -> None:
    header = b"FAF0" + type_code.encode("ascii") + f"{len(body):05d}".encode("ascii")
    sock.sendall(header + body)

def dds_recv(sock: socket.socket, timeout: float = 30.0):
    sock.settimeout(timeout)
    hdr = b""
    while len(hdr) < 10:
        chunk = sock.recv(10 - len(hdr))
        if not chunk:
            raise ConnectionError("Conexión cerrada")
        hdr += chunk
    if hdr[:4] != b"FAF0":
        raise ValueError(f"Encabezado inválido: {hdr[:4]}")
    type_code = hdr[4:5].decode("ascii")
    length = int(hdr[5:10].decode("ascii"))
    body = b""
    while len(body) < length:
        chunk = sock.recv(length - len(body))
        if not chunk:
            raise ConnectionError("Cuerpo incompleto")
        body += chunk
    return type_code, body

def parse_iddcp_response(body: bytes):
    if len(body) < 77:
        return None, None, None
    filename = body[:40]
    header = body[40:77]
    payload = body[77:]
    return filename, header, payload

def decode_domsat_header(header: bytes):
    dcp = header[0:8].decode("ascii", errors="replace")
    tstr = header[8:19].decode("ascii", errors="replace")
    yy = int(tstr[0:2]); ddd = int(tstr[2:5]); hh = int(tstr[5:7]); mm = int(tstr[7:9]); ss = int(tstr[9:11])
    year = 2000 + yy
    base = datetime.datetime(year, 1, 1, tzinfo=datetime.timezone.utc) + timedelta(days=ddd - 1)
    ts = base.replace(hour=hh, minute=mm, second=ss)
    return {
        "dcp": dcp,
        "timestamp": ts,
        "tstr": tstr,
        "type": header[19:20].decode("ascii", errors="replace"),
        "sig": header[20:22].decode("ascii", errors="replace"),
        "freq": header[22:24].decode("ascii", errors="replace"),
        "mod": header[24:25].decode("ascii", errors="replace"),
        "qual": header[25:26].decode("ascii", errors="replace"),
        "channel": header[26:29].decode("ascii", errors="replace"),
        "spacecraft": header[29:30].decode("ascii", errors="replace"),
        "uplink": header[30:32].decode("ascii", errors="replace"),
        "mlen": int(header[32:37].decode("ascii", errors="replace")),
    }

def build_criteria_text(dcp_id: str, since: str = "now - 2 hours", until: str = "now", retransmitted: str = "N") -> str:
    lines = ["#", "# LRGS Search Criteria", "#"]
    lines.append(f"DRS_SINCE: {since}")
    lines.append(f"DRS_UNTIL: {until}")
    lines.append(f"DCP_ADDRESS: {dcp_id}")
    lines.append(f"RETRANSMITTED: {retransmitted}")
    return "\n".join(lines) + "\n"

# ==================== DESCARGA ====================
def fetch_from_host(host: str, dcp_id: str, nombre: str, since: str = "now - 2 hours", until: str = "now", multi: bool = False) -> bool:
    print(f"[INFO] {host} -> {dcp_id} | {nombre} | since={since}")
    try:
        user, password = get_credentials(host)

        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.settimeout(20)
            sock.connect((host, 16003))

            dds_send(sock, "m", compute_auth_hello_body(user, password))
            t, _ = dds_recv(sock)
            if t != "m": return False

            criteria_text = build_criteria_text(dcp_id, since=since, until=until,
                                                retransmitted="Y" if multi else "N")
            body = (b" " * 50) + criteria_text.encode("ascii")
            dds_send(sock, "g", body)
            t, _ = dds_recv(sock)
            if t != "g": return False

            def _procesar_mensaje(resp, dcp_id, host):
                """Procesa un mensaje DCS individual. Retorna True si se procesó OK."""
                filename, header, payload = parse_iddcp_response(resp)
                if not header: return False

                raw_header = header
                raw_payload = payload

                meta = decode_domsat_header(header)
                if meta["dcp"].upper() != dcp_id.upper(): return False
                
                guardar_header_goes(dcp_id, meta)

                valores = []
                if payload:
                    try:
                        payload_clean = payload.decode('ascii')[4:].encode('ascii')
                        valores = ConvertFrom6BitsSutron(payload_clean)
                        id_asignado = ESTACIONES[dcp_id].get("id_asignado", "")
                        
                        insertar_datos_mapeados(dcp_id, id_asignado, meta["timestamp"], valores)

                        sensores_estacion = SENSORES.get(id_asignado, [])
                        
                        if not sensores_estacion:
                            print(f"[WARN MIS] No se encontraron sensores para IdAsignado='{id_asignado}' (DCP: {dcp_id})")
                            if id_asignado == "":
                                print(f"[WARN MIS] El IdAsignado está vacío. Verifique tabla NV_GoesSGD campo IdAsignado.")
                            elif id_asignado not in SENSORES:
                                print(f"[WARN MIS] El IdAsignado '{id_asignado}' no existe en keys de SENSORES (Total keys: {len(SENSORES)})")
                        else:
                            print(f"[DEBUG MIS] Generando MIS para {dcp_id} (IdAsignado: {id_asignado}) con {len(sensores_estacion)} sensores.")

                        if config.getboolean('settings', 'generar_mis', fallback=True):
                            generar_archivo_mis(dcp_id, meta, raw_header, raw_payload, valores, ESTACIONES[dcp_id], sensores_estacion)

                    except Exception as e:
                        print(f"[WARN] Error procesando mensaje: {e}")

                log_to_pg(dcp_id, True, meta["timestamp"], host)
                return True

            # Obtener primer mensaje
            dds_send(sock, "f", b"")
            t, resp = dds_recv(sock)
            if t != "f": return False

            if not _procesar_mensaje(resp, dcp_id, host):
                return False
            print(f"[OK] Mensaje 1 procesado")

            # En modo multi: obtener mensajes adicionales con "n" (next)
            if multi:
                msg_count = 1
                while True:
                    try:
                        dds_send(sock, "n", b"")
                        t, resp = dds_recv(sock, timeout=10.0)
                        if t != "n":
                            break
                        if _procesar_mensaje(resp, dcp_id, host):
                            msg_count += 1
                            print(f"[OK] Mensaje {msg_count} procesado")
                        else:
                            break
                    except (ConnectionError, socket.timeout, OSError):
                        break
                print(f"[RECOVER] {dcp_id}: {msg_count} mensajes recuperados")

            return True

    except Exception as e:
        print(f"[ERROR] {host}: {e}")
        return False

def fetch_messages_for_dcp(dcp_id: str, since: str = "now - 2 hours", until: str = "now", multi: bool = False):
    data = ESTACIONES.get(dcp_id)
    if not data:
        print(f"[WARN] {dcp_id} no en BD")
        return False
    nombre = data["nombre"]
    print(f"[START] {dcp_id} | {nombre}")
    global _fetch_ok_count, _fetch_fail_count
    for host in HOSTS:
        if fetch_from_host(host, dcp_id, nombre, since=since, until=until, multi=multi):
            _fetch_ok_count += 1
            return True
    else:
        print(f"[FAIL] Sin datos para {dcp_id}")
        log_to_pg(dcp_id, False, None, None)
        _fetch_fail_count += 1
        return False

def fetch_messages_for_dcp_wrapper(dcp_id):
    if dcp_id not in ESTACIONES:
        print(f"[WARN] {dcp_id} ya no existe en configuración, ignorando.")
        _retry_state.pop(dcp_id, None)
        return schedule.CancelJob
    data = ESTACIONES[dcp_id]
    nombre = data["nombre"]
    print(f"[INFO] {dcp_id} | {nombre} en minuto {datetime.datetime.now().minute:02d}")

    # Cancelar reintentos pendientes de ciclos anteriores
    schedule.clear(f"retry_{dcp_id}")
    _retry_state.pop(dcp_id, None)

    success = fetch_messages_for_dcp(dcp_id)

    if not success:
        _programar_reintento(dcp_id, intento=1)


# ==================== REINTENTO POR ESTACIÓN ====================
_retry_state = {}  # {dcp_id: {"intentos": int, "hora_fallida": datetime}}
RETRY_DELAYS = [int(x) for x in config.get('settings', 'retry_delays', fallback='2,5,10').split(',')]
MAX_RETRIES = config.getint('settings', 'retry_max_intentos', fallback=3)

def _programar_reintento(dcp_id, intento):
    """Programa un reintento one-shot para una estación que no recibió datos."""
    if intento > MAX_RETRIES:
        nombre = ESTACIONES.get(dcp_id, {}).get("nombre", "?")
        print(f"[RETRY AGOTADO] {dcp_id} | {nombre} → {MAX_RETRIES} intentos fallidos, "
              f"se reintentará en su próxima TX programada")
        _retry_state.pop(dcp_id, None)
        return

    delay = RETRY_DELAYS[intento - 1] if intento <= len(RETRY_DELAYS) else RETRY_DELAYS[-1]
    nombre = ESTACIONES.get(dcp_id, {}).get("nombre", "?")

    _retry_state[dcp_id] = {
        "intentos": intento,
        "hora_fallida": datetime.datetime.now(),
    }

    # Programar one-shot: schedule corre la función, devuelve CancelJob → se elimina
    schedule.every(delay).minutes.do(
        lambda did=dcp_id: _ejecutar_reintento(did)
    ).tag(f"retry_{dcp_id}")

    print(f"[RETRY {intento}/{MAX_RETRIES}] {dcp_id} | {nombre} → reintento en {delay} min")

def _ejecutar_reintento(dcp_id):
    """Ejecuta un reintento one-shot y programa el siguiente si falla."""
    if dcp_id not in ESTACIONES:
        _retry_state.pop(dcp_id, None)
        return schedule.CancelJob

    state = _retry_state.get(dcp_id)
    if not state:
        return schedule.CancelJob

    intento = state["intentos"]
    nombre = ESTACIONES[dcp_id]["nombre"]

    # Rotar hosts: usar orden invertido en reintentos impares para diversificar
    hosts_orden = list(reversed(HOSTS)) if intento % 2 == 0 else HOSTS
    print(f"[RETRY {intento}/{MAX_RETRIES}] {dcp_id} | {nombre} → intentando ({hosts_orden[0]}...)")

    success = False
    data = ESTACIONES[dcp_id]
    for host in hosts_orden:
        if fetch_from_host(host, dcp_id, nombre, since="now - 2 hours", until="now", multi=False):
            success = True
            break

    if success:
        print(f"[RETRY OK] {dcp_id} | {nombre} → datos recuperados en intento {intento}")
        _retry_state.pop(dcp_id, None)
    else:
        log_to_pg(dcp_id, False, None, None)
        # Limpiar este job antes de programar el siguiente
        schedule.clear(f"retry_{dcp_id}")
        _programar_reintento(dcp_id, intento + 1)

    return schedule.CancelJob

# ==================== RECUPERACIÓN AUTOMÁTICA DE HUECOS ====================
def detectar_huecos(horas=6):
    """Detecta huecos de datos en las últimas N horas para estaciones que transmiten.
    Solo analiza estaciones que tienen datos históricos en PostgreSQL.
    Retorna lista de estaciones con huecos encontrados."""
    conn = get_pg_conn()
    if not conn:
        print("[AUTO-RECOVER] No se pudo obtener conexión PG para detectar huecos.")
        return []
    estaciones_con_huecos = []
    try:
        cur = conn.cursor()

        # Primero: obtener estaciones que realmente transmiten (tienen datos en los últimos 30 días)
        cur.execute("""
            SELECT DISTINCT dcp_id
            FROM dcp_datos
            WHERE ts > NOW() - INTERVAL '30 days'
        """)
        estaciones_activas = {r[0] for r in cur.fetchall()}
        total_cfg = len(ESTACIONES)
        total_activas = len(estaciones_activas & set(ESTACIONES.keys()))
        print(f"[AUTO-RECOVER] {total_activas}/{total_cfg} estaciones con transmisión reciente (30d)")

        for dcp_id, data in ESTACIONES.items():
            # Saltar estaciones que nunca han transmitido o llevan >30 días sin datos
            if dcp_id not in estaciones_activas:
                continue

            nombre = data["nombre"]
            rango_tx = data.get("rango_transmision", 60)
            tolerancia_min = int(rango_tx * 1.5)

            cur.execute("""
                SELECT DISTINCT ts
                FROM dcp_datos
                WHERE dcp_id = %s
                  AND ts > NOW() - make_interval(hours := %s)
                ORDER BY ts
            """, (dcp_id, horas))
            rows = cur.fetchall()

            if not rows:
                # Estación activa pero sin datos en la ventana de análisis → hueco real
                estaciones_con_huecos.append({
                    "dcp_id": dcp_id, "nombre": nombre, "rango_tx": rango_tx,
                    "huecos": [{"tipo": "SIN DATOS", "desde": f"hace {horas}h",
                                "hasta": "ahora", "duracion_h": horas}]
                })
                continue

            timestamps = [r[0] for r in rows]
            tx_times = sorted(set(
                t.replace(minute=(t.minute // rango_tx) * rango_tx if rango_tx <= 60 else 0,
                          second=0, microsecond=0)
                for t in timestamps
            ))

            huecos = []
            for i in range(1, len(tx_times)):
                diff_min = (tx_times[i] - tx_times[i-1]).total_seconds() / 60
                if diff_min > tolerancia_min:
                    huecos.append({
                        "tipo": "HUECO",
                        "desde": tx_times[i-1].strftime("%Y-%m-%d %H:%M"),
                        "hasta": tx_times[i].strftime("%Y-%m-%d %H:%M"),
                        "duracion_h": round(diff_min / 60, 1)
                    })

            now = datetime.datetime.now(timestamps[0].tzinfo) if timestamps[0].tzinfo else datetime.datetime.utcnow()
            ultima = tx_times[-1]
            diff_final = (now - ultima).total_seconds() / 60
            if diff_final > tolerancia_min:
                huecos.append({
                    "tipo": "RETRASO",
                    "desde": ultima.strftime("%Y-%m-%d %H:%M"),
                    "hasta": "ahora",
                    "duracion_h": round(diff_final / 60, 1)
                })

            if huecos:
                estaciones_con_huecos.append({
                    "dcp_id": dcp_id, "nombre": nombre,
                    "rango_tx": rango_tx, "huecos": huecos
                })
        cur.close()
    except Exception as e:
        print(f"[AUTO-RECOVER] Error detectando huecos: {e}")
    finally:
        release_pg_conn(conn)
    return estaciones_con_huecos

def recuperacion_automatica():
    """Detecta huecos y recupera datos faltantes automáticamente.
    Se ejecuta periódicamente desde el scheduler."""
    recover_hours = config.getint('settings', 'auto_recover_hours', fallback=6)
    print(f"[AUTO-RECOVER] Analizando huecos en las últimas {recover_hours} horas...")
    estaciones_con_huecos = detectar_huecos(recover_hours)

    if not estaciones_con_huecos:
        print(f"[AUTO-RECOVER] Sin huecos detectados.")
        return

    # Solo recuperar estaciones con huecos reales (HUECO o SIN DATOS), no simples retrasos
    estaciones_a_recuperar = [
        est for est in estaciones_con_huecos
        if any(h["tipo"] in ("HUECO", "SIN DATOS") for h in est["huecos"])
    ]

    if not estaciones_a_recuperar:
        retrasos = len(estaciones_con_huecos)
        print(f"[AUTO-RECOVER] Solo {retrasos} estaciones con retraso (se recuperarán en su próxima TX).")
        return

    print(f"[AUTO-RECOVER] {len(estaciones_a_recuperar)} estaciones con huecos → recuperando...")
    for est in estaciones_a_recuperar:
        max_h = max(h["duracion_h"] for h in est["huecos"])
        recover_h = min(int(max_h + 2), recover_hours)
        since = f"now - {recover_h} hours"
        print(f"  [FIX] {est['dcp_id']} | {est['nombre']} → recover {recover_h}h "
              f"({len(est['huecos'])} huecos)")
        try:
            fetch_messages_for_dcp(est["dcp_id"], since=since, multi=True)
        except Exception as e:
            print(f"  [ERROR] {est['dcp_id']}: {e}")
    print(f"[AUTO-RECOVER] Recuperación finalizada.")

# ==================== RECARGA DE CONFIGURACIÓN ====================
_scheduled_dcp_ids = set()

def _programar_estacion(dcp_id, data):
    """Programa la descarga de una estación en el scheduler."""
    minuto = data["minuto"]
    nombre = data["nombre"]
    minuto_ejec = (minuto + 1) % 60
    carry = (minuto + 2) // 60

    if carry:
        schedule.every().hour.at(":00").do(lambda did=dcp_id: fetch_messages_for_dcp_wrapper(did)).tag(f"dcp_{dcp_id}")
        print(f"[SCHED+] {dcp_id} | {nombre} → :00 (próxima hora)")
    else:
        schedule.every().hour.at(f":{minuto_ejec:02d}").do(lambda did=dcp_id: fetch_messages_for_dcp_wrapper(did)).tag(f"dcp_{dcp_id}")
        print(f"[SCHED+] {dcp_id} | {nombre} → :{minuto_ejec:02d}")

def recargar_configuracion():
    """Recarga estaciones y sensores desde SQL Server.
    Si hay cambios (nuevas estaciones, eliminadas o modificadas), actualiza el scheduler."""
    global ESTACIONES, SENSORES, UMBRALES_PRECIP, CUENCAS_ESTACION, _scheduled_dcp_ids
    try:
        print("[RELOAD] Verificando cambios en SQL Server...")
        nuevas_estaciones = load_estaciones_from_sql()
        nuevos_sensores = load_sensores_from_sql()

        if not nuevas_estaciones:
            print("[RELOAD] No se pudieron cargar estaciones, manteniendo configuración actual.")
            return

        # Asignar posiciones a sensores nuevos
        for id_asignado, sensores in nuevos_sensores.items():
            rango_tx = 60
            for dcp_id, datos_estacion in nuevas_estaciones.items():
                if datos_estacion.get("id_asignado") == id_asignado:
                    rango_tx = datos_estacion.get("rango_transmision", 60)
                    break
            nuevos_sensores[id_asignado] = asignar_posiciones_sensores(sensores, rango_tx)

        # Detectar cambios
        ids_actuales = set(ESTACIONES.keys())
        ids_nuevos = set(nuevas_estaciones.keys())

        agregadas = ids_nuevos - ids_actuales
        eliminadas = ids_actuales - ids_nuevos
        # Detectar estaciones con minuto modificado
        modificadas = set()
        for dcp_id in ids_actuales & ids_nuevos:
            if ESTACIONES[dcp_id].get("minuto") != nuevas_estaciones[dcp_id].get("minuto"):
                modificadas.add(dcp_id)

        if not agregadas and not eliminadas and not modificadas:
            # Aún así actualizamos sensores (podrían cambiar cotas, límites, etc.)
            SENSORES = nuevos_sensores
            ESTACIONES = nuevas_estaciones
            UMBRALES_PRECIP = load_umbrales_precipitacion()
            CUENCAS_ESTACION = load_cuencas_from_sql()
            revalidar_datos_invalidos(SENSORES, ESTACIONES)
            print("[RELOAD] Sin cambios en estaciones. Sensores, umbrales y cuencas actualizados.")
            return

        # Aplicar cambios
        if eliminadas:
            for dcp_id in eliminadas:
                schedule.clear(f"dcp_{dcp_id}")
                _scheduled_dcp_ids.discard(dcp_id)
            print(f"[RELOAD] Estaciones eliminadas del scheduler: {eliminadas}")

        if modificadas:
            for dcp_id in modificadas:
                schedule.clear(f"dcp_{dcp_id}")
                _programar_estacion(dcp_id, nuevas_estaciones[dcp_id])
            print(f"[RELOAD] Estaciones reprogramadas: {modificadas}")

        if agregadas:
            for dcp_id in agregadas:
                _programar_estacion(dcp_id, nuevas_estaciones[dcp_id])
                _scheduled_dcp_ids.add(dcp_id)
            print(f"[RELOAD] Nuevas estaciones agregadas al scheduler: {agregadas}")

        ESTACIONES = nuevas_estaciones
        SENSORES = nuevos_sensores
        UMBRALES_PRECIP = load_umbrales_precipitacion()
        CUENCAS_ESTACION = load_cuencas_from_sql()
        revalidar_datos_invalidos(SENSORES, ESTACIONES)
        print(f"[RELOAD] Configuración actualizada. Total estaciones: {len(ESTACIONES)}")

    except Exception as e:
        print(f"[ERROR RELOAD] {e}")

# ==================== DESCARGA DE ASSETS LOCALES ====================
_STATIC_DIR = os.path.join(_base_dir, 'static', 'vendor')
_STATIC_ASSETS = {
    'jquery.min.js':       'https://cdn.jsdelivr.net/npm/jquery@3.7.1/dist/jquery.min.js',
    'semantic.min.css':    'https://cdn.jsdelivr.net/npm/fomantic-ui@2.9.3/dist/semantic.min.css',
    'semantic.min.js':     'https://cdn.jsdelivr.net/npm/fomantic-ui@2.9.3/dist/semantic.min.js',
    'chart.umd.min.js':   'https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js',
    'three.min.js':        'https://cdn.jsdelivr.net/npm/three@0.160.0/build/three.min.js',
}

def _ensure_static_assets():
    """Descarga librerías JS/CSS a static/vendor/ si no existen."""
    os.makedirs(_STATIC_DIR, exist_ok=True)
    # También crear directorio de temas de Fomantic UI
    themes_dir = os.path.join(_STATIC_DIR, 'themes', 'default', 'assets', 'fonts')
    os.makedirs(themes_dir, exist_ok=True)
    import urllib.request
    for filename, url in _STATIC_ASSETS.items():
        filepath = os.path.join(_STATIC_DIR, filename)
        if not os.path.exists(filepath):
            try:
                print(f"[STATIC] Descargando {filename}...")
                urllib.request.urlretrieve(url, filepath)
                print(f"[STATIC] ✓ {filename} descargado")
            except Exception as e:
                print(f"[STATIC] ✗ Error descargando {filename}: {e}")

# ==================== PANEL DE MONITOREO WEB ====================
_MONITOR_HTML = r'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>PIH - Monitor GOES</title>
<link rel="stylesheet" href="/static/vendor/semantic.min.css">
<script src="/static/vendor/chart.umd.min.js"></script>
<script src="/static/vendor/three.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{background:#0f1923;color:#e0e0e0;min-height:100vh;font-family:'Segoe UI',system-ui,sans-serif;overflow-x:hidden}
#threeBg{position:fixed;top:0;left:0;width:100%;height:100%;z-index:0;pointer-events:none}
.app-wrap{position:relative;z-index:1}
.top-bar{padding:14px 24px;border-bottom:2px solid #21ba45;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;box-shadow:0 2px 16px rgba(0,0,0,.4)}
.top-bar h2{margin:0;color:#21ba45;font-size:1.3em;letter-spacing:.5px}
.top-bar .subtitle{color:#5a9;font-size:.7em;display:block;letter-spacing:1px;font-weight:400}
.top-bar .meta{color:#6a8}
.main-grid{padding:16px;max-width:1400px;margin:0 auto}
.dark-card .header{color:#fff!important}
.stat-value{font-size:2.2em;font-weight:bold;color:#21ba45;transition:all .3s}
.stat-label{color:#6a8a9a;font-size:.8em;text-transform:uppercase;letter-spacing:1px}
.log-container{background:#0d1117;border-radius:6px;padding:12px;max-height:400px;overflow-y:auto;font-family:'JetBrains Mono','Fira Code','Courier New',monospace;font-size:.78em;line-height:1.6;border:1px solid #1a2332}
.log-line{border-bottom:1px solid #161b22;padding:2px 4px;transition:background .15s}
.log-line:hover{background:#161b22}
.log-ts{color:#484f58;margin-right:8px}
.log-ok{color:#3fb950}.log-fail{color:#f85149}.log-warn{color:#d29922}.log-info{color:#58a6ff}.log-retry{color:#bc8cff}
table.dark-table{background:#1a2332!important;color:#e0e0e0!important;border-radius:6px;overflow:hidden}
table.dark-table thead th{background:#0d1117!important;color:#8b949e!important;border-bottom:2px solid #21ba45!important;font-weight:600;text-transform:uppercase;font-size:.75em;letter-spacing:.5px}
table.dark-table td{border-top:1px solid #21262d!important;padding:8px 10px!important}
table.dark-table tr:hover td{background:#161b22!important}
.retry-badge{background:#8957e5!important}.online-badge{background:#238636!important}.offline-badge{background:#da3633!important}
#stationFilter{background:#0d1117;border:1px solid #30363d;color:#e0e0e0;padding:8px 14px;border-radius:6px;width:320px;transition:border-color .2s}
#stationFilter:focus{border-color:#21ba45;outline:none;box-shadow:0 0 0 3px rgba(33,186,69,.15)}
.ui.inverted.menu .item{color:#8b949e!important;transition:color .2s}
.ui.inverted.menu .item.active{color:#21ba45!important;border-color:#21ba45!important}
.ui.inverted.menu .item:hover{color:#3fb950!important}
.pulse{animation:pulse 2s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.4}}
@keyframes fadeIn{from{opacity:0;transform:translateY(16px)}to{opacity:1;transform:translateY(0)}}
@keyframes slideInLeft{from{opacity:0;transform:translateX(-30px)}to{opacity:1;transform:translateX(0)}}
@keyframes slideInRight{from{opacity:0;transform:translateX(30px)}to{opacity:1;transform:translateX(0)}}
@keyframes slideInUp{from{opacity:0;transform:translateY(30px)}to{opacity:1;transform:translateY(0)}}
@keyframes zoomIn{from{opacity:0;transform:scale(.6)}to{opacity:1;transform:scale(1)}}
@keyframes countUp{from{opacity:0;transform:scale(.8)}to{opacity:1;transform:scale(1)}}
@keyframes shimmer{0%{background-position:-200% 0}100%{background-position:200% 0}}
@keyframes glow{0%,100%{box-shadow:0 0 5px rgba(33,186,69,.2)}50%{box-shadow:0 0 20px rgba(33,186,69,.4),0 0 40px rgba(33,186,69,.1)}}
@keyframes borderGlow{0%,100%{border-color:#2a3a4a}50%{border-color:#21ba45}}
@keyframes float{0%,100%{transform:translateY(0)}50%{transform:translateY(-4px)}}
.fade-in{animation:fadeIn .5s ease-out both}
.slide-in-left{animation:slideInLeft .5s ease-out both}
.slide-in-right{animation:slideInRight .5s ease-out both}
.slide-in-up{animation:slideInUp .5s ease-out both}
.zoom-in{animation:zoomIn .4s ease-out both}
.anim-delay-1{animation-delay:.1s}
.anim-delay-2{animation-delay:.2s}
.anim-delay-3{animation-delay:.3s}
.anim-delay-4{animation-delay:.4s}
.stat-value{animation:countUp .5s ease-out}
.live-dot{display:inline-block;width:8px;height:8px;border-radius:50%;background:#21ba45;margin-right:6px;animation:pulse 1.5s infinite;box-shadow:0 0 6px rgba(33,186,69,.6)}
.action-card{background:#1a2332;border:1px solid #2a3a4a;border-radius:8px;padding:20px;transition:border-color .3s,box-shadow .3s,transform .3s}
.action-card:hover{border-color:#21ba45;box-shadow:0 0 20px rgba(33,186,69,.15);transform:translateY(-2px)}
.action-card h4{color:#3fb950;margin-top:0;display:flex;align-items:center;gap:8px}
.action-card h4 .icon-wrap{background:rgba(33,186,69,.1);border-radius:6px;padding:6px 8px;font-size:1.2em}
.action-help{background:#0d1117;border:1px solid #21262d;border-radius:6px;padding:12px 16px;margin-top:12px;font-size:.85em;color:#8b949e}
.action-help code{background:#161b22;padding:2px 6px;border-radius:3px;color:#58a6ff;font-size:.9em}
.chart-container{background:#1a2332;border:1px solid #2a3a4a;border-radius:8px;padding:16px;margin-bottom:16px}
.progress-bar{height:4px;background:#21262d;border-radius:2px;overflow:hidden;margin-top:6px}
.progress-fill{height:100%;background:linear-gradient(90deg,#21ba45,#3fb950);border-radius:2px;transition:width .5s ease}
.dark-card{background:#1a2332!important;border:1px solid #2a3a4a!important;color:#e0e0e0!important;transition:transform .3s,box-shadow .3s,border-color .3s}
.dark-card:hover{transform:translateY(-3px);box-shadow:0 8px 25px rgba(33,186,69,.15)!important;border-color:#21ba45!important}
.top-bar{backdrop-filter:blur(12px);background:linear-gradient(135deg,rgba(26,35,50,.95) 0%,rgba(13,79,60,.95) 100%)!important}
.shimmer-bar{background:linear-gradient(90deg,#21262d 0%,#2a3a4a 50%,#21262d 100%);background-size:200% 100%;animation:shimmer 2s infinite}
.glow-border{animation:borderGlow 3s ease-in-out infinite}
.floating{animation:float 3s ease-in-out infinite}
</style>
</head>
<body>
<canvas id="threeBg"></canvas>
<div class="app-wrap">
<div class="top-bar">
 <div style="display:flex;align-items:center;gap:14px">
  <div style="font-size:1.8em" class="floating">&#127754;</div>
  <h2>Plataforma Integral Hidrometrica<span class="subtitle">Monitor GOES &middot; Ingesta Satelital en Tiempo Real</span></h2>
  <span id="statusBadge" class="ui tiny green label" style="margin-left:8px">ACTIVO</span>
  <span class="meta" id="uptimeLabel">Uptime: --</span>
 </div>
 <div style="display:flex;align-items:center;gap:14px">
  <div style="text-align:right;line-height:1.4">
   <div style="display:flex;align-items:center;gap:6px;justify-content:flex-end"><span class="live-dot"></span><span style="color:#e0e0e0;font-size:.95em;font-weight:600" id="clockLocal">--:--:--</span><span style="color:#6a8a9a;font-size:.7em">LOCAL</span></div>
   <div style="display:flex;align-items:center;gap:6px;justify-content:flex-end"><span style="color:#8b949e;font-size:.85em" id="clockUTC">--:--:--</span><span style="color:#6a8a9a;font-size:.7em">UTC</span></div>
  </div>
  <button class="ui tiny yellow button" id="btnPause" onclick="toggleScheduler()" style="border-radius:6px">&#9208; Pausar</button>
 </div>
</div>
<div class="main-grid">
 <div class="ui four column stackable grid" style="margin-bottom:16px">
  <div class="column zoom-in anim-delay-1"><div class="ui card dark-card glow-border fluid"><div class="content" style="text-align:center;padding:18px"><div class="stat-value" id="statEstaciones">--</div><div class="stat-label">Estaciones</div><div class="progress-bar"><div class="progress-fill" id="prgEst" style="width:0%"></div></div></div></div></div>
  <div class="column zoom-in anim-delay-2"><div class="ui card dark-card glow-border fluid"><div class="content" style="text-align:center;padding:18px"><div class="stat-value" style="color:#58a6ff" id="statScheduled">--</div><div class="stat-label">Programadas</div><div class="progress-bar"><div class="progress-fill" id="prgSch" style="width:0%;background:linear-gradient(90deg,#1f6feb,#58a6ff)"></div></div></div></div></div>
  <div class="column zoom-in anim-delay-3"><div class="ui card dark-card glow-border fluid"><div class="content" style="text-align:center;padding:18px"><div class="stat-value" style="color:#bc8cff" id="statRetries">--</div><div class="stat-label">Reintentos Activos</div><div class="progress-bar"><div class="progress-fill" id="prgRet" style="width:0%;background:linear-gradient(90deg,#8957e5,#bc8cff)"></div></div></div></div></div>
  <div class="column zoom-in anim-delay-4"><div class="ui card dark-card glow-border fluid"><div class="content" style="text-align:center;padding:18px"><div class="stat-value" style="color:#d29922" id="statJobs">--</div><div class="stat-label">Jobs Pendientes</div><div class="progress-bar"><div class="progress-fill" id="prgJob" style="width:0%;background:linear-gradient(90deg,#9e6a03,#d29922)"></div></div></div></div></div>
 </div>
 <!-- GRAFICA movida a tab -->
 <div class="ui top attached inverted tabular menu" style="background:#0d1117;border-color:#21262d">
  <a class="item active" data-tab="stations">&#128225; Estaciones</a>
  <a class="item" data-tab="chart">&#128200; Actividad</a>
  <a class="item" data-tab="retries">&#128260; Reintentos</a>
  <a class="item" data-tab="bitacora">&#128203; Bitacora</a>
  <a class="item" data-tab="log">&#128421; Consola</a>
  <a class="item" data-tab="config">&#9881; Config</a>
  <a class="item" data-tab="actions">&#128640; Acciones</a>
 </div>
 <!-- TAB Actividad en Tiempo Real -->
 <div class="ui bottom attached inverted tab segment" data-tab="chart" style="background:#1a2332;border-color:#2a3a4a">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
   <span style="color:#8b949e;font-size:.85em;font-weight:600;text-transform:uppercase;letter-spacing:1px"><span class="live-dot"></span> Actividad en Tiempo Real</span>
   <span style="color:#484f58;font-size:.75em" id="chartRange">Ultimos 30 puntos</span>
  </div>
  <div style="height:calc(100vh - 380px);min-height:200px;max-height:500px">
   <canvas id="liveChart"></canvas>
  </div>
 </div>
 <!-- TAB Estaciones -->
 <div class="ui bottom attached inverted tab segment active" data-tab="stations" style="background:#1a2332;border-color:#2a3a4a">
  <div style="margin-bottom:12px;display:flex;justify-content:space-between;align-items:center">
   <input type="text" id="stationFilter" placeholder="&#128269; Buscar por DCP ID, nombre, cuenca..." oninput="filterStations()">
   <span class="meta" id="stationCount">--</span>
  </div>
  <div style="max-height:500px;overflow-y:auto;border-radius:6px">
   <table class="ui compact small unstackable table dark-table"><thead><tr><th>DCP ID</th><th>Nombre</th><th>ID Asignado</th><th>Min TX</th><th>Rango</th><th>Cuenca</th><th>Estado</th><th>Accion</th></tr></thead><tbody id="stationsBody"></tbody></table>
  </div>
 </div>
 <!-- TAB Reintentos -->
 <div class="ui bottom attached inverted tab segment" data-tab="retries" style="background:#1a2332;border-color:#2a3a4a"><div id="retriesContent"><p style="color:#6a8a9a">Cargando...</p></div></div>
 <!-- TAB Bitacora -->
 <div class="ui bottom attached inverted tab segment" data-tab="bitacora" style="background:#1a2332;border-color:#2a3a4a">
  <div style="margin-bottom:10px">
   <select id="bitacoraHoras" style="background:#0d1117;color:#e0e0e0;border:1px solid #30363d;padding:6px 10px;border-radius:4px"><option value="1">1 hora</option><option value="3">3 horas</option><option value="6" selected>6 horas</option><option value="12">12 horas</option><option value="24">24 horas</option></select>
   <button class="ui tiny blue button" onclick="loadBitacora()" style="border-radius:4px">Actualizar</button>
   <span class="meta" id="bitacoraCount" style="margin-left:12px">--</span>
  </div>
  <div style="max-height:500px;overflow-y:auto;border-radius:6px">
   <table class="ui compact small unstackable table dark-table"><thead><tr><th>Hora UTC</th><th>DCP ID</th><th>Nombre</th><th>Msg Timestamp</th><th>Servidor</th><th>Resultado</th></tr></thead><tbody id="bitacoraBody"></tbody></table>
  </div>
 </div>
 <!-- TAB Consola -->
 <div class="ui bottom attached inverted tab segment" data-tab="log" style="background:#1a2332;border-color:#2a3a4a">
  <div style="margin-bottom:8px;display:flex;align-items:center;gap:12px">
   <button class="ui tiny blue button" onclick="loadLog()" style="border-radius:4px">Actualizar</button>
   <label style="color:#6a8a9a"><input type="checkbox" id="autoScroll" checked> Auto-scroll</label>
   <label style="color:#6a8a9a"><input type="checkbox" id="autoRefreshLog" checked> Auto-refresh (5s)</label>
  </div>
  <div class="log-container" id="logContainer"></div>
 </div>
 <!-- TAB Config -->
 <div class="ui bottom attached inverted tab segment" data-tab="config" style="background:#1a2332;border-color:#2a3a4a"><div id="configContent"><p style="color:#6a8a9a">Cargando...</p></div></div>
 <!-- TAB Acciones -->
 <div class="ui bottom attached inverted tab segment" data-tab="actions" style="background:#1a2332;border-color:#2a3a4a">
  <div class="ui three column stackable grid">
   <div class="column">
    <div class="action-card">
     <h4><span class="icon-wrap">&#128260;</span> Recuperacion de Datos</h4>
     <p style="color:#8b949e;font-size:.85em;margin-bottom:14px">Busca mensajes GOES historicos en los servidores LRGS y los reinserta en la base de datos. Ideal para llenar huecos por cortes o fallas.</p>
     <div class="ui inverted form">
      <div class="field"><label style="color:#6a8a9a">DCP ID <span style="color:#484f58">(vacio = todas las estaciones)</span></label><input type="text" id="recoverDcpId" placeholder="Ej: E8906DD6" style="background:#0d1117;border:1px solid #30363d;border-radius:4px"></div>
      <div class="field"><label style="color:#6a8a9a">Horas hacia atras</label><input type="number" id="recoverHoras" value="12" min="1" max="168" style="background:#0d1117;border:1px solid #30363d;border-radius:4px"></div>
      <button class="ui fluid green button" onclick="doRecover()" style="border-radius:6px;margin-top:4px">&#128260; Iniciar Recuperacion</button>
     </div>
     <div id="recoverResult" style="margin-top:12px"></div>
     <div class="action-help">
      <strong style="color:#58a6ff">&#9432; Como funciona:</strong><br>
      1. Conecta a servidores LRGS (USGS)<br>
      2. Solicita mensajes de las ultimas <code>N horas</code><br>
      3. Decodifica y valida cada mensaje GOES<br>
      4. Inserta los datos faltantes en TimescaleDB<br>
      5. Los datos duplicados se ignoran automaticamente<br><br>
      <strong style="color:#d29922">&#9888; Nota:</strong> LRGS retiene ~72h de mensajes. Para datos mas antiguos usa archivos <code>.mis</code>
     </div>
    </div>
   </div>
   <div class="column">
    <div class="action-card">
     <h4><span class="icon-wrap">&#128225;</span> Fetch Manual</h4>
     <p style="color:#8b949e;font-size:.85em;margin-bottom:14px">Descarga inmediata de una estacion especifica, sin esperar al scheduler. Util para verificar que una estacion esta transmitiendo.</p>
     <div class="ui inverted form">
      <div class="field"><label style="color:#6a8a9a">DCP ID</label><input type="text" id="fetchDcpId" placeholder="Ej: E8906DD6" style="background:#0d1117;border:1px solid #30363d;border-radius:4px"></div>
      <div class="field"><label style="color:#6a8a9a">Horas</label><input type="number" id="fetchHoras" value="2" min="1" max="48" style="background:#0d1117;border:1px solid #30363d;border-radius:4px"></div>
      <button class="ui fluid blue button" onclick="doFetch()" style="border-radius:6px;margin-top:4px">&#128225; Ejecutar Fetch</button>
     </div>
     <div id="fetchResult" style="margin-top:12px"></div>
     <div class="action-help">
      <strong style="color:#58a6ff">&#9432; Diferencia con Recover:</strong><br>
      &bull; <strong>Fetch</strong>: descarga 1 estacion, ventana corta (1-2h)<br>
      &bull; <strong>Recover</strong>: puede ser todas las estaciones, ventana larga (hasta 168h)
     </div>
    </div>
   </div>
   <div class="column">
    <div class="action-card">
     <h4><span class="icon-wrap">&#9881;</span> Control del Servicio</h4>
     <p style="color:#8b949e;font-size:.85em;margin-bottom:14px">Controla el scheduler de descargas automaticas. Las acciones manuales (Fetch/Recover) siguen funcionando aunque el scheduler este pausado.</p>
     <button class="ui fluid yellow button" id="btnPauseMain" onclick="toggleScheduler()" style="border-radius:6px">&#9208; Pausar Scheduler</button>
     <div style="margin-top:16px;padding:12px;background:#0d1117;border-radius:6px;border:1px solid #21262d">
      <div style="display:flex;justify-content:space-between;margin-bottom:6px"><span style="color:#6a8a9a;font-size:.8em">Estado</span><span id="svcStatus" style="font-size:.85em;color:#3fb950">Activo</span></div>
      <div style="display:flex;justify-content:space-between;margin-bottom:6px"><span style="color:#6a8a9a;font-size:.8em">Modo</span><span style="font-size:.85em;color:#8b949e">Scheduler automatico</span></div>
      <div style="display:flex;justify-content:space-between"><span style="color:#6a8a9a;font-size:.8em">Monitor Web</span><span style="font-size:.85em;color:#58a6ff">Puerto 5555</span></div>
     </div>
     <div class="action-help" style="margin-top:12px">
      <strong style="color:#58a6ff">&#9432; Modos del sistema:</strong><br>
      &bull; <strong style="color:#3fb950">Activo</strong>: Descarga automatica cada hora<br>
      &bull; <strong style="color:#d29922">Pausado</strong>: Solo acciones manuales<br>
      &bull; <strong>Reintentos</strong>: Automaticos a +2, +5, +10 min<br>
      &bull; <strong>Recover auto</strong>: Cada 30 min busca huecos
     </div>
    </div>
   </div>
  </div>
  <!-- Seccion Archivos .mis -->
  <div style="margin-top:20px">
   <div class="action-card">
    <h4><span class="icon-wrap">&#128193;</span> Archivos .mis</h4>
    <p style="color:#8b949e;font-size:.85em;margin:-8px 0 14px 0">Dos funciones independientes: <strong style="color:#58a6ff">Generaci&oacute;n</strong> (exportar .mis para sistemas alternos) e <strong style="color:#d29922">Importaci&oacute;n</strong> (rellenar huecos de datos desde archivos .mis externos)</p>
    <div class="ui two column stackable grid">
     <div class="column">
      <h5 style="color:#58a6ff;margin-top:0;margin-bottom:10px;text-transform:uppercase;font-size:.8em;letter-spacing:1px">&#128228; Generaci&oacute;n &mdash; Exportar para sistema alterno</h5>
      <p style="color:#6a8a9a;font-size:.8em;margin:-4px 0 10px 0">Crea archivos .mis por cada fetch GOES para alimentar otros sistemas que usan .mis como fuente de datos.</p>
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:14px">
       <span style="color:#8b949e;font-size:.9em">Activa:</span>
       <div class="ui toggle checkbox" id="misToggleWrap"><input type="checkbox" id="misToggle" onchange="toggleMisGen(this.checked)"><label style="color:#e0e0e0"></label></div>
       <span id="misToggleLabel" style="font-size:.85em;color:#3fb950">Activa</span>
      </div>
      <div style="background:#0d1117;border:1px solid #21262d;border-radius:6px;padding:12px;margin-bottom:12px">
       <div style="display:flex;justify-content:space-between;margin-bottom:6px"><span style="color:#6a8a9a;font-size:.8em">Directorio salida</span><span id="misOutDir" style="font-size:.8em;color:#8b949e;word-break:break-all">--</span></div>
       <div style="display:flex;justify-content:space-between;margin-bottom:6px"><span style="color:#6a8a9a;font-size:.8em">Archivos generados</span><span id="misOutCount" style="font-size:.85em;color:#58a6ff">--</span></div>
       <div style="display:flex;justify-content:space-between;margin-bottom:6px"><span style="color:#6a8a9a;font-size:.8em">Directorio importacion</span><span id="misImportDir" style="font-size:.8em;color:#8b949e;word-break:break-all">--</span></div>
       <div style="display:flex;justify-content:space-between;margin-bottom:6px"><span style="color:#6a8a9a;font-size:.8em">Auto-import cada</span><span id="misImportInterval" style="font-size:.85em;color:#d29922">-- min</span></div>
       <div style="display:flex;justify-content:space-between;margin-bottom:6px"><span style="color:#6a8a9a;font-size:.8em">Procesados</span><span id="misProcCount" style="font-size:.85em;color:#3fb950">--</span></div>
       <div style="display:flex;justify-content:space-between"><span style="color:#6a8a9a;font-size:.8em">Errores</span><span id="misErrCount" style="font-size:.85em;color:#f85149">--</span></div>
      </div>
      <div class="action-help">
       <strong style="color:#58a6ff">&#9432; Generaci&oacute;n vs Importaci&oacute;n:</strong><br>
       &bull; <strong style="color:#58a6ff">Generaci&oacute;n</strong>: Exporta un <code>.mis</code> por cada fetch GOES &rarr; alimenta <em>otro sistema</em> que consume .mis como fuente de datos<br>
       &bull; <strong style="color:#d29922">Importaci&oacute;n</strong>: Lee archivos <code>.mis</code> externos para <em>rellenar huecos</em> espec&iacute;ficos en la base de datos de este sistema<br>
       &bull; Son funciones <strong>independientes</strong>: se puede desactivar la generaci&oacute;n sin afectar la importaci&oacute;n y viceversa<br>
       &bull; Procesados se mueven a <code>procesados/</code>, fallidos a <code>errores/</code><br>
       &bull; Duplicados se ignoran (UPSERT por timestamp+sensor)
      </div>
     </div>
     <div class="column">
      <h5 style="color:#d29922;margin-top:0;margin-bottom:10px;text-transform:uppercase;font-size:.8em;letter-spacing:1px">&#128229; Importaci&oacute;n &mdash; Rellenar huecos de datos</h5>
      <p style="color:#6a8a9a;font-size:.8em;margin:-4px 0 10px 0">Coloca archivos .mis en la carpeta de importaci&oacute;n para llenar huecos espec&iacute;ficos en la base de datos.</p>
      <h5 style="color:#8b949e;margin-top:0;margin-bottom:8px;text-transform:uppercase;font-size:.75em;letter-spacing:1px">Archivos pendientes</h5>
      <div id="misPendingList" style="max-height:250px;overflow-y:auto;background:#0d1117;border:1px solid #21262d;border-radius:6px;padding:8px;margin-bottom:12px">
       <p style="color:#484f58;text-align:center;padding:20px">Cargando...</p>
      </div>
      <div style="display:flex;gap:8px">
       <button class="ui fluid green button" onclick="importAllMis()" style="border-radius:6px">&#128194; Importar Todos</button>
       <button class="ui fluid blue button" onclick="loadMisStatus()" style="border-radius:6px">&#128260; Actualizar</button>
      </div>
      <div id="misImportResult" style="margin-top:12px"></div>
     </div>
    </div>
   </div>
  </div>
 </div>
</div>
</div><!-- /app-wrap -->
<script src="/static/vendor/jquery.min.js"></script>
<script src="/static/vendor/semantic.min.js"></script>
<script>
/* ===== THREE.JS - LLUVIA DE PARTICULAS ===== */
(function(){
 const canvas=document.getElementById('threeBg');
 const renderer=new THREE.WebGLRenderer({canvas,alpha:true,antialias:false});
 renderer.setPixelRatio(Math.min(window.devicePixelRatio,1.5));
 renderer.setSize(window.innerWidth,window.innerHeight);
 const scene=new THREE.Scene();
 const camera=new THREE.PerspectiveCamera(60,window.innerWidth/window.innerHeight,.1,1000);
 camera.position.z=50;
 // Particulas tipo lluvia/gotas
 const COUNT=600;
 const geo=new THREE.BufferGeometry();
 const pos=new Float32Array(COUNT*3);
 const vel=new Float32Array(COUNT);
 const sizes=new Float32Array(COUNT);
 for(let i=0;i<COUNT;i++){
  pos[i*3]=(Math.random()-0.5)*100;
  pos[i*3+1]=(Math.random()-0.5)*100;
  pos[i*3+2]=(Math.random()-0.5)*60;
  vel[i]=0.02+Math.random()*0.08;
  sizes[i]=0.8+Math.random()*1.5;
 }
 geo.setAttribute('position',new THREE.BufferAttribute(pos,3));
 geo.setAttribute('size',new THREE.BufferAttribute(sizes,1));
 // Shader material para puntos brillantes
 const mat=new THREE.ShaderMaterial({
  transparent:true,depthWrite:false,blending:THREE.AdditiveBlending,
  uniforms:{uTime:{value:0},uColor1:{value:new THREE.Color('#21ba45')},uColor2:{value:new THREE.Color('#58a6ff')}},
  vertexShader:`
   attribute float size;
   varying float vAlpha;
   void main(){
    vec4 mv=modelViewMatrix*vec4(position,1.0);
    gl_PointSize=size*(300.0/(-mv.z));
    gl_Position=projectionMatrix*mv;
    vAlpha=smoothstep(-50.0,0.0,mv.z)*0.6;
   }`,
  fragmentShader:`
   uniform vec3 uColor1;uniform vec3 uColor2;uniform float uTime;
   varying float vAlpha;
   void main(){
    float d=length(gl_PointCoord-0.5)*2.0;
    if(d>1.0)discard;
    float glow=1.0-d*d;
    vec3 col=mix(uColor1,uColor2,gl_PointCoord.y+sin(uTime*0.5)*0.3);
    gl_FragColor=vec4(col,glow*vAlpha*0.4);
   }`
 });
 const particles=new THREE.Points(geo,mat);
 scene.add(particles);
 // Lineas de conexion sutiles
 const lineGeo=new THREE.BufferGeometry();
 const linePts=[];
 for(let i=0;i<40;i++){
  const x1=(Math.random()-0.5)*80,y1=(Math.random()-0.5)*80,z1=(Math.random()-0.5)*30;
  const x2=x1+(Math.random()-0.5)*15,y2=y1+(Math.random()-0.5)*15,z2=z1+(Math.random()-0.5)*10;
  linePts.push(x1,y1,z1,x2,y2,z2);
 }
 lineGeo.setAttribute('position',new THREE.Float32BufferAttribute(linePts,3));
 const lineMat=new THREE.LineBasicMaterial({color:0x21ba45,transparent:true,opacity:0.06});
 scene.add(new THREE.LineSegments(lineGeo,lineMat));
 let time=0;
 function animate(){
  requestAnimationFrame(animate);
  time+=0.01;
  mat.uniforms.uTime.value=time;
  const p=geo.attributes.position.array;
  for(let i=0;i<COUNT;i++){
   p[i*3+1]-=vel[i];
   if(p[i*3+1]<-50){p[i*3+1]=50;p[i*3]=(Math.random()-0.5)*100}
  }
  geo.attributes.position.needsUpdate=true;
  particles.rotation.y=Math.sin(time*0.1)*0.05;
  renderer.render(scene,camera);
 }
 animate();
 window.addEventListener('resize',()=>{
  camera.aspect=window.innerWidth/window.innerHeight;
  camera.updateProjectionMatrix();
  renderer.setSize(window.innerWidth,window.innerHeight);
 });
})();
/* ===== FIN THREE.JS ===== */
$('.tabular.menu .item').tab();
let _stations=[],_isRunning=true;
const MAX_CHART_POINTS=30;
let _chartData={labels:[],ok:[],fail:[],retry:[]};
const _liveChart=(()=>{const ctx=document.getElementById('liveChart').getContext('2d');
 return new Chart(ctx,{type:'line',data:{labels:[],datasets:[
  {label:'OK',data:[],borderColor:'#3fb950',backgroundColor:'rgba(63,185,80,.1)',fill:true,tension:.35,pointRadius:2,borderWidth:2},
  {label:'Fallo',data:[],borderColor:'#f85149',backgroundColor:'rgba(248,81,73,.1)',fill:true,tension:.35,pointRadius:2,borderWidth:2},
  {label:'Reintentos',data:[],borderColor:'#bc8cff',backgroundColor:'rgba(188,140,255,.08)',fill:true,tension:.35,pointRadius:2,borderWidth:2}
 ]},options:{responsive:true,maintainAspectRatio:false,animation:{duration:600,easing:'easeOutQuart'},interaction:{intersect:false,mode:'index'},
  plugins:{legend:{labels:{color:'#8b949e',usePointStyle:true,pointStyle:'circle',padding:16,font:{size:11}}}},
  scales:{x:{ticks:{color:'#484f58',maxTicksLimit:8,font:{size:10}},grid:{color:'#21262d'}},y:{beginAtZero:true,ticks:{color:'#484f58',stepSize:1,font:{size:10}},grid:{color:'#21262d'}}}}})
})();
function updateChart(ok,fail,retry){
 const now=new Date().toLocaleTimeString('es-MX',{hour:'2-digit',minute:'2-digit',second:'2-digit'});
 _chartData.labels.push(now);_chartData.ok.push(ok);_chartData.fail.push(fail);_chartData.retry.push(retry);
 if(_chartData.labels.length>MAX_CHART_POINTS){_chartData.labels.shift();_chartData.ok.shift();_chartData.fail.shift();_chartData.retry.shift()}
 _liveChart.data.labels=_chartData.labels;_liveChart.data.datasets[0].data=_chartData.ok;_liveChart.data.datasets[1].data=_chartData.fail;_liveChart.data.datasets[2].data=_chartData.retry;
 _liveChart.update()}
setInterval(()=>{
 const now=new Date();
 const dl=now.toLocaleDateString('es-MX',{day:'2-digit',month:'short',year:'numeric'});
 const tl=now.toLocaleTimeString('es-MX',{hour:'2-digit',minute:'2-digit',second:'2-digit'});
 document.getElementById('clockLocal').textContent=dl+' '+tl;
 const du=now.toUTCString().replace(/ GMT$/,'');
 document.getElementById('clockUTC').textContent=now.toISOString().replace('T',' ').substring(0,19)+' GMT';
},1000);
let _prevOk=0,_prevFail=0;
function loadStatus(){fetch('/api/status').then(r=>r.json()).then(d=>{
 _isRunning=d.running;
 document.getElementById('uptimeLabel').textContent='Uptime: '+d.uptime;
 document.getElementById('statEstaciones').textContent=d.total_estaciones;
 document.getElementById('statScheduled').textContent=d.total_scheduled;
 document.getElementById('statRetries').textContent=d.retries_activos;
 document.getElementById('statJobs').textContent=d.pending_jobs;
 const t=d.total_estaciones||1;
 document.getElementById('prgEst').style.width='100%';
 document.getElementById('prgSch').style.width=Math.round((d.total_scheduled/t)*100)+'%';
 document.getElementById('prgRet').style.width=Math.min(Math.round((d.retries_activos/t)*100),100)+'%';
 document.getElementById('prgJob').style.width=Math.min(Math.round((d.pending_jobs/Math.max(t,1))*100),100)+'%';
 const ok=d.fetch_ok||0,fail=d.fetch_fail||0;
 updateChart(ok-_prevOk,fail-_prevFail,d.retries_activos);
 _prevOk=ok;_prevFail=fail;
 const b=document.getElementById('statusBadge'),p=document.getElementById('btnPause'),m=document.getElementById('btnPauseMain'),sv=document.getElementById('svcStatus');
 if(d.running){b.className='ui tiny green label';b.textContent='ACTIVO';p.innerHTML='&#9208; Pausar';p.className='ui tiny yellow button';m.innerHTML='&#9208; Pausar Scheduler';m.className='ui fluid yellow button';sv.textContent='Activo';sv.style.color='#3fb950'}
 else{b.className='ui tiny red label pulse';b.textContent='PAUSADO';p.innerHTML='&#9654; Reanudar';p.className='ui tiny green button';m.innerHTML='&#9654; Reanudar Scheduler';m.className='ui fluid green button';sv.textContent='Pausado';sv.style.color='#d29922'}
}).catch(()=>{})}
function loadStations(){fetch('/api/stations').then(r=>r.json()).then(d=>{_stations=d;renderStations(d);document.getElementById('stationCount').textContent=d.length+' estaciones'})}
function renderStations(data){
 const f=document.getElementById('stationFilter').value.toLowerCase();
 const fd=f?data.filter(s=>s.dcp_id.toLowerCase().includes(f)||s.nombre.toLowerCase().includes(f)||(s.cuenca||'').toLowerCase().includes(f)||(s.id_asignado||'').toLowerCase().includes(f)):data;
 let h='';for(const s of fd){
  const e=s.retry?'<span class="ui tiny label retry-badge">'+s.retry.intentos+'/3 &#8635; '+s.retry.hora_fallida+'</span>':(s.scheduled?'<span class="ui tiny label online-badge">OK</span>':'<span class="ui tiny label offline-badge">OFF</span>');
  h+='<tr class="fade-in"><td><b>'+s.dcp_id+'</b></td><td>'+s.nombre+'</td><td>'+s.id_asignado+'</td><td>:'+String(s.minuto_tx).padStart(2,'0')+'</td><td>'+s.rango_tx+'m</td><td>'+(s.cuenca||'-')+'</td><td>'+e+'</td><td><button class="ui tiny compact blue button" style="border-radius:4px" onclick="quickFetch(\''+s.dcp_id+'\')">Fetch</button></td></tr>'}
 document.getElementById('stationsBody').innerHTML=h;
 document.getElementById('stationCount').textContent=fd.length+'/'+data.length+' estaciones'}
function filterStations(){renderStations(_stations)}
function quickFetch(id){fetch('/api/fetch/'+id,{method:'POST'}).then(r=>r.json()).then(d=>{if(d.success)alert('Fetch iniciado: '+id)})}
function loadRetries(){fetch('/api/retries').then(r=>r.json()).then(data=>{
 if(!data.length){document.getElementById('retriesContent').innerHTML='<div style="text-align:center;padding:40px;color:#6a8a9a" class="fade-in"><div style="font-size:3em;margin-bottom:8px">&#10004;</div><h3 style="color:#3fb950">Sin reintentos activos</h3><p>Todas las estaciones respondieron correctamente</p></div>';return}
 let h='<table class="ui compact small table dark-table"><thead><tr><th>DCP ID</th><th>Nombre</th><th>Intento</th><th>Hora Falla</th><th>Accion</th></tr></thead><tbody>';
 for(const r of data)h+='<tr class="fade-in"><td><b>'+r.dcp_id+'</b></td><td>'+r.nombre+'</td><td><span class="ui tiny purple label">'+r.intentos+'/'+r.max_intentos+'</span></td><td>'+r.hora_fallida+'</td><td><button class="ui tiny compact green button" style="border-radius:4px" onclick="quickFetch(\''+r.dcp_id+'\')">Forzar</button></td></tr>';
 h+='</tbody></table>';document.getElementById('retriesContent').innerHTML=h})}
function loadBitacora(){const h=document.getElementById('bitacoraHoras').value;fetch('/api/bitacora?horas='+h+'&limit=200').then(r=>r.json()).then(data=>{
 if(data.error){document.getElementById('bitacoraBody').innerHTML='<tr><td colspan="6" style="color:#f85149">'+data.error+'</td></tr>';return}
 let html='';for(const r of data){const ic=r.exito?'<span style="color:#3fb950">&#10004;</span>':'<span style="color:#f85149">&#10008;</span>';const cl=r.exito?'':'style="background:#2d1215!important"';
  html+='<tr '+cl+'><td>'+(r.timestamp_utc||'-')+'</td><td><b>'+r.dcp_id+'</b></td><td>'+(r.nombre||'-')+'</td><td>'+(r.timestamp_msg||'-')+'</td><td>'+(r.servidor||'-')+'</td><td>'+ic+'</td></tr>'}
 document.getElementById('bitacoraBody').innerHTML=html;document.getElementById('bitacoraCount').textContent=data.length+' registros'})}
function esc(t){const d=document.createElement('div');d.textContent=t;return d.innerHTML}
function loadLog(){fetch('/api/log?limit=300').then(r=>r.json()).then(data=>{
 let h='';for(const e of data){let c='';const m=e.msg;
  if(m.includes('[OK]')||m.includes('RETRY OK'))c='log-ok';else if(m.includes('[FAIL]')||m.includes('[ERROR]'))c='log-fail';
  else if(m.includes('[WARN]')||m.includes('AGOTADO'))c='log-warn';else if(m.includes('[RETRY'))c='log-retry';
  else if(m.includes('[INFO]')||m.includes('[START]')||m.includes('[SCHED'))c='log-info';
  h+='<div class="log-line"><span class="log-ts">'+e.ts+'</span><span class="'+c+'">'+esc(m)+'</span></div>'}
 const ct=document.getElementById('logContainer');ct.innerHTML=h;if(document.getElementById('autoScroll').checked)ct.scrollTop=ct.scrollHeight})}
function loadConfig(){fetch('/api/config').then(r=>r.json()).then(data=>{
 let h='';for(const[s,vals]of Object.entries(data)){h+='<h4 style="color:#58a6ff;margin-top:16px;display:flex;align-items:center;gap:6px"><span style="color:#484f58">[</span>'+s+'<span style="color:#484f58">]</span></h4><div class="ui inverted form">';
  for(const[k,v]of Object.entries(vals)){const ip=k.toLowerCase().includes('password');
   h+='<div class="inline field" style="margin-bottom:4px"><label style="width:260px;color:#8b949e">'+k+'</label><input type="'+(ip?'password':'text')+'" value="'+esc(v)+'" style="background:#0d1117;border:1px solid #30363d;color:#e0e0e0;padding:6px 10px;width:300px;border-radius:4px" '+(ip?'disabled':'onchange="updateConfig(\''+s+"','"+k+"',this.value)\"")+"></div>"}
  h+='</div>'}document.getElementById('configContent').innerHTML=h})}
function updateConfig(s,k,v){fetch('/api/config',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({section:s,key:k,value:v})}).then(r=>r.json()).then(d=>{alert(d.success?'Actualizado: ['+s+'] '+k:'Error: '+(d.error||'?'))})}
function doRecover(){const id=document.getElementById('recoverDcpId').value.trim(),h=parseInt(document.getElementById('recoverHoras').value),b={horas:h};if(id)b.dcp_id=id;
 document.getElementById('recoverResult').innerHTML='<div class="ui blue message fade-in"><div class="ui tiny active inline loader"></div> Iniciando recuperacion...</div>';
 fetch('/api/recover',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(b)}).then(r=>r.json()).then(d=>{
  document.getElementById('recoverResult').innerHTML=d.success?'<div class="ui green message fade-in">&#10004; '+d.message+'</div>':'<div class="ui red message fade-in">&#10008; '+d.error+'</div>'})}
function doFetch(){const id=document.getElementById('fetchDcpId').value.trim(),h=parseInt(document.getElementById('fetchHoras').value);if(!id){alert('Ingresa DCP ID');return}
 document.getElementById('fetchResult').innerHTML='<div class="ui blue message fade-in"><div class="ui tiny active inline loader"></div> Descargando...</div>';
 fetch('/api/fetch/'+id+'?horas='+h,{method:'POST'}).then(r=>r.json()).then(d=>{
  document.getElementById('fetchResult').innerHTML=d.success?'<div class="ui blue message fade-in">&#10004; '+d.message+'</div>':'<div class="ui red message fade-in">&#10008; '+d.error+'</div>'})}
function toggleScheduler(){fetch('/api/scheduler',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({action:_isRunning?'pause':'resume'})}).then(r=>r.json()).then(()=>loadStatus())}
function refreshAll(){loadStatus();const a=document.querySelector('.tabular.menu .item.active');if(a){const t=a.getAttribute('data-tab');if(t==='stations')loadStations();else if(t==='retries')loadRetries()}}
loadStatus();loadStations();loadRetries();loadBitacora();loadLog();loadConfig();
setInterval(refreshAll,10000);
setInterval(()=>{if(document.getElementById('autoRefreshLog')&&document.getElementById('autoRefreshLog').checked){const a=document.querySelector('.tabular.menu .item.active');if(a&&a.getAttribute('data-tab')==='log')loadLog()}},5000);
setInterval(loadBitacora,30000);
// === MIS FUNCTIONS ===
function loadMisStatus(){fetch('/api/mis/status').then(r=>r.json()).then(d=>{
 document.getElementById('misOutDir').textContent=d.output_dir||'-';
 document.getElementById('misOutCount').textContent=d.output_count;
 document.getElementById('misImportDir').textContent=d.import_dir||'-';
 document.getElementById('misImportInterval').textContent=d.import_interval_min+' min';
 document.getElementById('misProcCount').textContent=d.procesados_count;
 document.getElementById('misErrCount').textContent=d.errores_count;
 const tgl=document.getElementById('misToggle'),lbl=document.getElementById('misToggleLabel');
 tgl.checked=d.generar_mis;lbl.textContent=d.generar_mis?'Activa':'Desactivada';lbl.style.color=d.generar_mis?'#3fb950':'#f85149';
 let h='';if(!d.pendientes.length){h='<p style="color:#484f58;text-align:center;padding:20px">&#10004; Sin archivos pendientes</p>'}
 else{for(const f of d.pendientes){h+='<div style="display:flex;justify-content:space-between;align-items:center;padding:4px 6px;border-bottom:1px solid #161b22"><span style="color:#e0e0e0;font-size:.8em">'+esc(f.nombre)+'</span><span style="display:flex;align-items:center;gap:6px"><span style="color:#484f58;font-size:.75em">'+f.size_kb+' KB</span><button class="ui tiny compact teal button" style="border-radius:4px;padding:4px 8px;font-size:.7em" onclick="importOneMis(\''+esc(f.nombre)+'\')">Importar</button></span></div>'}}
 document.getElementById('misPendingList').innerHTML=h})}
function toggleMisGen(enable){fetch('/api/mis/toggle',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({enable:enable})}).then(r=>r.json()).then(d=>{
 if(d.success){const lbl=document.getElementById('misToggleLabel');lbl.textContent=d.generar_mis?'Activa':'Desactivada';lbl.style.color=d.generar_mis?'#3fb950':'#f85149'}})}
function importOneMis(nombre){document.getElementById('misImportResult').innerHTML='<div class="ui blue message fade-in"><div class="ui tiny active inline loader"></div> Importando '+esc(nombre)+'...</div>';
 fetch('/api/mis/import',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({archivo:nombre})}).then(r=>r.json()).then(d=>{
  document.getElementById('misImportResult').innerHTML=d.success?'<div class="ui green message fade-in">&#10004; '+d.message+'</div>':'<div class="ui red message fade-in">&#10008; '+d.error+'</div>';
  setTimeout(loadMisStatus,2000)})}
function importAllMis(){document.getElementById('misImportResult').innerHTML='<div class="ui blue message fade-in"><div class="ui tiny active inline loader"></div> Importando todos los archivos...</div>';
 fetch('/api/mis/import',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({})}).then(r=>r.json()).then(d=>{
  document.getElementById('misImportResult').innerHTML=d.success?'<div class="ui green message fade-in">&#10004; '+d.message+'</div>':'<div class="ui red message fade-in">&#10008; '+d.error+'</div>';
  setTimeout(loadMisStatus,3000)})}
loadMisStatus();setInterval(loadMisStatus,30000);
</script>
</body>
</html>'''

_monitor_app = Flask(__name__, static_folder=os.path.join(_base_dir, 'static'))
_monitor_app.config['JSON_AS_ASCII'] = False
_scheduler_running = threading.Event()
_scheduler_running.set()  # Arranca activo

@_monitor_app.route('/')
def monitor_index():
    # Primero intentar archivo externo (para desarrollo), luego HTML embebido
    html_path = os.path.join(_base_dir, 'monitor.html')
    if os.path.exists(html_path):
        with open(html_path, 'r', encoding='utf-8') as f:
            return Response(f.read(), mimetype='text/html')
    return Response(_MONITOR_HTML, mimetype='text/html')

@_monitor_app.route('/api/status')
def api_status():
    uptime = datetime.datetime.now() - _start_time
    dias = uptime.days
    horas, rem = divmod(uptime.seconds, 3600)
    minutos = rem // 60
    return jsonify({
        "running": _scheduler_running.is_set(),
        "uptime": f"{dias}d {horas}h {minutos}m",
        "start_time": _start_time.strftime("%Y-%m-%d %H:%M:%S"),
        "total_estaciones": len(ESTACIONES),
        "total_scheduled": len(_scheduled_dcp_ids),
        "retries_activos": len(_retry_state),
        "pending_jobs": len(schedule.get_jobs()),
        "hosts_lrgs": HOSTS,
        "auto_recover_hours": config.getint('settings', 'auto_recover_hours', fallback=12),
        "recover_interval_min": config.getint('settings', 'auto_recover_interval_minutes', fallback=30),
        "reload_interval_min": config.getint('settings', 'reload_interval_minutes', fallback=5),
        "retry_max": MAX_RETRIES,
        "retry_delays": RETRY_DELAYS,
        "fetch_ok": _fetch_ok_count,
        "fetch_fail": _fetch_fail_count,
    })

@_monitor_app.route('/api/stations')
def api_stations():
    result = []
    for dcp_id, data in ESTACIONES.items():
        retry = _retry_state.get(dcp_id)
        result.append({
            "dcp_id": dcp_id,
            "nombre": data.get("nombre", "?"),
            "id_asignado": data.get("id_asignado", ""),
            "minuto_tx": data.get("minuto", -1),
            "rango_tx": data.get("rango_transmision", 60),
            "cuenca": data.get("cuenca", ""),
            "scheduled": dcp_id in _scheduled_dcp_ids,
            "retry": {
                "intentos": retry["intentos"],
                "hora_fallida": retry["hora_fallida"].strftime("%H:%M:%S"),
            } if retry else None,
        })
    result.sort(key=lambda x: x["minuto_tx"])
    return jsonify(result)

@_monitor_app.route('/api/retries')
def api_retries():
    result = []
    for dcp_id, state in _retry_state.items():
        data = ESTACIONES.get(dcp_id, {})
        result.append({
            "dcp_id": dcp_id,
            "nombre": data.get("nombre", "?"),
            "intentos": state["intentos"],
            "max_intentos": MAX_RETRIES,
            "hora_fallida": state["hora_fallida"].strftime("%H:%M:%S"),
        })
    return jsonify(result)

@_monitor_app.route('/api/log')
def api_log():
    limit = request.args.get('limit', 200, type=int)
    with _log_lock:
        entries = list(_log_buffer)[-limit:]
    return jsonify(entries)

@_monitor_app.route('/api/bitacora')
def api_bitacora():
    limit = request.args.get('limit', 100, type=int)
    horas = request.args.get('horas', 6, type=int)
    conn = get_pg_conn()
    if not conn:
        return jsonify({"error": "Sin conexión PG"}), 500
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT dcp_id, timestamp_utc, timestamp_msg, servidor, exito
            FROM bitacora_goes
            WHERE timestamp_utc > NOW() - make_interval(hours := %s)
            ORDER BY timestamp_utc DESC
            LIMIT %s
        """, (horas, limit))
        rows = cur.fetchall()
        cur.close()
        result = [{
            "dcp_id": r[0],
            "nombre": ESTACIONES.get(r[0], {}).get("nombre", "-"),
            "timestamp_utc": r[1].strftime("%Y-%m-%d %H:%M:%S") if r[1] else None,
            "timestamp_msg": r[2].strftime("%Y-%m-%d %H:%M:%S") if r[2] else None,
            "servidor": r[3],
            "exito": r[4],
        } for r in rows]
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        release_pg_conn(conn)

@_monitor_app.route('/api/config', methods=['GET'])
def api_config_get():
    cfg = {}
    for section in config.sections():
        cfg[section] = dict(config.items(section))
        # Ocultar contraseñas
        for key in cfg[section]:
            if 'password' in key.lower():
                cfg[section][key] = '••••••'
    return jsonify(cfg)

@_monitor_app.route('/api/config', methods=['POST'])
def api_config_set():
    data = request.get_json()
    if not data or 'section' not in data or 'key' not in data or 'value' not in data:
        return jsonify({"error": "Faltan campos: section, key, value"}), 400
    section = data['section']
    key = data['key']
    value = data['value']
    if 'password' in key.lower():
        return jsonify({"error": "No se pueden cambiar contraseñas desde la web"}), 403
    if not config.has_section(section):
        return jsonify({"error": f"Sección '{section}' no existe"}), 404
    config.set(section, key, str(value))
    config_path = os.path.join(_base_dir, 'config.ini')
    with open(config_path, 'w') as f:
        config.write(f)
    print(f"[MONITOR] Config actualizada: [{section}] {key} = {value}")
    return jsonify({"success": True, "message": f"[{section}] {key} = {value}"})

@_monitor_app.route('/api/fetch/<dcp_id>', methods=['POST'])
def api_fetch_station(dcp_id):
    dcp_id = dcp_id.upper()
    if dcp_id not in ESTACIONES:
        return jsonify({"error": f"Estación {dcp_id} no encontrada"}), 404
    horas = request.args.get('horas', 2, type=int)
    def _do():
        fetch_messages_for_dcp(dcp_id, since=f"now - {horas} hours", multi=horas > 2)
    t = threading.Thread(target=_do, daemon=True)
    t.start()
    return jsonify({"success": True, "message": f"Fetch iniciado: {dcp_id} ({horas}h)"})

@_monitor_app.route('/api/recover', methods=['POST'])
def api_recover():
    data = request.get_json() or {}
    horas = data.get('horas', 12)
    dcp_id = data.get('dcp_id', '').upper()
    if dcp_id:
        if dcp_id not in ESTACIONES:
            return jsonify({"error": f"Estación {dcp_id} no encontrada"}), 404
        def _do():
            fetch_messages_for_dcp(dcp_id, since=f"now - {horas} hours", multi=True)
        t = threading.Thread(target=_do, daemon=True)
        t.start()
        return jsonify({"success": True, "message": f"Recovery {dcp_id}: {horas}h"})
    else:
        def _do():
            recuperacion_automatica()
        t = threading.Thread(target=_do, daemon=True)
        t.start()
        return jsonify({"success": True, "message": f"Recovery global iniciada ({horas}h)"})

@_monitor_app.route('/api/scheduler', methods=['POST'])
def api_scheduler_control():
    data = request.get_json() or {}
    action = data.get('action', '')
    if action == 'pause':
        _scheduler_running.clear()
        print("[MONITOR] Scheduler PAUSADO por el usuario")
        return jsonify({"success": True, "running": False})
    elif action == 'resume':
        _scheduler_running.set()
        print("[MONITOR] Scheduler REANUDADO por el usuario")
        return jsonify({"success": True, "running": True})
    return jsonify({"error": "Acción no válida. Usa 'pause' o 'resume'"}), 400

@_monitor_app.route('/api/mis/status')
def api_mis_status():
    """Estado de la configuración .mis: directorios, archivos pendientes, etc."""
    import glob as glob_mod
    generar = config.getboolean('settings', 'generar_mis', fallback=True)
    import_dir = config.get('paths',
        'mis_import_dir_windows' if platform.system() == 'Windows' else 'mis_import_dir_mac',
        fallback='')
    output_dir = MIS_OUT_DIR
    import_interval = config.getint('settings', 'auto_import_interval_minutes', fallback=5)

    # Contar archivos en cada carpeta
    pendientes = []
    procesados_count = 0
    errores_count = 0
    output_count = 0

    if import_dir and os.path.isdir(import_dir):
        for f in sorted(glob_mod.glob(os.path.join(import_dir, "*.mis"))):
            sz = os.path.getsize(f)
            pendientes.append({"nombre": os.path.basename(f), "size_kb": round(sz / 1024, 1)})
        proc_dir = os.path.join(import_dir, "procesados")
        err_dir = os.path.join(import_dir, "errores")
        if os.path.isdir(proc_dir):
            procesados_count = len(glob_mod.glob(os.path.join(proc_dir, "*.mis")))
        if os.path.isdir(err_dir):
            errores_count = len(glob_mod.glob(os.path.join(err_dir, "*.mis")))

    if os.path.isdir(output_dir):
        output_count = len(glob_mod.glob(os.path.join(output_dir, "*.mis")))

    return jsonify({
        "generar_mis": generar,
        "output_dir": output_dir,
        "output_count": output_count,
        "import_dir": import_dir,
        "import_interval_min": import_interval,
        "pendientes": pendientes,
        "pendientes_count": len(pendientes),
        "procesados_count": procesados_count,
        "errores_count": errores_count,
    })

@_monitor_app.route('/api/mis/toggle', methods=['POST'])
def api_mis_toggle():
    """Activa/desactiva la generación de archivos .mis en tiempo real."""
    data = request.get_json() or {}
    enable = data.get('enable')
    if enable is None:
        return jsonify({"error": "Falta parámetro 'enable' (true/false)"}), 400
    val = 'true' if enable else 'false'
    config.set('settings', 'generar_mis', val)
    config_path = os.path.join(_base_dir, 'config.ini')
    with open(config_path, 'w') as f:
        config.write(f)
    print(f"[MONITOR] Generación .mis {'ACTIVADA' if enable else 'DESACTIVADA'}")
    return jsonify({"success": True, "generar_mis": enable})

@_monitor_app.route('/api/mis/import', methods=['POST'])
def api_mis_import():
    """Ejecuta importación manual de archivos .mis pendientes."""
    data = request.get_json() or {}
    archivo = data.get('archivo')  # nombre específico o None para todos
    import_dir = config.get('paths',
        'mis_import_dir_windows' if platform.system() == 'Windows' else 'mis_import_dir_mac',
        fallback='')
    if not import_dir or not os.path.isdir(import_dir):
        return jsonify({"error": f"Directorio de importación no existe: {import_dir}"}), 400

    def _run_import():
        if archivo:
            ruta = os.path.join(import_dir, archivo)
            if os.path.isfile(ruta):
                print(f"[MONITOR IMPORT] Importando: {archivo}")
                n = importar_archivo_mis(ruta, SENSORES, ESTACIONES)
                print(f"[MONITOR IMPORT] {archivo}: {n} registros")
                # Mover a procesados/errores
                dest_dir = os.path.join(import_dir, "procesados" if n > 0 else "errores")
                os.makedirs(dest_dir, exist_ok=True)
                import shutil
                shutil.move(ruta, os.path.join(dest_dir, archivo))
            else:
                print(f"[MONITOR IMPORT] No existe: {archivo}")
        else:
            print("[MONITOR IMPORT] Importando todos los pendientes...")
            importar_mis_automatico()

    t = threading.Thread(target=_run_import, daemon=True)
    t.start()
    msg = f"Importando {archivo}" if archivo else "Importando todos los archivos pendientes"
    return jsonify({"success": True, "message": msg})

def _iniciar_monitor_web(port=5555):
    """Inicia el servidor Flask en un hilo daemon."""
    import logging
    log = logging.getLogger('werkzeug')
    log.setLevel(logging.WARNING)
    print(f"[MONITOR] Panel web disponible en http://localhost:{port}")
    _monitor_app.run(host='0.0.0.0', port=port, debug=False, use_reloader=False)

# ==================== SCHEDULER ====================
def main():
    global _scheduled_dcp_ids
    reload_minutes = config.getint('settings', 'reload_interval_minutes', fallback=5)
    recover_interval = config.getint('settings', 'auto_recover_interval_minutes', fallback=30)

    print("[INFO] Programando descargas (+1 minuto)...")
    for dcp_id, data in ESTACIONES.items():
        _programar_estacion(dcp_id, data)
        _scheduled_dcp_ids.add(dcp_id)

    # Programar recarga periódica de configuración
    schedule.every(reload_minutes).minutes.do(recargar_configuracion)
    print(f"[INFO] Recarga automática de configuración cada {reload_minutes} minutos.")

    # Programar recuperación automática de huecos
    schedule.every(recover_interval).minutes.do(recuperacion_automatica)
    print(f"[INFO] Recuperación automática de huecos cada {recover_interval} minutos.")

    # Programar importación automática de archivos .mis
    import_interval = config.getint('settings', 'auto_import_interval_minutes', fallback=5)
    mis_import_dir = config.get('paths', 'mis_import_dir_windows' if platform.system() == 'Windows' else 'mis_import_dir_mac', fallback='')
    if mis_import_dir:
        os.makedirs(mis_import_dir, exist_ok=True)
        schedule.every(import_interval).minutes.do(importar_mis_automatico)
        print(f"[INFO] Importación automática de .mis cada {import_interval} min desde: {mis_import_dir}")

    # Programar importación automática de Excel FunVasos
    fv_interval = config.getint('settings', 'funvasos_import_interval_minutes', fallback=5)
    fv_inbox_key = 'funvasos_inbox_windows' if platform.system() == 'Windows' else 'funvasos_inbox_mac'
    fv_inbox_dir = config.get('paths', fv_inbox_key, fallback='')
    if fv_inbox_dir:
        os.makedirs(fv_inbox_dir, exist_ok=True)
        schedule.every(fv_interval).minutes.do(importar_funvasos_automatico)
        print(f"[INFO] Importación automática de FunVasos cada {fv_interval} min desde: {fv_inbox_dir}")

    # Ejecutar una primera detección al arrancar (después de 2 minutos para dar tiempo al primer ciclo)
    schedule.every(2).minutes.do(lambda: (recuperacion_automatica(), schedule.CancelJob)[-1]).tag("recover_inicial")

    # Descargar assets estáticos si no existen
    _ensure_static_assets()

    # Iniciar panel de monitoreo web en hilo daemon
    monitor_port = config.getint('settings', 'monitor_port', fallback=5555)
    monitor_thread = threading.Thread(target=_iniciar_monitor_web, args=(monitor_port,), daemon=True)
    monitor_thread.start()

    print("[INFO] Scheduler activo...")
    while True:
        if _scheduler_running.is_set():
            schedule.run_pending()
        time.sleep(30)

# ==================== CARGA INICIAL ====================
_ensure_pg_tables()
ESTACIONES = load_estaciones_from_sql()
SENSORES = load_sensores_from_sql()
UMBRALES_PRECIP = load_umbrales_precipitacion()
CUENCAS_ESTACION = load_cuencas_from_sql()

# Asignar posiciones considerando el rango de transmisión de cada estación
for id_asignado, sensores in SENSORES.items():
    # Buscar el rango de transmisión de esta estación
    rango_tx = 60  # Default
    for dcp_id, datos_estacion in ESTACIONES.items():
        if datos_estacion.get("id_asignado") == id_asignado:
            rango_tx = datos_estacion.get("rango_transmision", 60)
            break
    
    SENSORES[id_asignado] = asignar_posiciones_sensores(sensores, rango_tx)

# Re-validar datos que pudieron marcarse como inválidos con límites anteriores
revalidar_datos_invalidos(SENSORES, ESTACIONES)

# ==================== SERVICIO WINDOWS ====================
_SERVICE_NAME = "MyCloudTimescale"
_SERVICE_DISPLAY = "MyCloud GOES TimescaleDB"
_SERVICE_DESC = "Servicio de ingesta de datos GOES satelitales hacia TimescaleDB"

if platform.system() == 'Windows':
    try:
        import win32serviceutil
        import win32service
        import win32event
        import servicemanager

        class MyCloudService(win32serviceutil.ServiceFramework):
            _svc_name_ = _SERVICE_NAME
            _svc_display_name_ = _SERVICE_DISPLAY
            _svc_description_ = _SERVICE_DESC

            def __init__(self, args):
                win32serviceutil.ServiceFramework.__init__(self, args)
                self.hWaitStop = win32event.CreateEvent(None, 0, 0, None)
                self._running = True

            def SvcStop(self):
                self.ReportServiceStatus(win32service.SERVICE_STOP_PENDING)
                self._running = False
                _scheduler_running.clear()
                win32event.SetEvent(self.hWaitStop)
                print("[SERVICE] Deteniendo servicio...")

            def SvcDoRun(self):
                servicemanager.LogMsg(
                    servicemanager.EVENTLOG_INFORMATION_TYPE,
                    servicemanager.PYS_SERVICE_STARTED,
                    (self._svc_name_, ''))
                print(f"[SERVICE] {_SERVICE_DISPLAY} iniciado.")
                try:
                    self._run_main()
                except Exception as ex:
                    print(f"[SERVICE] Error fatal: {ex}")
                    servicemanager.LogErrorMsg(f"{_SERVICE_NAME}: {ex}")

            def _run_main(self):
                global _scheduled_dcp_ids
                reload_minutes = config.getint('settings', 'reload_interval_minutes', fallback=5)
                recover_interval = config.getint('settings', 'auto_recover_interval_minutes', fallback=30)

                for dcp_id, data in ESTACIONES.items():
                    _programar_estacion(dcp_id, data)
                    _scheduled_dcp_ids.add(dcp_id)

                schedule.every(reload_minutes).minutes.do(recargar_configuracion)
                schedule.every(recover_interval).minutes.do(recuperacion_automatica)

                import_interval = config.getint('settings', 'auto_import_interval_minutes', fallback=5)
                mis_import_dir = config.get('paths', 'mis_import_dir_windows', fallback='')
                if mis_import_dir:
                    os.makedirs(mis_import_dir, exist_ok=True)
                    schedule.every(import_interval).minutes.do(importar_mis_automatico)

                fv_interval = config.getint('settings', 'funvasos_import_interval_minutes', fallback=5)
                fv_inbox_dir = config.get('paths', 'funvasos_inbox_windows', fallback='')
                if fv_inbox_dir:
                    os.makedirs(fv_inbox_dir, exist_ok=True)
                    schedule.every(fv_interval).minutes.do(importar_funvasos_automatico)

                schedule.every(2).minutes.do(lambda: (recuperacion_automatica(), schedule.CancelJob)[-1]).tag("recover_inicial")

                _ensure_static_assets()
                monitor_port = config.getint('settings', 'monitor_port', fallback=5555)
                monitor_thread = threading.Thread(target=_iniciar_monitor_web, args=(monitor_port,), daemon=True)
                monitor_thread.start()

                while self._running:
                    if _scheduler_running.is_set():
                        schedule.run_pending()
                    # Revisar stop cada 5 segundos para respuesta rápida
                    rc = win32event.WaitForSingleObject(self.hWaitStop, 5000)
                    if rc == win32event.WAIT_OBJECT_0:
                        break
                print(f"[SERVICE] {_SERVICE_DISPLAY} detenido.")

        _HAS_WIN32 = True
    except ImportError:
        _HAS_WIN32 = False
else:
    _HAS_WIN32 = False

# ==================== ENTRADA ====================
if __name__ == "__main__":
    import sys

    if "--service" in sys.argv and _HAS_WIN32:
        # Ejecutado por el SCM de Windows (no usar directamente)
        servicemanager.Initialize()
        servicemanager.PrepareToHostSingle(MyCloudService)
        servicemanager.StartServiceCtrlDispatcher()
    elif any(cmd in sys.argv for cmd in ("install", "remove", "start", "stop", "update")):
        # Gestión del servicio: install, remove, start, stop, update
        if not _HAS_WIN32:
            print("[ERROR] pywin32 no disponible. Instala con: pip install pywin32")
            sys.exit(1)
        # Inyectar --startup auto en install para que arranque con Windows
        if "install" in sys.argv and "--startup" not in sys.argv:
            idx = sys.argv.index("install")
            sys.argv.insert(idx + 1, "--startup")
            sys.argv.insert(idx + 2, "auto")
        win32serviceutil.HandleCommandLine(MyCloudService)
    elif "--gaps" in sys.argv:
        # Modo detección de huecos: --gaps [horas] [--fix]
        # Ejemplos:
        #   python mycloud_all_timescale.py --gaps 6          (detectar huecos últimas 6 horas)
        #   python mycloud_all_timescale.py --gaps 12 --fix   (detectar y recuperar automáticamente)
        args = [a for a in sys.argv[1:] if a not in ("--gaps", "--fix")]
        horas = 6  # default
        auto_fix = "--fix" in sys.argv
        for a in args:
            try:
                horas = int(a)
            except ValueError:
                pass

        print(f"[GAPS] Analizando huecos en las últimas {horas} horas...")
        estaciones_con_huecos = detectar_huecos(horas)

        # Mostrar resultados
        if not estaciones_con_huecos:
            print(f"[GAPS] Sin huecos detectados en las últimas {horas} horas.")
        else:
            print(f"\n{'='*80}")
            print(f"  HUECOS DETECTADOS: {len(estaciones_con_huecos)} estaciones con datos faltantes")
            print(f"{'='*80}")
            for est in estaciones_con_huecos:
                print(f"\n  {est['dcp_id']} | {est['nombre']} (TX cada {est['rango_tx']} min)")
                for h in est["huecos"]:
                    print(f"    [{h['tipo']}] {h['desde']} → {h['hasta']} ({h['duracion_h']}h)")

            if auto_fix:
                print(f"\n[GAPS] Recuperando datos faltantes...")
                for est in estaciones_con_huecos:
                    max_h = max(h["duracion_h"] for h in est["huecos"])
                    recover_hours = min(int(max_h + 2), horas)
                    since = f"now - {recover_hours} hours"
                    print(f"  [FIX] {est['dcp_id']} | {est['nombre']} → recover {recover_hours}h")
                    fetch_messages_for_dcp(est["dcp_id"], since=since, multi=True)
                print("[GAPS] Recuperación finalizada.")
            else:
                print(f"\n  Para recuperar automáticamente, ejecuta:")
                print(f"    python mycloud_all_timescale.py --gaps {horas} --fix")
                print(f"\n  O recuperar estación específica:")
                print(f"    python mycloud_all_timescale.py --recover {horas} <DCP_ID>")

    elif "--recover" in sys.argv:
        # Modo recuperación: --recover [horas] [dcp_id]
        # Ejemplos:
        #   python mycloud_all_timescale.py --recover 5          (todas las estaciones, últimas 5 horas)
        #   python mycloud_all_timescale.py --recover 3 E8906DD6 (solo una estación, últimas 3 horas)
        args = [a for a in sys.argv[1:] if a != "--recover"]
        horas = 4  # default
        target_dcp = None
        for a in args:
            try:
                horas = int(a)
            except ValueError:
                target_dcp = a.upper()
        since = f"now - {horas} hours"
        print(f"[RECOVER] Recuperando últimas {horas} horas (since='{since}')...")
        if target_dcp:
            print(f"[RECOVER] Solo estación: {target_dcp}")
            fetch_messages_for_dcp(target_dcp, since=since, multi=True)
        else:
            total = len(ESTACIONES)
            for i, dcp_id in enumerate(ESTACIONES, 1):
                print(f"[RECOVER] [{i}/{total}] {dcp_id}")
                fetch_messages_for_dcp(dcp_id, since=since, multi=True)
        print("[RECOVER] Finalizado.")
    elif "--manual" in sys.argv:
        print("[MANUAL] Ejecutando todas las estaciones...")
        for dcp_id in ESTACIONES:
            fetch_messages_for_dcp(dcp_id)
        print("[MANUAL] Finalizado.")
    elif "--import" in sys.argv:
        # Modo importación de archivos .mis históricos
        # Ejemplos:
        #   python mycloud_all_timescale.py --import /ruta/a/directorio
        #   python mycloud_all_timescale.py --import /ruta/a/archivo.mis
        #   python mycloud_all_timescale.py --import /ruta/a/directorio "E89*.mis"
        args = [a for a in sys.argv[1:] if a != "--import"]
        if not args:
            print("[IMPORT] Uso:")
            print("  python mycloud_all_timescale.py --import <directorio>")
            print("  python mycloud_all_timescale.py --import <archivo.mis>")
            print("  python mycloud_all_timescale.py --import <directorio> \"patron*.mis\"")
            sys.exit(1)

        ruta = args[0]
        patron = args[1] if len(args) > 1 else "*.mis"

        if os.path.isfile(ruta):
            print(f"[IMPORT] Importando archivo: {ruta}")
            n = importar_archivo_mis(ruta, SENSORES, ESTACIONES)
            print(f"[IMPORT] {n} registros importados.")
        elif os.path.isdir(ruta):
            importar_directorio_mis(ruta, SENSORES, ESTACIONES, patron)
        else:
            print(f"[IMPORT] Ruta no encontrada: {ruta}")
            sys.exit(1)
    elif "--help" in sys.argv or "-h" in sys.argv:
        print(f"""
MyCloud GOES TimescaleDB - Ingesta de datos satelitales
========================================================
Uso:
  python mycloud_all_timescale.py                     Modo consola (scheduler)
  python mycloud_all_timescale.py --gaps 6            Detectar huecos (últimas 6h)
  python mycloud_all_timescale.py --gaps 12 --fix     Detectar y recuperar huecos
  python mycloud_all_timescale.py --recover 5         Recuperar últimas 5h (todas)
  python mycloud_all_timescale.py --recover 3 DCP_ID  Recuperar estación específica
  python mycloud_all_timescale.py --manual             Fetch todas las estaciones
  python mycloud_all_timescale.py --import <ruta>      Importar archivos .mis

Servicio Windows (requiere pywin32):
  python mycloud_all_timescale.py install              Instalar servicio (auto-start)
  python mycloud_all_timescale.py start                Iniciar servicio
  python mycloud_all_timescale.py stop                 Detener servicio  
  python mycloud_all_timescale.py remove               Desinstalar servicio
  python mycloud_all_timescale.py update               Actualizar configuración del servicio

Monitor web: http://localhost:{config.getint('settings', 'monitor_port', fallback=5555)}
""")
    else:
        main()

    # for dcp_id in ESTACIONES:
    #         fetch_messages_for_dcp(dcp_id)
    # print("[MANUAL] Finalizado.")