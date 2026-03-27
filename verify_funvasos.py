import openpyxl
import psycopg2

# Quick test: verify the funvasos Excel parsing logic matches what C# will do
wb = openpyxl.load_workbook('/Users/subgerenciagrijalva/Downloads/FIN240326.xlsx', data_only=True)
ws = wb['FIN']

# Check date extraction: row 4, col H
date_cell_r4h = ws.cell(row=4, column=8).value
date_cell_r3h = ws.cell(row=3, column=8).value
print(f"R4.H = {date_cell_r4h} (type: {type(date_cell_r4h).__name__})")
print(f"R3.H = {date_cell_r3h} (type: {type(date_cell_r3h).__name__})")

# Check Angostura section (rows 15-38, data in cols A-P)
sections = {
    'Angostura': (15, 38),
    'Chicoasén': (48, 70),
    'Malpaso': (81, 103),
    'Tapón Juan Grijalva': (114, 136),
    'Peñitas': (147, 170),
}

for presa, (start, end) in sections.items():
    count = 0
    for row in range(start, end + 1):
        hora = ws.cell(row=row, column=1).value  # A
        elev = ws.cell(row=row, column=2).value   # B
        if hora is not None and elev is not None:
            count += 1
    print(f"{presa}: {count} horas con datos")
    # Show first row
    row = start
    print(f"  Row {row}: Hora={ws.cell(row=row,column=1).value}, Elev={ws.cell(row=row,column=2).value}, Almac={ws.cell(row=row,column=3).value}")

# Check TimescaleDB table exists
conn = psycopg2.connect(host='atlas16.ddns.net', dbname='mycloud_timescale', user='postgres', password='***REDACTED-PG-PASSWORD***')
cur = conn.cursor()
cur.execute("SELECT count(*) FROM public.funvasos_horario")
print(f"\nFunvasos rows in DB: {cur.fetchone()[0]}")
cur.close()
conn.close()

wb.close()
print("\nVerification complete!")
