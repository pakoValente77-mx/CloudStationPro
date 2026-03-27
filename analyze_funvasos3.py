import openpyxl
wb = openpyxl.load_workbook('DocumentRepository/funvasos/FIN160226.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Check rows 6-18 for column headers
print('=== ROWS 6-18 (all non-empty) ===')
for r in range(6, 19):
    for c in range(1, 30):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            print(f'  R{r} {cl}(c{c}): {repr(v)}')

# Get the header row for dates
print('\n=== ROW 18 date headers Z onwards ===')
for c in range(26, 60):
    v = ws.cell(row=18, column=c).value
    if v is not None:
        from openpyxl.utils import get_column_letter
        cl = get_column_letter(c)
        print(f'  {cl}18: {repr(v)}')

# Check far-right summary columns OA-OI  
# OA=col 391, OB=392, etc
print('\n=== Columns OA-OI (summary) header rows 11-18 ===')
for col_name, col_idx in [('OA',391),('OB',392),('OC',393),('OD',394),('OE',395),('OF',396),('OG',397),('OH',398),('OI',399)]:
    for r in range(11, 19):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None:
            print(f'  R{r} {col_name}(c{col_idx}): {repr(v)}')

# Check actual OA-OI data for first dam rows 19-22
print('\n=== Columns OA-OI data rows 19-22 ===')
for col_name, col_idx in [('OA',391),('OB',392),('OC',393),('OD',394),('OE',395),('OF',396),('OG',397),('OH',398),('OI',399)]:
    for r in range(19, 23):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None:
            print(f'  R{r} {col_name}: {repr(v)}')

# Check columns after OK
print('\n=== Columns OJ-ON (far right) headers ===')
for col_idx in range(399, 422):
    for r in range(11, 19):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(col_idx)
            print(f'  R{r} {cl}(c{col_idx}): {repr(v)}')

# All unique B values (dam numbering)
print('\n=== All B values (dam numbers) ===')
for r in range(1, ws.max_row+1):
    v = ws.cell(row=r, column=2).value
    if v is not None:
        c = ws.cell(row=r, column=3).value  # name
        i = ws.cell(row=r, column=9).value  # code
        m = ws.cell(row=r, column=13).value # type
        x = ws.cell(row=r, column=24).value # variable
        print(f'  R{r}: B={v}, C={repr(c)}, I={repr(i)}, M={repr(m)}, X={repr(x)}')

# Check the structure of rows 297-318 (bottom of the sheet)
print('\n=== BOTTOM ROWS 295-318 ===')
for r in range(295, 319):
    for c in range(1, 30):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            print(f'  R{r} {cl}: {repr(v)}')
