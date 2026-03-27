import openpyxl
wb = openpyxl.load_workbook('DocumentRepository/funvasos/FIN160226.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]
print(f'Sheet: {wb.sheetnames[0]}, Rows: {ws.max_row}, Cols: {ws.max_column}')

# Column A values
print('\n=== COL A ===')
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v is not None:
        print(f'  R{r}: {v}')

# Headers - rows 1-5 all columns with values
print('\n=== ROWS 1-5 (all non-empty) ===')
for r in range(1, 6):
    for c in range(1, min(ws.max_column+1, 400)):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            print(f'  R{r} {cl}{r}(col{c}): {repr(v)}')

# Key data rows - first dam block
print('\n=== ROWS 18-55 KEY COLS (B,C-H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y) ===')
for r in range(18, 56):
    vals = []
    for c in range(1, 26):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            vals.append(f'{cl}={repr(v)}')
    if vals:
        print(f'  R{r}: {", ".join(vals)}')

# Check data area - Z column onwards for a few data rows
print('\n=== DATA CELLS Z-AE for rows 19-22 (first dam daily values) ===')
for r in range(19, 23):
    for c in range(26, 32):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            from openpyxl.utils import get_column_letter
            cl = get_column_letter(c)
            print(f'  R{r} {cl}: {repr(v)}')
